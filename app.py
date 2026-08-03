"""
MSP Property Dashboard — Streamlit App
Shared multi-user dashboard with Google Sheets backend.
"""
import streamlit as st
import base64
import gzip
import html as html_lib
import json
from io import BytesIO
import os
import re
import shutil
import subprocess
import tempfile
import time
from datetime import datetime, date, timedelta
from pathlib import Path
from uuid import uuid4

import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
import openpyxl
from openpyxl.styles import Alignment, Font as _XLFont
from openpyxl.worksheet.datavalidation import DataValidation
import json

try:
    import fitz
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False

try:
    from lease_builder import (
        BOOKMARK_LABELS,
        build_word_document,
        discover_templates,
        find_kp_references,
        humanize_bookmark,
        inspect_template,
        KP_LINE_BREAK,
        KP_LINK_COLOR,
        KP_TOKEN_RE,
        kp_value_plain,
        normalize_kp_tokens,
        normalize_kp_value,
        load_clause_library,
        publish_template_docx,
        PUBLISHED_PREFIX,
    )
    import lease_format as lf
    import lease_content as lc
    import lease_markup as lm
    import lease_render as lr
    import lease_docs as ld
    import lease_store as lstore
    import lease_space as lsp

    # Streamlit re-runs this script without re-importing a module it already
    # holds, so a deploy can leave a new app.py talking to an old lease_*.py.
    # The symptom is a redacted AttributeError deep inside a tab. Checking the
    # names up front turns that into an actionable message instead.
    for _module, _names in (
        (lf, ("normalize_profiles", "profile_settings", "resolve_profile_name",
              "migrate_template_formatting", "describe_settings", "normalize_settings",
              "settings_diff", "default_settings", "validate_settings",
              "DEFAULT_PROFILE_NAME", "DEFAULTS")),
        (lc, ("load_content", "extract_from_docx", "write_content")),
        (lm, ("parse_blocks", "to_html", "to_markup")),
        (lr, ("render_lease", "bookmark_names", "dangling_anchors")),
        (ld, ("normalize_store", "migrate_stores", "copy_document",
              "build_document", "describe_document", "apply_choice",
              "normalize_choice", "choice_options", "NO_CHOICE")),
        (lsp, ("space_records", "find_space", "resolve", "resolve_provisions",
               "token_names", "field_label", "unresolved", "SPACE_TOKEN_RE")),
        (lstore, ("build_store", "LeaseStore", "LocalBackend", "GitHubBackend",
                  "StoreError", "slugify")),
    ):
        _missing = [_name for _name in _names if not hasattr(_module, _name)]
        if _missing:
            raise ImportError(
                f"cannot import name {_missing[0]!r} from {_module.__name__!r} "
                f"(module is stale — {len(_missing)} name(s) missing)"
            )

    LEASE_BUILDER_AVAILABLE = True
    LEASE_BUILDER_ERROR = ""
except ImportError as exc:
    # Keep the real reason: this fires for a missing package and for a name that
    # lease_builder no longer exports, and those need different fixes.
    LEASE_BUILDER_AVAILABLE = False
    LEASE_BUILDER_ERROR = f"{type(exc).__name__}: {exc}"

from st_aggrid import AgGrid, DataReturnMode, GridOptionsBuilder, GridUpdateMode, JsCode

# --- CONFIG ---
SHEET_ID = "1Jqdnf9JFPLBoYirN7ZpBiicDo13meTcEdaI19usbZss"
SERVICE_ACCOUNT_FILE = None
SERVICE_ACCOUNT_INFO = None
COLUMN_CONFIG_FILE = os.path.join(os.path.dirname(__file__), "column_widths.json")
EXPENSE_CONFIG_FILE = os.path.join(os.path.dirname(__file__), "building_expenses.json")

# Try file first, then Streamlit secrets
for p in [
    os.path.join(os.path.dirname(__file__), "service_account.json"),
    "/home/node/.openclaw/workspace/config/gcp_service_account.json",
]:
    if os.path.exists(p):
        SERVICE_ACCOUNT_FILE = p
        break

if not SERVICE_ACCOUNT_FILE:
    try:
        SERVICE_ACCOUNT_INFO = dict(st.secrets["gcp_service_account"])
    except Exception:
        pass

TENANCY_FILE = None
for p in [
    os.path.join(os.path.dirname(__file__), "data", "MSP Tenancy.xlsx"),
    "/home/node/OpenClaw/Tenant Database/MSP Tenancy.xlsx",
    r"C:\Dropbox\OpenClaw\Tenant Database\MSP Tenancy.xlsx",
]:
    if os.path.exists(p):
        TENANCY_FILE = p
        break

SOP_PDF = None
for p in [
    os.path.join(os.path.dirname(__file__), "data", "Marion_St_SOP_Manual.pdf"),
    "/home/node/OpenClaw/Share Jason/General Procedures/Marion_St_SOP_Manual.pdf",
]:
    if os.path.exists(p):
        SOP_PDF = p
        break

BUILDING_MAP = {
    "114 Central": {"code": "MSP114", "dest_folder": "114 Central", "share": "0_114Share"},
    "15 South": {"code": "MSP15", "dest_folder": "15 South", "share": "0_15Share"},
    "36 South": {"code": "MSP36", "dest_folder": "36 South", "share": "0_36Share"},
    "1280 Springfield": {"code": "MSP1280", "dest_folder": "1280 Springfield", "share": "0_1280Share"},
}

ACTIVE_PROPS_ROOT = None
for _p in [
    "/home/node/OpenClaw/Share Jason/ACTIVE PROPERTIES",
    r"C:\Dropbox\ASRA Investments\Marion St Properties\ACTIVE PROPERTIES",
]:
    if os.path.exists(_p):
        ACTIVE_PROPS_ROOT = Path(_p)
        break

TODAY = date.today()

# --- PAGE CONFIG ---
st.set_page_config(
    page_title="MSP Property Dashboard",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# Custom CSS
st.markdown("""
<style>
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] { padding: 8px 20px; font-weight: 600; }
    .metric-card { background: #161b22; border: 1px solid #2d333b; border-radius: 8px; padding: 16px; text-align: center; }
    .metric-value { font-size: 28px; font-weight: 700; color: #58a6ff; font-family: monospace; }
    .metric-label { font-size: 12px; color: #8b949e; margin-top: 4px; }
    .building-header { background: #1a2332; padding: 8px 12px; border-radius: 6px; margin: 16px 0 8px; border-left: 3px solid #58a6ff; }
    .badge-mtm { background: #f8514926; color: #f85149; padding: 2px 8px; border-radius: 10px; font-size: 12px; font-weight: 600; }
    .badge-short { background: #d2992244; color: #d29922; padding: 2px 8px; border-radius: 10px; font-size: 12px; font-weight: 600; }
    .badge-ok { background: #23883826; color: #3fb950; padding: 2px 8px; border-radius: 10px; font-size: 12px; font-weight: 600; }
    .badge-long { background: #1f6feb26; color: #58a6ff; padding: 2px 8px; border-radius: 10px; font-size: 12px; font-weight: 600; }
    .sop-card { background: #161b22; border: 1px solid #2d333b; border-radius: 8px; padding: 20px; margin-bottom: 12px; cursor: pointer; transition: all 0.2s; }
    .sop-card:hover { border-color: #58a6ff; }
    /* Body text is 11pt everywhere. */
    [data-testid="stAppViewContainer"], [data-testid="stSidebar"] { font-size: 11pt; }
    [data-testid="stAppViewContainer"] p,
    [data-testid="stAppViewContainer"] li,
    [data-testid="stAppViewContainer"] label,
    [data-testid="stAppViewContainer"] input,
    [data-testid="stAppViewContainer"] textarea,
    [data-testid="stAppViewContainer"] button,
    [data-testid="stAppViewContainer"] div[data-testid="stMarkdownContainer"] p,
    [data-testid="stSidebar"] p, [data-testid="stSidebar"] label { font-size: 11pt !important; }
    /* Grids are 10pt. Declared after the 11pt rules so they win. */
    .ag-right-aligned-header .ag-header-cell-label { justify-content: flex-end !important; }
    .ag-theme-streamlit, .ag-theme-streamlit .ag-cell,
    .ag-theme-streamlit .ag-header-cell-label,
    .ag-theme-streamlit input, .ag-theme-streamlit .ag-select,
    div[data-testid="stDataFrame"] table,
    div[data-testid="stDataFrame"] td, div[data-testid="stDataFrame"] th,
    div[data-testid="stDataEditor"] table,
    div[data-testid="stDataEditor"] td, div[data-testid="stDataEditor"] th { font-size: 10pt !important; }
    .ag-theme-streamlit .ag-cell { padding: 2px 6px !important; }
    .ag-theme-streamlit .ag-row { height: 30px !important; }
    .ag-theme-streamlit .ag-header-row { height: 34px !important; }
    .sop-section { border: 1px solid rgba(255,255,255,0.1); border-radius: 6px; padding: 8px 12px; background: rgba(255,255,255,0.02); margin-bottom: 10px; }
    .sop-heading { font-weight: 600; margin-top: 8px; text-transform: uppercase; letter-spacing: 0.5px; font-size: 0.85rem; color: #c8d4ff; }
    .sop-text { margin: 2px 0 4px; line-height: 1.2; }
    .sop-steps { margin: 4px 0 8px 18px; padding-left: 12px; line-height: 1.25; }
    .sop-steps li { margin-bottom: 2px; }
    .sop-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 8px; margin: 6px 0 10px; }
    .sop-grid-card { border: 1px solid rgba(255,255,255,0.08); border-radius: 6px; padding: 6px 10px; background: rgba(255,255,255,0.03); }
    .sop-grid-title { font-size: 0.8rem; text-transform: uppercase; letter-spacing: 0.4px; color: #9fb7ff; margin-bottom: 4px; }
    .sop-grid-card ul { margin: 0; padding-left: 16px; line-height: 1.2; }
    .sop-table { width: 100%; border-collapse: collapse; margin: 8px 0 12px; font-size: 0.85rem; }
    .sop-table th { background: rgba(100,140,255,0.15); color: #c8d4ff; text-align: left; padding: 6px 10px; border: 1px solid rgba(255,255,255,0.1); font-weight: 600; }
    .sop-table td { padding: 5px 10px; border: 1px solid rgba(255,255,255,0.08); }
    .sop-table tr:nth-child(even) { background: rgba(255,255,255,0.03); }
    .sop-table tr:hover { background: rgba(100,140,255,0.08); }
    .mobile-card { background: #161b22; border: 1px solid #2d333b; border-radius: 8px; padding: 12px 16px; margin-bottom: 8px; }
    .mobile-card-header { font-size: 15px; font-weight: 700; color: #58a6ff; margin-bottom: 6px; border-bottom: 1px solid #2d333b; padding-bottom: 6px; }
    .mobile-card-row { display: flex; justify-content: space-between; padding: 2px 0; font-size: 13px; }
    .mobile-card-label { color: #8b949e; }
    .mobile-card-value { color: #e6edf3; font-weight: 500; text-align: right; }
    .mobile-card-vacant { border-left: 3px solid #f85149; }
</style>
""", unsafe_allow_html=True)

# --- MOBILE VIEW TOGGLE ---
if 'mobile_view' not in st.session_state:
    st.session_state.mobile_view = False


def show_mobile_cards(df, key_col=None, header_col='Tenant', highlight_cols=None, skip_cols=None):
    """Render a DataFrame as stacked mobile-friendly cards."""
    if skip_cols is None:
        skip_cols = set()
    if highlight_cols is None:
        highlight_cols = ['Monthly', 'Annual', 'MTE', 'Status']
    for _, row in df.iterrows():
        header = str(row.get(header_col, ''))
        is_vacant = 'VACANT' in header.upper() if header else False
        card_class = 'mobile-card mobile-card-vacant' if is_vacant else 'mobile-card'
        rows_html = []
        for col in df.columns:
            if col == header_col or col in skip_cols:
                continue
            val = row[col]
            if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() in ('', '-', '—'):
                continue
            bold = 'font-weight:700;' if col in highlight_cols else ''
            rows_html.append(
                f'<div class="mobile-card-row">'
                f'<span class="mobile-card-label">{col}</span>'
                f'<span class="mobile-card-value" style="{bold}">{val}</span>'
                f'</div>'
            )
        html = f'<div class="{card_class}"><div class="mobile-card-header">{header}</div>{"".join(rows_html)}</div>'
        st.markdown(html, unsafe_allow_html=True)


def ensure_json_file(path, default):
    if not os.path.exists(path):
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(default, f)
        except Exception:
            return default
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data if isinstance(data, dict) else default
    except Exception:
        return default


def _read_gsheet_config(tab_name):
    """Read a JSON config blob stored in cell A1 of a Google Sheet tab. Uses session_state cache."""
    cache_key = f"_gsheet_cfg_{tab_name}"
    if cache_key in st.session_state:
        return st.session_state[cache_key]
    try:
        sheet = get_gsheet()
        if not sheet:
            return None
        ws = sheet.worksheet(tab_name)
        val = ws.acell('A1').value
        if val:
            if val.startswith("__GZIP64__"):
                payload = base64.b64decode(val[len("__GZIP64__"):])
                val = gzip.decompress(payload).decode("utf-8")
            data = json.loads(val)
            st.session_state[cache_key] = data
            return data
    except Exception:
        pass
    return None


def _write_gsheet_config(tab_name, data):
    """Write a JSON config blob to cell A1 of a Google Sheet tab. Creates tab if needed."""
    try:
        sheet = get_gsheet()
        if not sheet:
            return False
        try:
            ws = sheet.worksheet(tab_name)
        except Exception:
            ws = sheet.add_worksheet(title=tab_name, rows=1, cols=1)
        payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
        if len(payload) > 45000:
            payload = "__GZIP64__" + base64.b64encode(gzip.compress(payload.encode("utf-8"), compresslevel=9)).decode("ascii")
        ws.update_acell('A1', payload)
        # Clear session_state cache for this tab
        cache_key = f"_gsheet_cfg_{tab_name}"
        if cache_key in st.session_state:
            del st.session_state[cache_key]
        return True
    except Exception:
        return False


def read_column_config():
    # Try Google Sheets first (persistent across deploys), fall back to local JSON
    data = _read_gsheet_config('Config: Columns')
    if data and isinstance(data, dict):
        return data
    return ensure_json_file(COLUMN_CONFIG_FILE, {})


def write_column_config(data):
    # Write to Google Sheets (primary) and local JSON (backup)
    _write_gsheet_config('Config: Columns', data)
    try:
        with open(COLUMN_CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass


def read_expense_config():
    default = {name: 0 for name in BUILDING_MAP.keys()}
    data = _read_gsheet_config('Config: Expenses')
    if data and isinstance(data, dict):
        return data
    return ensure_json_file(EXPENSE_CONFIG_FILE, default)


def write_expense_config(data):
    _write_gsheet_config('Config: Expenses', data)
    try:
        with open(EXPENSE_CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass


def get_column_width_overrides(tab_key):
    if not tab_key:
        return {}
    data = read_column_config()
    cfg = data.get(tab_key, {})
    clean = {}
    for col, width in cfg.items():
        if width == '' or width is None:
            continue
        try:
            clean[col] = int(width)
        except (ValueError, TypeError):
            continue
    return clean


def get_column_order(tab_key, default_cols):
    """Return column list reordered per config. Columns in config come first (in config order),
    then any remaining columns not in config keep their original order."""
    if not tab_key:
        return default_cols
    data = read_column_config()
    cfg = data.get(tab_key, {})
    if not cfg:
        return default_cols
    config_cols = [c for c in cfg.keys() if c in default_cols]
    remaining = [c for c in default_cols if c not in config_cols]
    return config_cols + remaining


def save_column_width_overrides(tab_key, overrides):
    data = read_column_config()
    data[tab_key] = overrides
    write_column_config(data)


def render_column_config_editor(tab_key, columns):
    if not tab_key or not columns:
        return
    existing = read_column_config().get(tab_key, {})
    # Show columns in config order first (preserves reordering), then any new ones
    ordered_cols = [c for c in existing.keys() if c in columns]
    remaining = [c for c in columns if c not in ordered_cols]
    all_cols = ordered_cols + remaining
    default_lines = []
    for col in all_cols:
        width = existing.get(col, '')
        default_lines.append(f"{col},{width}")
    with st.expander(f"⚙️ Column Config — {tab_key.title()} (order + width)"):
        st.write("Reorder lines to change column order. Set width after comma (blank = auto). Save to apply.")
        text = st.text_area(
            "column_widths", value='\n'.join(default_lines), height=150,
            key=f"cfg_text_{tab_key}", label_visibility="collapsed"
        )
        if st.button("Save Column Config", key=f"cfg_save_{tab_key}"):
            from collections import OrderedDict
            new_cfg = OrderedDict()
            for line in text.splitlines():
                parts = [p.strip() for p in line.split(',')]
                if not parts or not parts[0]:
                    continue
                col_name = parts[0]
                width = None
                if len(parts) >= 2 and parts[1]:
                    try:
                        width = int(parts[1])
                    except ValueError:
                        width = None
                new_cfg[col_name] = width if width else ""
            save_column_width_overrides(tab_key, dict(new_cfg))
            st.success("Saved! Column order + widths updated.")
            st.rerun()


def render_expense_editor():
    cfg = read_expense_config()
    with st.expander("💸 Building Expense Config"):
        st.caption("Enter annual operating expenses per building. Used for CAM reimbursement and NOI calcs.")
        updated = {}
        for building in BUILDING_MAP.keys():
            val = cfg.get(building, 0) or 0
            updated[building] = st.number_input(
                f"{building} Annual Expenses", min_value=0.0, value=float(val), step=1000.0,
                key=f"expense_{building}"
            )
        if st.button("Save Expenses", key="save_expenses"):
            write_expense_config({k: float(v) for k, v in updated.items()})
            st.success("Building expenses saved. Reload the page to apply.")
            st.rerun()


def show_grid(df, key, height=None, fit_columns=True, pinned_bottom=None, column_configs=None, tab_key=None):
    """Render a DataFrame as an AG Grid with compact columns and dark theme."""
    import pandas as pd
    # Reorder columns based on config
    if tab_key:
        ordered_cols = get_column_order(tab_key, list(df.columns))
        df = df[ordered_cols]

    # Mobile view: render cards instead of grid
    if st.session_state.get('mobile_view', False):
        header_col = 'Tenant' if 'Tenant' in df.columns else ('Last Tenant' if 'Last Tenant' in df.columns else df.columns[0])
        skip = {'TTE Label', 'NNN', 'Is_NNN'}
        show_mobile_cards(df, header_col=header_col, skip_cols=skip)
        return

    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        sortable=True,
        filterable=True,
        resizable=True,
        suppressSizeToFit=False,
        autoWidth=True,
        wrapHeaderText=True,
        autoHeaderHeight=True,
        cellStyle={'text-align': 'right'},
        headerClass='ag-right-aligned-header',
    )

    # Set narrow column widths based on content type
    for col in df.columns:
        sample = df[col].astype(str)
        max_len = max(sample.str.len().max(), len(col))
        if max_len <= 4:
            width = 60
        elif max_len <= 8:
            width = 80
        elif max_len <= 15:
            width = 110
        elif max_len <= 25:
            width = 150
        else:
            width = 200
        gb.configure_column(col, width=width, minWidth=50)

    # Make Tenant column wider
    if 'Tenant' in df.columns:
        gb.configure_column('Tenant', width=160, minWidth=120)
    if 'Building' in df.columns:
        gb.configure_column('Building', width=130, minWidth=100)

    # Render 'View' column as a clickable link (opens PDF in new tab)
    if 'View' in df.columns:
        # Use HTML cell renderer with proper escaping for the URL
        view_renderer = JsCode("""
            class ViewCellRenderer {
                init(params) {
                    this.eGui = document.createElement('div');
                    this.eGui.style.textAlign = 'center';
                    if (params.value) {
                        this.eGui.innerHTML = '<a href="' + params.value + '" target="_blank" rel="noopener" style="color:#58a6ff;text-decoration:none;font-weight:600;">📄 View</a>';
                    } else {
                        this.eGui.innerHTML = '';
                    }
                }
                getGui() {
                    return this.eGui;
                }
            }
        """)
        gb.configure_column('View', cellRenderer=view_renderer, width=90, minWidth=70, suppressSizeToFit=True, filter=False, sortable=False)

    if column_configs:
        for col, cfg in column_configs.items():
            if col in df.columns:
                gb.configure_column(col, **cfg)

    overrides = get_column_width_overrides(tab_key)
    has_overrides = bool(overrides)
    for col, width in overrides.items():
        if col in df.columns:
            gb.configure_column(col, width=width, initialWidth=width, minWidth=width, maxWidth=width + 50, suppressSizeToFit=True, suppressAutoSize=True)

    grid_options = gb.build()

    # If user has configured widths, don't let fit_columns override them
    if has_overrides:
        fit_columns = False
        grid_options['suppressColumnVirtualisation'] = True

    # Add pinned bottom row for totals
    if pinned_bottom is not None:
        grid_options['pinnedBottomRowData'] = pinned_bottom
        grid_options['getRowStyle'] = JsCode("""
            function(params) {
                if (params.node.rowPinned === 'bottom') {
                    return {'background-color': '#1a3a5c', 'font-weight': '600', 'border-top': '2px solid #58a6ff'};
                }
            }
        """)

    # Calculate height
    if height is None:
        row_count = len(df) + (1 if pinned_bottom else 0)
        height = min(max(row_count * 30 + 40, 100), 600)

    try:
        AgGrid(
            df,
            gridOptions=grid_options,
            height=height,
            theme='streamlit',
            update_mode=GridUpdateMode.NO_UPDATE,
            fit_columns_on_grid_load=fit_columns,
            key=key,
            allow_unsafe_jscode=True,
        )
    except Exception:
        # Fallback to native Streamlit dataframe if AG Grid fails
        st.dataframe(df, height=height, use_container_width=True)


# --- GOOGLE SHEETS CONNECTION ---
@st.cache_resource(ttl=300)
def get_gspread_client():
    scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
    if SERVICE_ACCOUNT_FILE:
        creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=scopes)
    elif SERVICE_ACCOUNT_INFO:
        creds = Credentials.from_service_account_info(SERVICE_ACCOUNT_INFO, scopes=scopes)
    else:
        return None
    return gspread.authorize(creds)

@st.cache_resource(ttl=300)
def get_gsheet():
    """Open the dashboard spreadsheet without allowing a Google outage to crash the app."""
    try:
        gc = get_gspread_client()
    except Exception as exc:
        print(f"Google Sheets authorization unavailable: {type(exc).__name__}: {exc}")
        return None
    if gc is None:
        return None

    last_error = None
    for attempt in range(3):
        try:
            return gc.open_by_key(SHEET_ID)
        except gspread.exceptions.APIError as exc:
            last_error = exc
            response = getattr(exc, "response", None)
            status_code = getattr(response, "status_code", None)
            # Retry only transient quota/server failures. Permission errors should fail fast.
            if status_code not in (429, 500, 502, 503, 504) or attempt == 2:
                break
            time.sleep(1.5 * (2 ** attempt))
        except Exception as exc:
            last_error = exc
            break

    if last_error is not None:
        response = getattr(last_error, "response", None)
        status_code = getattr(response, "status_code", "unknown")
        print(f"Google Sheets unavailable (HTTP {status_code}): {type(last_error).__name__}")
    return None


@st.cache_data(ttl=120)
def get_activity_data():
    sheet = get_gsheet()
    if not sheet:
        return []
    try:
        ws = sheet.worksheet('Vacancy Activity')
        records = ws.get_all_records()
        return records
    except Exception as e:
        st.error(f"Error reading activity data: {e}")
        return []


def add_activity_entry(entry):
    sheet = get_gsheet()
    if not sheet:
        return False
    try:
        ws = sheet.worksheet('Vacancy Activity')
        ws.append_row([
            entry['date'], entry['space'], entry['prospect'], entry['broker'],
            entry['type'], entry['feedback'], entry.get('added_by', ''),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ], value_input_option='USER_ENTERED')
        get_activity_data.clear()
        return True
    except Exception as e:
        st.error(f"Error adding entry: {e}")
        return False


def delete_activity_entry(row_idx):
    """Delete a row from Vacancy Activity (row_idx is 0-based from records, +2 for header+1-index)."""
    sheet = get_gsheet()
    if not sheet:
        return False
    try:
        ws = sheet.worksheet('Vacancy Activity')
        ws.delete_rows(row_idx + 2)
        get_activity_data.clear()
        return True
    except Exception as e:
        st.error(f"Error deleting: {e}")
        return False


@st.cache_data(ttl=120)
def get_marketing_data():
    sheet = get_gsheet()
    if not sheet:
        return []
    try:
        ws = sheet.worksheet('Marketing')
        return ws.get_all_records()
    except Exception:
        return []


def add_marketing_entry(entry):
    sheet = get_gsheet()
    if not sheet:
        return False
    try:
        ws = sheet.worksheet('Marketing')
        ws.append_row([
            entry['space'], entry['description'], entry['url'],
            entry.get('added_by', ''), datetime.now().strftime('%Y-%m-%d'),
            'Active'
        ], value_input_option='USER_ENTERED')
        get_marketing_data.clear()
        return True
    except Exception as e:
        st.error(f"Error: {e}")
        return False


def delete_marketing_entry(row_idx):
    sheet = get_gsheet()
    if not sheet:
        return False
    try:
        ws = sheet.worksheet('Marketing')
        ws.delete_rows(row_idx + 2)
        return True
    except Exception:
        return False


# --- TENANCY DATA ---
@st.cache_data(ttl=300)
def load_tenancy():
    if not TENANCY_FILE:
        return [], {}, []
    wb = openpyxl.load_workbook(TENANCY_FILE, data_only=True)
    ws = wb['Sheet1']

    # Summary rows (10-40)
    summary_header = [ws.cell(10, c).value for c in range(1, 25)]
    summaries = []
    for r in range(11, 41):
        row = {}
        for c, h in enumerate(summary_header, 1):
            if h:
                row[h] = ws.cell(r, c).value
        if row.get('Buidling'):
            summaries.append(row)

    # Term detail rows: auto-detect the header row (the one starting with 'ID').
    # Historically row 45, but inserts can shift it (e.g. to 46). Search 43-50.
    detail_header_row = 45
    for hr in range(43, 51):
        if str(ws.cell(hr, 1).value).strip() == 'ID':
            detail_header_row = hr
            break
    detail_header = [ws.cell(detail_header_row, c).value for c in range(1, 24)]
    details = {}
    for r in range(detail_header_row + 1, ws.max_row + 1):
        row = {}
        for c, h in enumerate(detail_header, 1):
            if h:
                val = ws.cell(r, c).value
                row[h] = val.strip() if isinstance(val, str) else val
        tid = row.get('ID')
        if not tid or tid == '-':
            continue
        if tid not in details:
            details[tid] = []
        details[tid].append(row)

    wb.close()

    # Building percentages come from New Sqft in the MSP tenancy summary.
    # This gives every tenant—including gross leases—a share of the building,
    # and the shares add to 100% within each building.
    def _norm_key(value):
        return str(value).strip().lower() if value is not None else ''

    summary_by_tenant = {}
    building_sqft = {}
    for s in summaries:
        b = _norm_key(s.get('Buidling'))
        tenant = _norm_key(s.get('Tenant Name'))
        unit = _norm_key(s.get('Unit'))
        try:
            new_sqft = float(s.get('New Sqft', 0) or 0)
        except (TypeError, ValueError):
            new_sqft = 0.0
        if b and new_sqft > 0:
            building_sqft[b] = building_sqft.get(b, 0.0) + new_sqft
            try:
                new_cam = float(s.get('New CAM', 0) or 0)
            except (TypeError, ValueError):
                new_cam = 0.0
            summary_by_tenant[(b, unit, tenant)] = {
                'sqft': new_sqft,
                'new_cam': new_cam,
            }

    # Build tenant list
    tenants = []
    def to_date(val):
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        return None

    for tid, rows in details.items():
        first = rows[0]
        rows_sorted = sorted(rows, key=lambda r: to_date(r.get('Start Date')) or date.min)
        current_row = rows_sorted[0]
        next_row = None
        # Find current period: tightest fit (start <= TODAY <= end with smallest end date)
        best_end = None
        has_current_period = False
        for r in rows_sorted:
            start = to_date(r.get('Start Date'))
            end = to_date(r.get('End Date'))
            if start and end and start <= TODAY <= end:
                has_current_period = True
                if best_end is None or end < best_end:
                    best_end = end
                    current_row = r

        # Future tenant: no period contains TODAY and the earliest start is in the future
        earliest_start = None
        for r in rows_sorted:
            s = to_date(r.get('Start Date'))
            if s and (earliest_start is None or s < earliest_start):
                earliest_start = s
        is_future = (not has_current_period) and (earliest_start is not None) and (earliest_start > TODAY)
        # For a future tenant, anchor display to the first (commencement) period
        if is_future:
            current_row = rows_sorted[0]
            commence_date = earliest_start
            # Use first period that actually carries rent (>0) for the display figures,
            # so a $0 free-rent first month doesn't make the row look empty.
            for r in rows_sorted:
                try:
                    m = float(r.get('Monthly', 0) or 0)
                except (TypeError, ValueError):
                    m = 0
                if m > 0:
                    current_row = r
                    break
        else:
            commence_date = None
        # Find next row: the period starting right after current_row's end
        current_end = to_date(current_row.get('End Date'))
        for r in rows_sorted:
            start = to_date(r.get('Start Date'))
            end = to_date(r.get('End Date'))
            if end and current_end and end > current_end:
                if start and start <= TODAY:
                    # This is a future year with start=commencement — use end date ordering
                    if not next_row or end < to_date(next_row.get('End Date')):
                        next_row = r
                elif start and start > TODAY:
                    if not next_row or start < to_date(next_row.get('Start Date')):
                        next_row = r

        last_orig_end = None
        for r in rows:
            if r.get('Term') == 'Original' and r.get('End Date'):
                ed = r['End Date']
                if isinstance(ed, (datetime, date)):
                    if last_orig_end is None or ed > last_orig_end:
                        last_orig_end = ed

        tte_days = None
        tte_months = 0
        tte_label = 'MTM'
        if last_orig_end:
            if isinstance(last_orig_end, datetime):
                last_orig_end = last_orig_end.date()
            tte_days = (last_orig_end - TODAY).days
            if tte_days < 0:
                tte_months = 0
                tte_label = 'MTM'
            else:
                tte_months = max(0, round(tte_days / 30.44))
                tte_label = str(tte_months)
        else:
            tte_days = -1
            tte_months = 0
            tte_label = 'MTM'

        option_years = len([r for r in rows if r.get('Term', '').startswith('Option')])

        # Find first option expiration date (if exercised, when would it expire)
        first_option_exp = None
        option_rows = [r for r in rows if r.get('Term', '').startswith('Option')]
        if option_rows:
            # Sort by Term (Option_1, Option_2, etc.) and get the first one's end date
            option_rows_sorted = sorted(option_rows, key=lambda r: r.get('Term', ''))
            first_opt = option_rows_sorted[0]
            opt_end = to_date(first_opt.get('End Date'))
            if opt_end:
                first_option_exp = opt_end

        try:
            gross_mod = float(current_row.get('Gross Mod', 0) or 0)
        except (TypeError, ValueError):
            gross_mod = 0
        try:
            monthly = float(current_row.get('Monthly', 0) or 0) + gross_mod
        except (TypeError, ValueError):
            monthly = gross_mod
        try:
            annual = float(current_row.get('Annual', 0) or 0) + (gross_mod * 12)
        except (TypeError, ValueError):
            annual = gross_mod * 12
        try:
            sqft = float(first.get('Sqft', 0) or 0)
        except (TypeError, ValueError):
            sqft = 0
        psf = (annual / sqft) if sqft and sqft > 1 else 0
        building_key = _norm_key(first.get('Building'))
        summary_rec = summary_by_tenant.get((
            building_key, _norm_key(first.get('Space')), _norm_key(first.get('Tenant'))
        ), {})
        summary_sqft = summary_rec.get('sqft', sqft) if isinstance(summary_rec, dict) else summary_rec
        building_pct = (summary_sqft / building_sqft[building_key]
                        if building_key in building_sqft and building_sqft[building_key] > 0 else 0)
        new_cam_pct = (summary_rec.get('new_cam', 0) if isinstance(summary_rec, dict) else 0)
        try:
            sec_dep = float(first.get('Sec Dep', 0) or 0)
        except (TypeError, ValueError):
            sec_dep = 0

        # Exp Date = last original period end date
        exp_dt = last_orig_end
        exp_str = exp_dt.strftime('%m/%d/%Y') if isinstance(exp_dt, (datetime, date)) else 'MTM'

        # Next Anniversary = end date of current period
        current_end = to_date(current_row.get('End Date'))
        next_anniv = current_end if current_end and current_end > TODAY else None

        # Next rent = from the period AFTER current
        next_gross_mod = (next_row.get('Gross Mod', 0) or 0) if next_row else 0
        next_monthly = ((next_row.get('Monthly', 0) or 0) + next_gross_mod) if next_row else None
        delta_monthly = (next_monthly - monthly) if (next_row and next_monthly is not None) else None
        anniv_months = None
        if next_anniv:
            diff_days = (next_anniv - TODAY).days
            anniv_months = max(0, round(diff_days / 30.44))

        # Find nearest future cancel date
        cancel_date = None
        for r in rows:
            cd = to_date(r.get('Cancel Date'))
            if cd and cd > TODAY:
                if cancel_date is None or cd < cancel_date:
                    cancel_date = cd
        cancel_str = cancel_date.strftime('%m/%d/%Y') if cancel_date else '—'

        # Get tenant type (Retail/Office/etc) from summary rows
        tenant_type = ''
        tenant_name = first.get('Tenant', '')
        for s in summaries:
            if s.get('Tenant Name') == tenant_name and s.get('Buidling') == first.get('Building'):
                tenant_type = s.get('Type', '') or ''
                break

        tenants.append({
            'Building': first.get('Building', ''),
            'Space': first.get('Space', ''),
            'Tenant': first.get('Tenant', ''),
            'Type': tenant_type,
            'Floor': first.get('Floor', ''),
            'SF': sqft,
            'Lease': first.get('Type', ''),
            'Monthly': round(monthly, 2),
            'Annual': round(annual, 2),
            'PSF': round(psf, 2),
            'Building_Pct': building_pct,
            # New CAM is the workbook's expense-allocation percentage.
            'CAM_Pct': new_cam_pct,
            'Is_NNN': True if str(first.get('Type', '')).upper().startswith('NNN') else False,
            'TTE': '0' if tte_label == 'MTM' else tte_label,
            'TTE_Label': tte_label,
            'TTE_Months': tte_months,
            'TTE_Days': tte_days if tte_days is not None else 9999,
            'Options': f"{option_years}yr" if option_years > 0 else '-',
            'First_Option_Exp': first_option_exp.strftime('%m/%d/%Y') if first_option_exp else '-',
            'Exp Date': exp_str,
            'Cancel Date': cancel_str,
            'Escalation': current_row.get('Escalation', 0) or 0,
            'Sec Dep': round(sec_dep, 2),
            'Next Anniv': next_anniv,
            'Anniv_Months': anniv_months,
            'Next Monthly': round(next_monthly, 2) if next_monthly is not None else None,
            'Delta Monthly': round(delta_monthly, 2) if delta_monthly is not None else None,
            'Future': is_future,
            'Commence': commence_date.strftime('%m/%d/%Y') if commence_date else '',
            'Commence_Date': commence_date,
        })

    tenants.sort(key=lambda t: (t['Building'], t['TTE_Days']))
    return tenants, details, summaries


# --- COI SCANNING ---
def extract_cert_info(pdf_path):
    if not HAS_PYMUPDF:
        return None, None
    try:
        doc = fitz.open(pdf_path)
        page = doc[0]
        full_text = page.get_text("text")
        all_dates = []
        for m in re.finditer(r'(\d{1,2}/\d{1,2}/\d{4})', full_text):
            try:
                dt = datetime.strptime(m.group(1), '%m/%d/%Y')
                all_dates.append(dt)
            except ValueError:
                pass
        exp_date = max(all_dates) if all_dates else None

        insured_name = None
        blocks = page.get_text("blocks")
        text_blocks = sorted([b for b in blocks if b[6] == 0], key=lambda b: (b[1], b[0]))
        insured_label_y = None
        for b in text_blocks:
            first_line = b[4].strip().split('\n')[0].strip().upper()
            if b[0] < page.rect.width * 0.5 and first_line in ('INSURED', 'NAMED INSURED'):
                insured_label_y = b[1]
                break
        if insured_label_y is not None:
            for b in text_blocks:
                text = b[4].strip().split('\n')[0].strip()
                if b[1] <= insured_label_y or b[0] > page.rect.width * 0.4:
                    continue
                if b[1] > insured_label_y + 50:
                    break
                if len(text) < 3 or re.match(r'^\d+\s', text):
                    continue
                if re.match(r'^[A-Za-z\s]+,?\s*[A-Z]{2}\s+\d{5}', text):
                    continue
                if any(kw in text.upper() for kw in ('INSURER', 'CERTIFICATE', 'REVISION')):
                    continue
                insured_name = text
                break
        doc.close()
        return exp_date, insured_name
    except Exception:
        return None, None


# Tenant name aliases: maps alternate / rebranded / d-b-a insured names (as they
# appear on COI certs) to the canonical roster tenant name. Add entries here when a
# tenant's insurance is issued under a different legal name than the roster shows.
# Keys and values are matched case-insensitively.
TENANT_ALIASES = {
    # City Property USA rebranded to Village Practice Management (15 South, Unit 1).
    # "village practice man" covers the 20-char-truncated filename form too.
    "village practice man": "City Property USA NJ, LLC",
    "village practice management": "City Property USA NJ, LLC",
    "village practice management company": "City Property USA NJ, LLC",
    "village practice management company, llc": "City Property USA NJ, LLC",
}


def _apply_tenant_alias(name):
    """Return the canonical roster name if `name` (or a prefix of it) is a known alias."""
    if not name:
        return name
    key = name.strip().lower().rstrip(' .,')
    if key in TENANT_ALIASES:
        return TENANT_ALIASES[key]
    # also try alias keys as substrings (cert name may carry extra text)
    for alias, canonical in TENANT_ALIASES.items():
        if alias in key:
            return canonical
    return name


def fuzzy_match_tenant(tenant_name, cert_name):
    if not tenant_name or not cert_name:
        return False
    # Normalize known aliases on BOTH sides to the canonical roster name.
    tenant_name = _apply_tenant_alias(tenant_name)
    cert_name = _apply_tenant_alias(cert_name)
    t, c = tenant_name.upper().strip(), cert_name.upper().strip()
    if t in c or c in t:
        return True
    skip = {'LLC', 'INC', 'CORP', 'LTD', 'DBA', 'THE', 'OF', 'AND', '&', 'PDF', 'COI', 'CERTIFICATE', 'INSURANCE'}
    t_words = [w for w in re.split(r'[\s_\-\.]+', t) if w not in skip and len(w) > 2]
    c_words = [w for w in re.split(r'[\s_\-\.]+', c) if w not in skip and len(w) > 2]
    if t_words and c_words and t_words[0] == c_words[0]:
        return True
    return len(set(t_words) & set(c_words)) >= 1


COI_RAW_BASE = "https://raw.githubusercontent.com/asollog61/msp-dashboard/master/data/coi"


def build_coi_url(building_name, filename):
    """Build a clickable GitHub raw URL for a COI PDF (URL-encoded)."""
    from urllib.parse import quote
    if not filename:
        return ''
    return f"{COI_RAW_BASE}/{quote(building_name)}/{quote(filename)}"


@st.cache_data(ttl=3600)
def scan_coi_files():
    """Returns (coi_data, building_coi_data, pm_coi_data): tenant certs, building-level certs, and PM certs per building."""
    coi_data = {}
    building_coi_data = {}
    pm_coi_data = {}
    # Check for COIs bundled in data/coi/{building}/ (Streamlit Cloud) or ACTIVE_PROPERTIES (local)
    bundled_coi = Path(os.path.dirname(__file__)) / "data" / "coi"
    for building_name, mapping in BUILDING_MAP.items():
        cert_dir = None
        # Priority 1: bundled COI folder (works on Streamlit Cloud)
        candidate = bundled_coi / building_name
        if candidate.exists():
            cert_dir = candidate
        # Priority 2: ACTIVE_PROPERTIES local path
        elif ACTIVE_PROPS_ROOT:
            cert_dir = ACTIVE_PROPS_ROOT / mapping["dest_folder"] / mapping["share"] / "Certificates of Insurance"
            if not cert_dir.exists():
                parent = ACTIVE_PROPS_ROOT / mapping["dest_folder"] / mapping["share"]
                cert_dir = None
                if parent.exists():
                    for d in parent.iterdir():
                        if d.is_dir() and 'certificate' in d.name.lower() and 'insurance' in d.name.lower():
                            cert_dir = d
                            break
        certs = []
        building_certs = []
        pm_certs = []
        if cert_dir and cert_dir.exists():
            for f in sorted(cert_dir.iterdir()):
                if not f.is_file() or f.suffix.lower() != '.pdf':
                    continue
                fname_lower = f.name.lower()
                # Building-level COI: filename contains _building (case-insensitive)
                is_building = '_building' in fname_lower
                exp_date, insured_name = extract_cert_info(str(f))
                if not insured_name:
                    m = re.match(r'Exp_(\d{8})_MSP\d+_(.+)\.pdf', f.name, re.IGNORECASE)
                    if m:
                        if not exp_date:
                            try:
                                exp_date = datetime.strptime(m.group(1), '%Y%m%d')
                            except ValueError:
                                pass
                        insured_name = m.group(2).replace('_', ' ')
                # PM COI: insured name contains "Proventus" (property manager)
                is_pm = insured_name and 'proventus' in insured_name.lower()
                rec = {'filename': f.name, 'exp_date': exp_date, 'insured_name': insured_name}
                if is_pm:
                    pm_certs.append(rec)
                elif is_building:
                    building_certs.append(rec)
                else:
                    certs.append(rec)
        coi_data[building_name] = certs
        building_coi_data[building_name] = building_certs
        pm_coi_data[building_name] = pm_certs
    return coi_data, building_coi_data, pm_coi_data


# --- YARDI PDF PARSING ---
def normalize_space(space_str, building=None):
    """Normalize Yardi space format to dashboard format.
    For 114 Central: 102->2, 104->4, 106->6, 108->8 (first floor even units)
    For 2-X format: 2-1->201, 2-2->202, etc.
    For other formats: keep as-is (101, 201, 1286, A-1, etc.)
    """
    s = str(space_str).strip()
    
    # Handle 2-1 format (second floor units)
    if '-' in s:
        parts = s.split('-')
        if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
            return f"{parts[0]}{parts[1].zfill(2)}"
    
    # Keep 3-digit units as-is (102, 104, 106, 108, 201, etc.)
    # Previously stripped leading "10" but spreadsheet now uses full 3-digit format
    
    # For other formats (101, 201, 1286, A-1, O-1, etc.), keep as-is
    return s


@st.cache_data(ttl=3600)
def _get_yardi_dir():
    """Return the Yardi PDF directory path, or None."""
    for p in [
        os.path.join(os.path.dirname(__file__), "data", "Yardi"),
        "/home/node/OpenClaw/Share Jason/General Procedures/msp-dashboard-src/data/Yardi/",
    ]:
        if os.path.exists(p):
            return Path(p)
    return None


YARDI_FILENAME_MAP = {
    "1280-springfield": "1280 Springfield",
    "15-south": "15 South",
    "36-south": "36 South",
    "114-central": "114 Central",
}

MONTH_ORDER = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}


def _filter_latest_per_building(pdfs):
    """Given a list of PDF paths, return only the most recent per building.
    Filenames start with Mon-YYYY (e.g., May-2026_...). Returns list of Paths."""
    building_best = {}  # building_key -> (sort_key, path)
    for pdf_path in pdfs:
        fname = pdf_path.stem.lower()
        # Identify building
        building_key = None
        for key in YARDI_FILENAME_MAP:
            if key in fname:
                building_key = key
                break
        if not building_key:
            continue
        # Parse date from filename: Mon-YYYY
        match = re.match(r'^([a-z]{3})-(\d{4})', fname)
        if match:
            mon_str, year_str = match.group(1), match.group(2)
            sort_key = int(year_str) * 100 + MONTH_ORDER.get(mon_str, 0)
        else:
            sort_key = 0
        prev = building_best.get(building_key)
        if not prev or sort_key > prev[0]:
            building_best[building_key] = (sort_key, pdf_path)
    return [v[1] for v in building_best.values()]


def parse_yardi_deposit_activity(latest_only=True):
    """Parse Security Deposit Activity pages from Yardi PDFs.
    Returns dict keyed by 'Building|Space' with Deposits On Hand amount.
    If latest_only=True, only parses the most recent report per building."""
    if not HAS_PYMUPDF:
        return {}

    yardi_dir = _get_yardi_dir()
    if not yardi_dir or not yardi_dir.exists():
        return {}

    filename_map = YARDI_FILENAME_MAP

    deposit_data = {}

    all_pdfs = sorted(yardi_dir.glob("*.pdf"))
    pdf_list = _filter_latest_per_building(all_pdfs) if latest_only else all_pdfs

    for pdf_path in pdf_list:
        filename_lower = pdf_path.stem.lower()
        building = None
        for key, name in filename_map.items():
            if key in filename_lower:
                building = name
                break
        if not building:
            continue

        try:
            doc = fitz.open(str(pdf_path))
            for page in doc:
                text = page.get_text()
                if 'Security Deposit Activity' not in text:
                    continue

                lines = [l.strip() for l in text.split('\n') if l.strip()]
                i_line = 0
                while i_line < len(lines):
                    line = lines[i_line]
                    if '(Current)' in line or '(Past)' in line:
                        tenant_name = line
                        # Collect numeric values following the tenant line
                        nums = []
                        j = i_line + 1
                        while j < len(lines) and len(nums) < 7:
                            try:
                                val = float(lines[j].replace(',', ''))
                                nums.append(val)
                                j += 1
                            except (ValueError, AttributeError):
                                break

                        # Find unit number - first try inline (unit + tenant on same line)
                        unit = ''
                        inline_match = re.match(r'^([A-Za-z0-9_-]+)\s+', line)
                        if inline_match:
                            candidate = inline_match.group(1)
                            # Verify it's a unit-like token (has digits or is short)
                            if len(candidate) <= 8 and any(c.isdigit() for c in candidate):
                                unit = candidate

                        # Fallback: look backwards for short alphanumeric token on prior line
                        if not unit:
                            skip_words = {'avenue', 'st', 'st.', 'south', 'central', 'ave', 'ave.',
                                          'd', '36', '15', '1280', '114', 'springfiel', 'property',
                                          'unit', 'tenant', 'prior', 'current', 'deposits', 'total',
                                          'on', 'hand', 'billed', 'receipts', 'forfeited', 'deposit'}
                            for k in range(i_line - 1, max(i_line - 6, 0), -1):
                                candidate = lines[k].strip()
                                if (re.match(r'^[A-Za-z0-9_-]+$', candidate) and
                                        len(candidate) <= 8 and
                                        candidate.lower() not in skip_words):
                                    unit = candidate
                                    break

                        if unit and len(nums) >= 5:
                            deposits_on_hand = nums[4]
                            normalized = normalize_space(unit, building)
                            key = f"{building}|{normalized}"
                            # Only use (Current) entries, or add to existing if (Past)
                            if '(Current)' in tenant_name:
                                deposit_data[key] = deposits_on_hand
                            elif key not in deposit_data:
                                deposit_data[key] = deposits_on_hand

                        i_line = j
                    else:
                        i_line += 1

            doc.close()
        except Exception:
            pass

    return deposit_data


@st.cache_data(ttl=300)
def parse_yardi_rent_rolls(latest_only=True):
    """Parse Yardi monthly statement PDFs from data/Yardi/ folder.
    Returns dict keyed by "Building|Space" with {monthly, cam, expiration, tenant}.
    If latest_only=True, only parses the most recent report per building.
    """
    if not HAS_PYMUPDF:
        return {}
    
    yardi_dir = _get_yardi_dir()
    if not yardi_dir or not yardi_dir.exists():
        return {}
    
    filename_map = YARDI_FILENAME_MAP
    
    yardi_data = {}
    
    all_pdfs = sorted(yardi_dir.glob("*.pdf"))
    pdf_list = _filter_latest_per_building(all_pdfs) if latest_only else all_pdfs

    for pdf_path in pdf_list:
        # Extract building name from filename
        filename_lower = pdf_path.stem.lower()
        building = None
        for key, name in filename_map.items():
            if key in filename_lower:
                building = name
                break
        
        if not building:
            continue
        
        try:
            doc = fitz.open(str(pdf_path))
            
            # Page 3 (0-indexed page 2) is the Monthly Rent Roll
            if len(doc) < 3:
                doc.close()
                continue
            
            page = doc[2]
            
            # Check if this is the Monthly Rent Roll page
            # Look for key headers: "Unit", "Tenant Name", "Actual Rent"
            text = page.get_text()
            if not all(keyword in text for keyword in ["Unit", "Tenant Name", "Actual Rent"]):
                doc.close()
                continue
            
            # Extract text blocks - each tenant is a block
            blocks = page.get_text("blocks")
            
            for block in blocks:
                if block[6] != 0:  # Not a text block
                    continue
                
                content = block[4].strip()
                lines = [l.strip() for l in content.split('\n') if l.strip()]
                
                if len(lines) < 5:
                    continue
                
                # First line should be unit number
                unit = lines[0]
                
                # Skip header rows, totals, and summary sections
                if unit.upper() in ('UNIT', 'TOTAL', 'SUMMARY', 'CURRENT/NOTICE/VACANT', 'FUTURE', 'OCCUPIED', 'TOTALS:', 'GROUPS'):
                    continue
                
                # Skip if doesn't look like a unit number (allow named units like EXTERIOR, ROOF, PARKING)
                NAMED_UNITS = {'EXTERIOR', 'ROOF', 'PARKING', 'ATM', 'PAD', 'BASEMENT', 'STORAGE', 'SIGN', 'BILLBOARD', 'EASEMENT'}
                if not any(c.isdigit() for c in unit) and unit.upper() not in NAMED_UNITS:
                    continue
                
                # Parse the tenant line
                # Expected format: Unit, SqFt, Tenant Name (may be multi-line), Monthly, ..., CAM, ..., dates
                
                # Second item should be SqFt (numeric with .00)
                if len(lines) < 3:
                    continue
                
                # Verify second item looks like a number (SqFt)
                sqft_str = lines[1]
                try:
                    sqft = float(sqft_str.replace(',', ''))
                except (ValueError, AttributeError):
                    continue
                
                # Extract tenant name (may span multiple lines for d/b/a)
                # Start from index 2, stop when we hit monthly rent (a larger number)
                tenant_lines = []
                idx = 2
                while idx < len(lines):
                    line = lines[idx]
                    # Check if this looks like the monthly rent (larger number, usually 3+ digits before decimal)
                    if re.match(r'^[\d,]+\.\d{2}$', line):
                        try:
                            val = float(line.replace(',', ''))
                            # If it's > 100, likely monthly rent; if < 100, might be part of tenant name
                            if val >= 100 or (idx > 2 and val > 0):
                                break
                        except (ValueError, AttributeError):
                            pass
                    tenant_lines.append(line)
                    idx += 1
                
                tenant_name = ' '.join(tenant_lines).strip()
                
                # Skip VACANT units
                if tenant_name.upper() == 'VACANT':
                    continue
                
                # Skip if tenant name is empty or looks like garbage
                if not tenant_name or len(tenant_name) < 2:
                    continue
                
                # Now extract monthly rent (next item after tenant name)
                if idx >= len(lines):
                    continue
                
                monthly_str = lines[idx]
                try:
                    monthly = float(monthly_str.replace(',', ''))
                    # Sanity check - monthly rent should be > 0 (allow named units like easements with $0)
                    if monthly <= 0 and unit.upper() not in NAMED_UNITS:
                        continue
                except (ValueError, AttributeError):
                    continue
                
                # Find CAM amount - typically the 5th numeric field after monthly rent
                # Format: Monthly, PSF, Tenant_Deposit, Other_Deposit, CAM, CAM_PSF, Misc, dates
                # Look for a reasonable CAM value (skip PSF which is small, skip deposits which might be 0)
                cam = 0.0
                numeric_fields = []
                for i in range(idx + 1, min(idx + 10, len(lines))):
                    try:
                        val = float(lines[i].replace(',', ''))
                        numeric_fields.append(val)
                    except (ValueError, AttributeError):
                        continue
                
                # CAM is typically at index 3 or 4 in numeric_fields (after PSF, 2 deposits)
                # Look for the first value > 10 that's reasonable relative to monthly rent
                for val in numeric_fields[2:6]:  # Skip first 2 (PSF, first deposit), check next 4
                    if val > 10 and val < monthly * 3:
                        cam = val
                        break
                
                # Find lease expiration date (format: MM/DD/YYYY)
                exp_date = None
                for i in range(idx + 1, len(lines)):
                    match = re.search(r'(\d{2}/\d{2}/\d{4})', lines[i])
                    if match:
                        date_str = match.group(1)
                        try:
                            exp_date = datetime.strptime(date_str, '%m/%d/%Y').date()
                            # Take the LAST date found (lease expiration, not move-in)
                        except ValueError:
                            pass
                
                # Normalize space number
                normalized_space = normalize_space(unit, building)
                key = f"{building}|{normalized_space}"
                
                # Extract security deposit (numeric_fields[1] = Tenant Deposit, after PSF)
                deposit = 0.0
                if len(numeric_fields) >= 2:
                    deposit = numeric_fields[1]

                yardi_data[key] = {
                    'monthly': monthly,
                    'cam': cam,
                    'deposit': deposit,
                    'expiration': exp_date,
                    'tenant': tenant_name,
                    'raw_unit': unit,
                }
            
            doc.close()
            
        except Exception as e:
            # Silently skip problematic PDFs
            pass
    
    return yardi_data


def compute_yardi_diffs(tenants, yardi_data):
    """Compare dashboard tenants with Yardi data.
    Returns dict keyed by "Building|Space" with comma-separated diff string.
    """
    diffs = {}
    
    for tenant in tenants:
        building = tenant.get('Building', '').strip()
        space = str(tenant.get('Space', '')).strip()
        
        if not building or not space:
            continue
        
        key = f"{building}|{space}"
        
        # Look up in Yardi data
        yardi_entry = yardi_data.get(key)
        
        if not yardi_entry:
            # Try alternate space formats
            # If space is "2", try "102"
            if space.isdigit() and len(space) == 1:
                alt_key = f"{building}|10{space}"
                yardi_entry = yardi_data.get(alt_key)
            # If space is "201", try "2-1"
            elif space.isdigit() and len(space) == 3:
                first = space[0]
                rest = space[1:]
                alt_key = f"{building}|{first}-{rest}"
                yardi_entry = yardi_data.get(alt_key)
        
        if not yardi_entry:
            diffs[key] = "Not in Yardi"
            continue
        
        # Compare monthly rent, CAM, and expiration
        diff_parts = []
        
        # Monthly rent
        dash_monthly = tenant.get('Monthly', 0) or 0
        yardi_monthly = yardi_entry.get('monthly', 0) or 0
        if abs(dash_monthly - yardi_monthly) > 0.01:
            diff_parts.append(f"Monthly: ${dash_monthly:,.0f} vs ${yardi_monthly:,.0f}")
        
        # CAM amount
        # Dashboard stores CAM as percentage of building expenses, Yardi stores actual monthly CAM
        # For comparison, we need building expenses
        # For now, just show if there's a CAM difference
        # We'll calculate actual CAM reimbursement in render function
        
        # Expiration date
        dash_exp_str = tenant.get('Exp Date', '').strip()
        yardi_exp = yardi_entry.get('expiration')
        
        if dash_exp_str and dash_exp_str != 'MTM' and yardi_exp:
            try:
                dash_exp = datetime.strptime(dash_exp_str, '%m/%d/%Y').date()
                if dash_exp != yardi_exp:
                    diff_parts.append(f"Exp: {dash_exp.strftime('%m/%d/%Y')} vs {yardi_exp.strftime('%m/%d/%Y')}")
            except (ValueError, AttributeError):
                pass
        
        diffs[key] = ', '.join(diff_parts) if diff_parts else ''
    
    return diffs


# --- SOP EXTRACTION ---
@st.cache_data(ttl=86400)
def _extract_sop_content(pdf_path, cache_token):
    if not pdf_path or not HAS_PYMUPDF:
        return [], {}
    doc = fitz.open(pdf_path)

    # Extract tables from all pages
    all_tables = []
    for page in doc:
        try:
            tabs = page.find_tables()
            for tab in tabs:
                cells = tab.extract()
                if cells and len(cells) > 1:
                    all_tables.append({'page': page.number, 'y': tab.bbox[1], 'rows': cells})
        except Exception:
            pass

    # Build table HTML lookup keyed by header signature
    table_map = {}
    for tbl in all_tables:
        rows = tbl['rows']
        header = rows[0]
        body = rows[1:]
        th = ''.join(f"<th>{(c or '').strip()}</th>" for c in header)
        tr_list = []
        for row in body:
            td = ''.join(f"<td>{(c or '').strip()}</td>" for c in row)
            tr_list.append(f"<tr>{td}</tr>")
        html = f"<table class='sop-table'><thead><tr>{th}</tr></thead><tbody>{''.join(tr_list)}</tbody></table>"
        # Key by first header cell text for matching
        key = (header[0] or '').strip().lower()
        if key:
            table_map[key] = {'html': html, 'cell_texts': set()}
            for row in rows:
                for cell in row:
                    if cell and len(cell.strip()) > 2:
                        table_map[key]['cell_texts'].add(cell.strip())

    # Standard text extraction
    full_text = ""
    for page in doc:
        full_text += page.get_text() + "\n"
    doc.close()

    section_defs = [
        ("Roles & Definitions", "👥"),
        ("SOP 100 — New Lease or Addendum", "📝"), ("SOP 200 — Insurance Reconciliation", "🛡️"),
        ("SOP 300 — Security Deposit Reconciliation", "💰"), ("SOP 400 — Tenant Move-Out", "🚪"),
        ("SOP 500 — Recurring Reports", "📊"), ("SOP 600 — Vacancy & Lead Management", "🏠"),
        ("Appendix — Numbering, Revision Log & Open Items", "📎"),
    ]

    positions = []
    for title, icon in section_defs:
        idx = full_text.find(title)
        if idx >= 0:
            last_idx = idx
            search_from = idx + 1
            while True:
                next_idx = full_text.find(title, search_from)
                if next_idx < 0:
                    break
                last_idx = next_idx
                search_from = next_idx + 1
            positions.append((last_idx, title, icon))

    positions.sort(key=lambda x: x[0])
    sections = []
    for i, (pos, title, icon) in enumerate(positions):
        end = positions[i + 1][0] if i + 1 < len(positions) else len(full_text)
        content = full_text[pos + len(title):end].strip()
        clean_lines = []
        for line in content.split('\n'):
            line = line.strip()
            if line.startswith('Marion St Properties') or line.startswith('Draft v1.0') or line.startswith('Page '):
                continue
            if '...' in line and len(line) > 40:
                continue
            clean_lines.append(line)
        sections.append({'title': title, 'icon': icon, 'content': '\n'.join(clean_lines).strip()})
    return sections, table_map


def _match_table_for_section(lines, table_map):
    """Check if any extracted table belongs in this section's content. Returns (table_key, set of lines to remove)."""
    matches = []
    for key, tbl in table_map.items():
        cell_texts = tbl['cell_texts']
        matched_indices = set()
        for i, line in enumerate(lines):
            stripped = line.strip()
            if stripped in cell_texts or any(ct in stripped for ct in cell_texts if len(ct) > 5):
                matched_indices.add(i)
        # If we matched enough cells, this table belongs here
        if len(matched_indices) >= 2:
            matches.append((key, matched_indices))
    return matches


def extract_sop_content():
    if not SOP_PDF:
        return [], {}
    try:
        cache_token = os.path.getmtime(SOP_PDF)
    except OSError:
        cache_token = datetime.now().timestamp()
    return _extract_sop_content(SOP_PDF, cache_token)


# =====================
# RENDER TABS
# =====================

def render_sop_tab():
    if not SOP_PDF:
        st.markdown("### 📋 Marion St Properties — Standard Operating Procedures")
        st.warning("SOP Manual PDF not found. Place Marion_St_SOP_Manual.pdf in the data/ folder.")
        return

    # --- Report buttons (SOP index for the PDF; also attach the manual itself) ---
    def _sop_sections():
        import pandas as pd
        sops = [
            ("SOP 100", "New Lease or Addendum"),
            ("SOP 200", "Insurance Reconciliation"),
            ("SOP 300", "Security Deposit Reconciliation"),
            ("SOP 400", "Tenant Move-Out"),
        ]
        pages = ''
        try:
            if HAS_PYMUPDF:
                _d = fitz.open(SOP_PDF)
                pages = f"{len(_d)} pages"
                _d.close()
        except Exception:
            pass
        df = pd.DataFrame([{"SOP": s, "Procedure": n} for s, n in sops])
        return [(f"Standard Operating Procedures Index ({pages})", df)]
    render_report_buttons("sop", "SOPs", _sop_sections,
                          meta="Full SOP manual available in the dashboard SOPs tab.")

    st.markdown("### 📋 Marion St Properties — Standard Operating Procedures")

    if st.session_state.get('mobile_view', False) and HAS_PYMUPDF:
        # Mobile: page-by-page image view with navigation
        import base64
        doc = fitz.open(SOP_PDF)
        total_pages = len(doc)

        if 'sop_page' not in st.session_state:
            st.session_state.sop_page = 0

        page_num = st.session_state.sop_page
        page_num = max(0, min(page_num, total_pages - 1))

        # Render current page as image
        page = doc[page_num]
        mat = fitz.Matrix(1.25, 1.25)  # 1.25x for mobile
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        b64_img = base64.b64encode(img_bytes).decode()
        doc.close()

        # Navigation buttons on top
        col_prev, col_num, col_next = st.columns([1, 2, 1])
        if col_prev.button("⬅️ Prev", disabled=(page_num == 0), key="sop_prev"):
            st.session_state.sop_page = page_num - 1
            st.rerun()
        col_num.markdown(f"<div style='text-align:center; padding-top:8px;'>{page_num + 1} / {total_pages}</div>", unsafe_allow_html=True)
        if col_next.button("Next ➡️", disabled=(page_num >= total_pages - 1), key="sop_next"):
            st.session_state.sop_page = page_num + 1
            st.rerun()

        st.markdown(f'<img src="data:image/png;base64,{b64_img}" style="width:100%; border-radius:6px; border:1px solid rgba(255,255,255,0.1);">', unsafe_allow_html=True)
    else:
        # Desktop: page-by-page image view (same as mobile, higher res)
        import base64
        if HAS_PYMUPDF:
            doc = fitz.open(SOP_PDF)
            total_pages = len(doc)

            if 'sop_page' not in st.session_state:
                st.session_state.sop_page = 0

            page_num = st.session_state.sop_page
            page_num = max(0, min(page_num, total_pages - 1))

            page = doc[page_num]
            mat = fitz.Matrix(2.0, 2.0)
            pix = page.get_pixmap(matrix=mat)
            # Crop whitespace/margins from the rendered page
            img_bytes = pix.tobytes("png")
            # Use fitz to detect content bbox and crop
            try:
                content_rect = page.get_text("dict", flags=0).get("width", 0)
                # Simpler: use CropBox or trim via pixmap
                # Get the actual content bounding box
                clip = page.rect
                # Try trimming margins: typical PDF has ~72pt margins on letter (612x792)
                pw, ph = page.rect.width, page.rect.height
                margin_x = pw * 0.06  # ~6% margins
                margin_top = ph * 0.04
                margin_bot = ph * 0.04
                trimmed = fitz.Rect(margin_x, margin_top, pw - margin_x, ph - margin_bot)
                pix2 = page.get_pixmap(matrix=mat, clip=trimmed)
                img_bytes = pix2.tobytes("png")
            except Exception:
                pass  # Fall back to full page render
            b64_img = base64.b64encode(img_bytes).decode()
            doc.close()

            col_prev, col_num, col_next, col_dl = st.columns([1, 2, 1, 2])
            if col_prev.button("⬅️ Prev", disabled=(page_num == 0), key="sop_prev_d"):
                st.session_state.sop_page = page_num - 1
                st.rerun()
            col_num.markdown(f"<div style='text-align:center; padding-top:8px;'>{page_num + 1} / {total_pages}</div>", unsafe_allow_html=True)
            if col_next.button("Next ➡️", disabled=(page_num >= total_pages - 1), key="sop_next_d"):
                st.session_state.sop_page = page_num + 1
                st.rerun()
            with open(SOP_PDF, "rb") as f:
                col_dl.download_button("⬇️ Download PDF", data=f.read(), file_name="Marion_St_SOP_Manual.pdf", mime="application/pdf", key="sop_dl")

            st.markdown(f'<img src="data:image/png;base64,{b64_img}" style="width:100%; max-width:900px; border-radius:6px; border:1px solid rgba(255,255,255,0.1); display:block; margin:0 auto;">', unsafe_allow_html=True)
        else:
            with open(SOP_PDF, "rb") as f:
                pdf_bytes = f.read()
            b64 = base64.b64encode(pdf_bytes).decode()
            st.markdown(f'<iframe src="data:application/pdf;base64,{b64}" width="100%" height="1600px" style="border:1px solid rgba(255,255,255,0.1);border-radius:6px;"></iframe>', unsafe_allow_html=True)


def render_tenancy_tab():
    tenants, details, summaries = load_tenancy()
    if not tenants:
        st.warning("MSP Tenancy.xlsx not found.")
        return

    expense_cfg = read_expense_config()

    # Get vacant spaces to flag in tenancy view
    vacant_keys, _vacant_meta = build_vacancy_lookup(tenants, include_auto=False)

    # Parse Yardi data and compute diffs
    yardi_data = parse_yardi_rent_rolls()
    yardi_diffs = compute_yardi_diffs(tenants, yardi_data)

    # Separate future (not-yet-commenced) tenants from current ones
    future_tenants = [t for t in tenants if t.get('Future')]
    # Summary metrics (current tenants only)
    active = [t for t in tenants if not t.get('Future')]
    # Add vacancy status and Yardi diff
    for t in active:
        key = f"{t['Building']}|{t['Space']}"
        t['Status'] = '🔴 VACANT' if key in vacant_keys else ''
        t['Yardi Diff'] = yardi_diffs.get(key, '')
    total_sf = sum(t['SF'] for t in active if t['SF'] > 1)
    total_annual = sum(t['Annual'] for t in active)
    total_monthly = sum(t['Monthly'] for t in active)
    mtm_count = sum(1 for t in active if t.get('TTE_Label') == 'MTM')

    # Calculate total CAM reimbursement across all buildings
    total_cam_reimb = 0
    for bldg_name in BUILDING_MAP.keys():
        bldg_expense = float(expense_cfg.get(bldg_name, 0) or 0)
        bldg_tenants = [t for t in active if t['Building'] == bldg_name]
        for t in bldg_tenants:
            if t.get('Is_NNN') and isinstance(t.get('CAM_Pct'), (int, float)):
                total_cam_reimb += bldg_expense * t['CAM_Pct']

    total_expenses = sum(float(expense_cfg.get(b, 0) or 0) for b in BUILDING_MAP.keys())
    total_gross_rev = total_annual + total_cam_reimb
    total_noi = total_gross_rev - total_expenses
    wavg_net_psf = total_annual / total_sf if total_sf > 0 else 0
    wavg_gross_psf = total_gross_rev / total_sf if total_sf > 0 else 0

    # --- Report buttons: per-building tenancy view mirroring the dashboard ---
    def _ten_sections():
        import pandas as pd

        def _exp_str(t):
            e = t.get('Exp Date')
            if isinstance(e, (datetime, date)):
                return e.strftime('%m/%d/%Y')
            return '-' if not e else str(e)

        def _mte(t):
            lbl = t.get('TTE_Label')
            if lbl == 'MTM':
                return 'MTM'
            m = t.get('TTE_Months')
            if m is None or (isinstance(m, float) and pd.isna(m)):
                return ''
            try:
                return str(int(m))
            except Exception:
                return str(m)

        secs = []

        def _space_sort_key(t):
            # Natural ordering: 2 before 10, while preserving labels like 102-104.
            raw = str(t.get('Space', '') or '').strip()
            parts = re.split(r'(\d+)', raw)
            return tuple(int(p) if p.isdigit() else p.lower() for p in parts)

        # One section per building, in BUILDING_MAP order; tenants by space.
        ordered_bldgs = [b for b in BUILDING_MAP if any(t['Building'] == b for t in active)]
        for b in ordered_bldgs:
            b_tenants = sorted(
                (t for t in active if t['Building'] == b),
                key=lambda t: (_space_sort_key(t), str(t.get('Building', '')).lower()),
            )
            b_expense = float(expense_cfg.get(b, 0) or 0)
            rows = []
            b_monthly = b_annual = b_gross_annual = 0.0
            b_sf = 0.0
            building_pct_total = sum(float(t.get('Building_Pct') or 0) for t in b_tenants)
            for t in b_tenants:
                sf = t.get('SF', 0) or 0
                monthly = t.get('Monthly', 0) or 0
                annual = t.get('Annual', 0) or 0
                building_pct = float(t.get('Building_Pct') or 0)
                # Allocate the full building expense using the workbook's
                # Building %; this applies to NNN and gross leases alike.
                # New CAM is already the workbook's expense allocation %.
                expense_annual = b_expense * float(t.get('CAM_Pct') or 0)
                expense_monthly = expense_annual / 12
                cam_reimb = expense_annual if t.get('Is_NNN') else 0
                gross_annual = annual + cam_reimb
                noi_annual = gross_annual - expense_annual
                net_psf = t.get('PSF', 0) or 0
                gross_psf = net_psf + ((cam_reimb / sf) if sf and sf > 0 else 0)
                b_monthly += monthly
                b_annual += annual
                b_gross_annual += gross_annual
                if sf > 1:
                    b_sf += sf
                rows.append({
                    'Space': str(t.get('Space', '')),
                    'Tenant': t.get('Tenant', ''),
                    'Type': t.get('Type', ''),
                    'SF': f"{sf:,.0f}" if isinstance(sf, (int, float)) and sf > 1 else '',
                    'Building %': f"{building_pct / building_pct_total:.1%}" if building_pct_total else '0.0%',
                    'New CAM %': f"{float(t.get('CAM_Pct') or 0):.1%}",
                    'Gross Annual': f"${gross_annual:,.0f}",
                    'Expense Annual': f"${expense_annual:,.0f}",
                    'NOI Annual': f"${noi_annual:,.0f}",
                    '': '',
                    'Monthly Rent': f"${monthly:,.0f}",
                    'Net PSF': f"${net_psf:,.2f}",
                    'Gross PSF': f"${gross_psf:,.2f}",
                    'Exp Date': _exp_str(t),
                    'MTE': _mte(t),
                })
            b_noi = b_gross_annual - b_expense
            b_net_psf = b_annual / b_sf if b_sf > 0 else 0
            b_gross_psf = b_gross_annual / b_sf if b_sf > 0 else 0
            # TOTAL row
            rows.append({
                'Space': '', 'Tenant': 'TOTAL', 'Type': '',
                'SF': f"{b_sf:,.0f}",
                'Building %': '100.0%',
                'New CAM %': f"{sum(float(t.get('CAM_Pct') or 0) for t in b_tenants):.1%}",
                'Gross Annual': f"${b_gross_annual:,.0f}",
                'Expense Annual': f"${b_expense:,.0f}",
                'NOI Annual': f"${b_noi:,.0f}",
                '': '',
                'Monthly Rent': f"${b_monthly:,.0f}",
                'Net PSF': f"${b_net_psf:,.2f}", 'Gross PSF': f"${b_gross_psf:,.2f}",
                'Exp Date': '', 'MTE': '',
            })
            code = BUILDING_MAP.get(b, {}).get('code', '')
            label = f"{b} ({code})" if code else b
            subtitle = (f"Monthly ${b_monthly:,.0f} · Annual ${b_annual:,.0f} · "
                        f"Gross Rev ${b_gross_annual:,.0f} · Expenses ${b_expense:,.0f} · "
                        f"NOI ${b_noi:,.0f}")
            secs.append((label, pd.DataFrame(rows), subtitle))
        return secs

    render_report_buttons(
        "tenancy", "Current Tenancy", _ten_sections,
        meta=f"Portfolio SF {total_sf:,.0f} · Gross Rev ${total_gross_rev:,.0f} · "
             f"Expenses ${total_expenses:,.0f} · NOI ${total_noi:,.0f} · "
             f"Wtd Avg Net ${wavg_net_psf:,.2f}/SF · Wtd Avg Gross ${wavg_gross_psf:,.2f}/SF")

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Portfolio SF", f"{total_sf:,.0f}")
    c2.metric("Gross Revenue", f"${total_gross_rev:,.0f}")
    c3.metric("Expenses", f"${total_expenses:,.0f}")
    c4.metric("NOI", f"${total_noi:,.0f}")
    c5.metric("Wtd Avg Gross $/SF", f"${wavg_gross_psf:,.2f}")
    c6.metric("Wtd Avg Net $/SF", f"${wavg_net_psf:,.2f}")

    # Filter
    buildings = ['All'] + sorted(set(t['Building'] for t in active))
    selected = st.selectbox("Filter by Building", buildings, key="tenancy_filter")

    filtered = active if selected == 'All' else [t for t in active if t['Building'] == selected]

    # Display by building
    current_bldg = None
    for t in filtered:
        if t['Building'] != current_bldg:
            current_bldg = t['Building']
            bldg_tenants = [x for x in filtered if x['Building'] == current_bldg]
            bldg_annual = sum(x['Annual'] for x in bldg_tenants)
            _bldg_header_placeholder = st.empty()

            import pandas as pd
            df = pd.DataFrame(bldg_tenants)
            building_expense = float(expense_cfg.get(current_bldg, 0) or 0)
            df['CAM_Reimb'] = df.apply(
                lambda row: (building_expense * row['CAM_Pct']) if row.get('Is_NNN') and isinstance(row.get('CAM_Pct'), (int, float)) else 0,
                axis=1
            )
            df['Gross_Annual'] = df['Annual'] + df['CAM_Reimb']
            df['CAM_PSF'] = df.apply(
                lambda row: (row['CAM_Reimb'] / row['SF']) if row['SF'] and row['SF'] > 0 else 0,
                axis=1
            )
            df['Gross_PSF'] = df['PSF'] + df['CAM_PSF']

            df['Gross_Annual'] = df['Annual'] + df['CAM_Reimb']

            display_cols = ['Space', 'Tenant', 'Type', 'SF', 'Lease', 'Monthly', 'Annual', 'Gross_Annual', 'PSF', 'Gross_PSF', 'CAM_Pct', 'CAM_Reimb', 'TTE_Months', 'Exp Date', 'Cancel Date', 'Options', 'First_Option_Exp', 'Escalation', 'Next Anniv', 'Anniv_Months', 'Next Monthly', 'Delta Monthly', 'Status', 'Yardi Diff', 'TTE_Label', 'Is_NNN']

            # Calculate building totals
            b_monthly = df['Monthly'].sum()
            b_annual = df['Annual'].sum()
            b_sf = df.loc[df['SF'] > 1, 'SF'].sum()
            b_wavg_psf = b_annual / b_sf if b_sf > 0 else 0
            b_gross_annual = df['Gross_Annual'].sum()
            b_cam_reimb = df['CAM_Reimb'].sum()
            b_noi = b_gross_annual - building_expense

            _bldg_header_placeholder.markdown(
                f'<div class="building-header"><strong>{current_bldg}</strong> — '
                f'{len(bldg_tenants)} tenants · '
                f'Gross Rev: ${b_gross_annual:,.0f} · '
                f'Expenses: ${building_expense:,.0f} · '
                f'NOI: ${b_noi:,.0f}</div>',
                unsafe_allow_html=True
            )

            # Create display dataframe
            display_df = df[display_cols].copy()
            display_df.rename(columns={
                'CAM_Pct': 'CAM %',
                'CAM_Reimb': 'CAM Reimb',
                'Gross_Annual': 'Gross Annual',
                'Gross_PSF': 'Gross PSF',
                'TTE_Months': 'MTE',
                'First_Option_Exp': '1st Opt Exp',
                'Next Anniv': 'Next Anniversary',
                'Anniv_Months': 'Anniv Δ',
                'Next Monthly': 'New Rent',
                'Delta Monthly': 'Δ Monthly',
                'TTE_Label': 'TTE Label',
                'Is_NNN': 'NNN',
            }, inplace=True)
            display_df['CAM %'] = display_df.apply(
                lambda row: f"{(row['CAM %']*100):.1f}%" if row['NNN'] and isinstance(row['CAM %'], (int, float)) and row['CAM %'] not in (None, 0)
                else '-'
                , axis=1)
            display_df['CAM Reimb'] = display_df['CAM Reimb'].apply(lambda x: f"${x:,.0f}" if x and x != 0 else '$0')
            display_df['Gross Annual'] = display_df['Gross Annual'].apply(lambda x: f"${x:,.0f}" if x else '$0')
            display_df['Gross PSF'] = display_df['Gross PSF'].apply(lambda x: f"${x:,.2f}" if x else '$0')
            display_df['MTE'] = display_df['MTE'].fillna(0).astype(int)
            display_df['Anniv Δ'] = display_df['Anniv Δ'].apply(
                lambda x: int(x) if x is not None and not (isinstance(x, float) and pd.isna(x)) else None
            )
            display_df['Escalation'] = display_df['Escalation'].apply(lambda x: f"{x:.1%}" if x and x > 0 else '-')
            display_df['Monthly'] = display_df['Monthly'].apply(lambda x: f"${x:,.0f}")
            display_df['Annual'] = display_df['Annual'].apply(lambda x: f"${x:,.0f}")
            display_df['PSF'] = display_df['PSF'].apply(lambda x: f"${x:,.2f}")
            display_df['Space'] = display_df['Space'].astype(str)
            display_df['SF'] = display_df['SF'].apply(lambda x: f"{x:,.0f}" if isinstance(x, (int, float)) and x > 1 else '')
            display_df['Next Anniversary'] = display_df['Next Anniversary'].apply(
                lambda d: d.strftime('%m/%d/%Y') if isinstance(d, (datetime, date)) else ('-' if not d else str(d))
            )
            display_df['New Rent'] = display_df['New Rent'].apply(
                lambda x: f"${x:,.0f}" if x is not None and not (isinstance(x, float) and pd.isna(x)) else '$0'
            )
            display_df['Δ Monthly'] = display_df['Δ Monthly'].apply(
                lambda x: f"{'+' if x and x > 0 else ''}${x:,.0f}" if x is not None and not (isinstance(x, float) and pd.isna(x)) else '$0'
            )

            # Pinned totals row
            totals_row = [{
                'Space': '', 'Tenant': 'TOTAL', 'Type': '', 'SF': f"{b_sf:,.0f}",
                'Lease': '', 'Monthly': f"${b_monthly:,.0f}", 'Annual': f"${b_annual:,.0f}",
                'PSF': f"${b_wavg_psf:,.2f}", 'Gross Annual': f"${b_gross_annual:,.0f}", 'Gross PSF': '',
                'CAM %': '', 'CAM Reimb': f"${b_cam_reimb:,.0f}",
                'MTE': '', 'Exp Date': '', 'Cancel Date': '', 'Options': '', '1st Opt Exp': '', 'Escalation': '', 'Next Anniversary': '', 'Anniv Δ': '',
                'New Rent': '', 'Δ Monthly': '', 'Status': '', 'Yardi Diff': '', 'TTE Label': '', 'NNN': '',
            }]

            tte_formatter = JsCode("""
                function(params) {
                    if (params.value === null || params.value === undefined) { return '-'; }
                    return params.value.toString();
                }
            """)
            mte_cell_style = JsCode("""
                function(params) {
                    if (params.value === null || params.value === undefined || params.node.rowPinned) {
                        return {'text-align': 'right'};
                    }
                    var v = parseInt(params.value);
                    if (isNaN(v)) return {'text-align': 'right'};
                    var maxMte = 120;
                    var ratio = Math.min(v / maxMte, 1.0);
                    var r, g, b;
                    if (ratio <= 0.5) {
                        var t = ratio / 0.5;
                        r = Math.round(200 - t * 100);
                        g = Math.round(50 + t * 100);
                        b = Math.round(50);
                    } else {
                        var t = (ratio - 0.5) / 0.5;
                        r = Math.round(100 - t * 70);
                        g = Math.round(150 + t * 55);
                        b = Math.round(50 + t * 20);
                    }
                    return {
                        'background-color': 'rgba(' + r + ',' + g + ',' + b + ',0.35)',
                        'color': '#e6edf3',
                        'font-weight': '600',
                        'text-align': 'right'
                    };
                }
            """)
            anniv_formatter = JsCode("""
                function(params) {
                    if (params.value === null || params.value === undefined) { return '-'; }
                    return params.value.toString();
                }
            """)
            column_configs = {
                'Space': {'type': ['textColumn']},
                'MTE': {'type': ['numericColumn'], 'valueFormatter': tte_formatter, 'cellStyle': mte_cell_style},
                'Anniv Δ': {'type': ['numericColumn'], 'valueFormatter': anniv_formatter},
                'TTE Label': {'hide': True},
                'NNN': {'hide': True},
            }

            show_grid(display_df, key=f"tenancy_{current_bldg}", pinned_bottom=totals_row,
                      column_configs=column_configs, tab_key="tenancy")

    # Portfolio totals
    st.divider()
    port_sf = sum(t['SF'] for t in filtered if t['SF'] > 1 and t['Tenant'] != 'Easement')
    port_annual = sum(t['Annual'] for t in filtered if t['Tenant'] != 'Easement')
    port_cam_reimb = 0
    for bldg_name in BUILDING_MAP.keys():
        bldg_expense = float(expense_cfg.get(bldg_name, 0) or 0)
        for t in filtered:
            if t['Building'] == bldg_name and t.get('Is_NNN') and isinstance(t.get('CAM_Pct'), (int, float)):
                port_cam_reimb += bldg_expense * t['CAM_Pct']
    port_gross_rev = port_annual + port_cam_reimb
    port_expenses = sum(float(expense_cfg.get(b, 0) or 0) for b in BUILDING_MAP.keys())
    port_noi = port_gross_rev - port_expenses
    port_wavg_net_psf = port_annual / port_sf if port_sf > 0 else 0
    port_wavg_gross_psf = port_gross_rev / port_sf if port_sf > 0 else 0
    pc1, pc2, pc3, pc4, pc5, pc6 = st.columns(6)
    pc1.metric("Total SF", f"{port_sf:,.0f}")
    pc2.metric("Gross Revenue", f"${port_gross_rev:,.0f}")
    pc3.metric("Expenses", f"${port_expenses:,.0f}")
    pc4.metric("NOI", f"${port_noi:,.0f}")
    pc5.metric("Wtd Avg Gross $/SF", f"${port_wavg_gross_psf:,.2f}")
    pc6.metric("Wtd Avg Net $/SF", f"${port_wavg_net_psf:,.2f}")

    # --- FUTURE TENANTS (leases not yet commenced) ---
    fut = future_tenants if selected == 'All' else [t for t in future_tenants if t['Building'] == selected]
    if fut:
        import pandas as pd
        st.divider()
        st.markdown(f"### 🔜 Future Tenants ({len(fut)})")
        st.caption("Executed leases that have not yet commenced. Excluded from current portfolio totals above.")
        # Sort by commencement date (soonest first)
        fut_sorted = sorted(fut, key=lambda t: t.get('Commence_Date') or date.max)
        fdf = pd.DataFrame(fut_sorted)
        fdf['Space'] = fdf['Space'].astype(str)
        fdf['SF'] = fdf['SF'].apply(lambda x: f"{x:,.0f}" if isinstance(x, (int, float)) and x > 1 else '')
        fdf['Monthly'] = fdf['Monthly'].apply(lambda x: f"${x:,.0f}")
        fdf['Annual'] = fdf['Annual'].apply(lambda x: f"${x:,.0f}")
        fdf['PSF'] = fdf['PSF'].apply(lambda x: f"${x:,.2f}")
        fut_cols = ['Building', 'Space', 'Tenant', 'Type', 'SF', 'Lease', 'Commence', 'Monthly', 'Annual', 'PSF', 'Exp Date', 'Options', 'First_Option_Exp']
        fut_cols = [c for c in fut_cols if c in fdf.columns]
        fdisplay = fdf[fut_cols].copy()
        fdisplay.rename(columns={'Commence': 'Commences', 'First_Option_Exp': '1st Opt Exp'}, inplace=True)
        show_grid(fdisplay, key="future_tenants", tab_key="tenancy")

    render_column_config_editor(
        'tenancy',
        ['Space', 'Tenant', 'Type', 'SF', 'Lease', 'Monthly', 'Annual', 'Gross Annual', 'PSF', 'Gross PSF',
         'CAM %', 'CAM Reimb', 'MTE', 'Exp Date', 'Cancel Date', 'Options', '1st Opt Exp', 'Escalation',
         'Next Anniversary', 'Anniv Δ', 'New Rent', 'Δ Monthly', 'Status', 'Yardi Diff']
    )
    render_expense_editor()


def get_vacant_spaces():
    """Get manually marked vacant spaces from Google Sheets."""
    sheet = get_gsheet()
    if not sheet:
        return []
    try:
        try:
            ws = sheet.worksheet('Vacant Spaces')
        except gspread.exceptions.WorksheetNotFound:
            ws = sheet.add_worksheet('Vacant Spaces', rows=100, cols=8)
            ws.update(values=[['Building', 'Space', 'Tenant', 'SF', 'Vacancy Date', 'Marked By', 'Notes', 'Status']], range_name='A1:H1')
        return ws.get_all_records()
    except Exception:
        return []


def build_vacancy_lookup(tenants=None, include_auto=True):
    """Return (set(keys), metadata dict) for vacant spaces keyed by Building|Space."""
    records = get_vacant_spaces()
    keys = set()
    meta = {}
    for rec in records:
        b = (rec.get('Building') or '').strip()
        space = str(rec.get('Space') or '').strip()
        if not b or not space:
            continue
        key = f"{b}|{space}"
        keys.add(key)
        meta[key] = {
            'tenant': rec.get('Tenant', ''),
            'vacancy_date': rec.get('Vacancy Date') or rec.get('Date Marked')
        }

    if include_auto and tenants:
        for t in tenants:
            # Never flag easements as vacant
            if 'easement' in (t.get('Tenant') or '').lower():
                continue
            auto_flag = (t.get('Monthly', 0) <= 0.01 and t.get('Annual', 0) <= 0.01) or t.get('Status') == '🔴 VACANT'
            if not auto_flag:
                continue
            b = (t.get('Building') or '').strip()
            space = str(t.get('Space') or '').strip()
            if not b or not space:
                continue
            key = f"{b}|{space}"
            keys.add(key)
            meta.setdefault(key, {'tenant': t.get('Tenant', '')})
    return keys, meta


def add_vacant_space(entry):
    sheet = get_gsheet()
    if not sheet:
        return False
    try:
        try:
            ws = sheet.worksheet('Vacant Spaces')
        except gspread.exceptions.WorksheetNotFound:
            ws = sheet.add_worksheet('Vacant Spaces', rows=100, cols=8)
            ws.update(values=[['Building', 'Space', 'Tenant', 'SF', 'Vacancy Date', 'Marked By', 'Notes', 'Status']], range_name='A1:H1')
        ws.append_row([
            entry['building'], entry['space'], entry['tenant'], entry['sf'],
            entry.get('vacancy_date', datetime.now().strftime('%Y-%m-%d')),
            entry.get('marked_by', ''), entry.get('notes', ''), 'Vacant'
        ], value_input_option='USER_ENTERED')
        return True
    except Exception as e:
        st.error(f"Error: {e}")
        return False


def remove_vacant_space(row_idx):
    sheet = get_gsheet()
    if not sheet:
        return False
    try:
        ws = sheet.worksheet('Vacant Spaces')
        ws.delete_rows(row_idx + 2)
        return True
    except Exception:
        return False


def _months_vacant(value):
    """Return elapsed vacancy time in months for report/display use."""
    try:
        if isinstance(value, datetime):
            vac_date = value.date()
        elif isinstance(value, date):
            vac_date = value
        else:
            vac_date = datetime.strptime(str(value)[:10], '%Y-%m-%d').date()
        days = max(0, (TODAY - vac_date).days)
        return round(days / 30.44, 1)
    except (TypeError, ValueError):
        return '—'


def render_vacancy_tab():
    tenants, _, _ = load_tenancy()
    active = [t for t in tenants]

    # Get vacant lookup (manual only)
    vacant_keys, vacant_meta = build_vacancy_lookup(active, include_auto=False)
    vacant_records = get_vacant_spaces()
    manual_keys = set(f"{v.get('Building', '')}|{v.get('Space', '')}" for v in vacant_records)

    # --- Report buttons (vacant + at-risk sections for the PDF) ---
    def _vac_sections():
        import pandas as pd
        vac_rows = []
        for v in vacant_records:
            vac_rows.append({
                'Building': v.get('Building', ''), 'Space': str(v.get('Space', '')),
                'Last Tenant': v.get('Tenant', ''), 'SF': v.get('SF', ''),
                'Vacant Since': v.get('Vacancy Date', v.get('Date Marked', '')),
                'Months Vacant': _months_vacant(v.get('Vacancy Date', v.get('Date Marked', ''))),
                'Notes': v.get('Notes', ''),
            })
        risk_rows = []
        for t in active:
            if (t.get('TTE_Label') == 'MTM' or (0 < t.get('TTE_Days', 0) <= 365)) \
                    and f"{t['Building']}|{t['Space']}" not in manual_keys:
                risk_rows.append({
                    'Building': t.get('Building', ''), 'Space': str(t.get('Space', '')),
                    'Tenant': t.get('Tenant', ''),
                    'SF': f"{t.get('SF', 0):,.0f}" if isinstance(t.get('SF'), (int, float)) else t.get('SF', ''),
                    'Monthly': f"${t.get('Monthly', 0):,.0f}",
                    'Net PSF': f"${t.get('PSF', 0):,.2f}",
                    'MTE': t.get('TTE_Months', ''), 'Exp Date': t.get('Exp Date', ''),
                })
        # Group matching leads under each currently vacant space and each space
        # expiring within three months. Each space gets its own clearly labeled
        # section with the five most recently active matching leads.
        lead_sections = []
        try:
            retail_leads, office_leads = load_lead_sheet()
            lead_frames = []
            for section_name, lead_df in (("Retail", retail_leads), ("Office", office_leads)):
                if lead_df is None or lead_df.empty:
                    continue
                work = lead_df.copy()
                work.insert(0, "Lead Type", section_name)
                activity_cols = [c for c in work.columns if "date" in str(c).lower()]
                if activity_cols:
                    parsed = [pd.to_datetime(work[c], errors="coerce") for c in activity_cols]
                    work["_Recent Date"] = pd.concat(parsed, axis=1).max(axis=1)
                else:
                    work["_Recent Date"] = pd.NaT
                lead_frames.append(work)

            all_leads = pd.concat(lead_frames, ignore_index=True) if lead_frames else pd.DataFrame()

            def _unit_tokens(value):
                return set(re.findall(r"\d+[A-Za-z]?", str(value or "").upper()))

            def _same_space(building, space, lead_building, lead_unit):
                b1 = re.sub(r"[^A-Z0-9]", "", str(building or "").upper())
                b2 = re.sub(r"[^A-Z0-9]", "", str(lead_building or "").upper())
                units = _unit_tokens(space) & _unit_tokens(lead_unit)
                return bool(b1 and b1 == b2 and units)

            targets = []
            seen_targets = set()
            for v in vacant_records:
                building = v.get('Building', '')
                space = str(v.get('Space', ''))
                key = (str(building).strip().lower(), space.strip().lower())
                if key not in seen_targets:
                    targets.append((building, space, f"VACANT · Since {v.get('Vacancy Date', v.get('Date Marked', ''))}"))
                    seen_targets.add(key)
            for t in active:
                if 0 < t.get('TTE_Days', 0) <= 90:
                    building = t.get('Building', '')
                    space = str(t.get('Space', ''))
                    key = (str(building).strip().lower(), space.strip().lower())
                    if key not in seen_targets:
                        targets.append((building, space, f"EXPIRING · {t.get('Exp Date', '')} · {t.get('TTE_Months', '')} months"))
                        seen_targets.add(key)

            lead_cols = ["Lead Type", "Tenant", "Unit", "Sqft", "Broker", "Phone", "Email",
                         "Date Contacted", "Showing", "Date Shown", "LOI", "Date LOI"]
            for building, space, status in targets:
                label = f"{building} — Space {space} · {status} · 5 Most Recent Matching Leads"
                if all_leads.empty:
                    matched = pd.DataFrame(columns=lead_cols)
                else:
                    matched = all_leads[
                        all_leads.apply(lambda r: _same_space(building, space, r.get('Building'), r.get('Unit')), axis=1)
                    ].copy()
                    matched = matched.sort_values("_Recent Date", ascending=False, na_position="last").head(5)
                    available = [c for c in lead_cols if c in matched.columns]
                    matched = matched[available].copy() if available else pd.DataFrame(columns=lead_cols)
                    for col in matched.columns:
                        if "date" in str(col).lower():
                            matched[col] = matched[col].fillna("").astype(str).str.slice(0, 19)
                if matched.empty:
                    matched = pd.DataFrame([{"Tenant": "No matching leads found"}])
                lead_sections.append((label, matched))
        except Exception:
            lead_sections = []

        return [("Currently Vacant Spaces", pd.DataFrame(vac_rows)),
                ("At Risk — Expiring Within 12 Months", pd.DataFrame(risk_rows)),
                *lead_sections]
    render_report_buttons("vacancy", "Vacancy", _vac_sections)

    # --- Mark space vacant ---
    st.markdown("### ✏️ Mark a Space Vacant")
    all_spaces = sorted(set(f"{t['Building']} #{t['Space']} — {t['Tenant']}" for t in active))

    with st.form("mark_vacant", clear_on_submit=True):
        vc1, vc2, vc3 = st.columns(3)
        selected_space = vc1.selectbox("Select Space", [""] + all_spaces)
        vacancy_date = vc2.date_input("Vacancy Date", value=date.today())
        marked_by = vc3.text_input("Your Name")
        notes = st.text_input("Notes (optional)")
        mark_btn = st.form_submit_button("🔴 Mark as Vacant", type="primary")

        if mark_btn and selected_space:
            parts = selected_space.split(' — ')
            bldg_space = parts[0]
            tenant_name = parts[1] if len(parts) > 1 else ''
            bldg_parts = bldg_space.split(' #')
            building = bldg_parts[0]
            space = bldg_parts[1] if len(bldg_parts) > 1 else ''

            sf = 0
            for t in active:
                if t['Building'] == building and str(t['Space']) == space:
                    sf = t['SF']
                    break

            success = add_vacant_space({
                'building': building, 'space': space, 'tenant': tenant_name,
                'sf': sf, 'marked_by': marked_by, 'notes': notes,
                'vacancy_date': str(vacancy_date),
            })
            if success:
                st.success(f"Marked {building} #{space} as vacant!")
                st.cache_resource.clear()
                st.rerun()

    st.divider()

    # --- Currently Vacant ---
    st.markdown("### 🔴 Currently Vacant Spaces")

    if vacant_records:
        import pandas as pd
        vacant_display = []
        for v in vacant_records:
            # Calculate days vacant
            vac_date_str = v.get('Vacancy Date', v.get('Date Marked', ''))
            months_vacant = _months_vacant(vac_date_str)

            vacant_display.append({
                'Building': v.get('Building', ''),
                'Space': str(v.get('Space', '')),
                'Last Tenant': v.get('Tenant', ''),
                'SF': v.get('SF', ''),
                'Vacant Since': vac_date_str,
                'Months Vacant': months_vacant,
                'Notes': v.get('Notes', ''),
            })

        if vacant_display:
            vdf = pd.DataFrame(vacant_display)
            vdf['Space'] = vdf['Space'].astype(str)
            show_grid(vdf, key="vacant_spaces", tab_key="vacancy")

            # Remove buttons for manual entries
            st.caption("Remove a manually marked space:")
            for i, v in enumerate(vacant_records):
                col1, col2 = st.columns([8, 1])
                col1.text(f"{v.get('Building', '')} #{v.get('Space', '')} — {v.get('Tenant', '')}")
                if col2.button("✅ Leased / Not Vacant", key=f"unvacant_{i}"):
                    remove_vacant_space(i)
                    st.cache_resource.clear()
                    st.rerun()
    else:
        st.success("✅ No currently vacant spaces — 100% occupied!")

    # At risk — exclude spaces already shown in Currently Vacant
    st.markdown("### 🟡 At Risk — Expiring Within 12 Months")
    at_risk = [t for t in active if (t.get('TTE_Label') == 'MTM' or (0 < t['TTE_Days'] <= 365))
               and f"{t['Building']}|{t['Space']}" not in manual_keys]
    if at_risk:
        import pandas as pd
        risk_df = pd.DataFrame(at_risk)[['Building', 'Space', 'Tenant', 'SF', 'Monthly', 'PSF', 'TTE_Months', 'Exp Date']]
        risk_df.rename(columns={'TTE_Months': 'MTE', 'PSF': 'Net PSF'}, inplace=True)
        risk_df['Monthly'] = risk_df['Monthly'].apply(lambda x: f"${x:,.0f}")
        risk_df['Net PSF'] = risk_df['Net PSF'].apply(lambda x: f"${x:,.2f}")
        show_grid(
            risk_df,
            key="at_risk",
            column_configs={'MTE': {'type': ['numericColumn']}},
            tab_key="vacancy"
        )
    else:
        st.success("✅ No tenants expiring within 12 months.")

    # all_spaces used by Marketing section below
    all_spaces = sorted(set(f"{t['Building']} #{t['Space']}" for t in active))

    st.divider()

    # --- MARKETING ---
    st.markdown("### 📢 Current Marketing")

    # File upload
    st.markdown("**Upload a brochure / flyer (PDF, image, etc.)**")
    uploaded_file = st.file_uploader(
        "Drag & drop or browse", type=["pdf", "png", "jpg", "jpeg", "doc", "docx"],
        key="mktg_upload", label_visibility="collapsed"
    )

    with st.form("add_marketing", clear_on_submit=True):
        mc1, mc2 = st.columns(2)
        mktg_space = mc1.selectbox("Space", [""] + all_spaces, key="mktg_space")
        mktg_desc = mc2.text_input("Description (e.g. LoopNet Listing)")
        mc3, mc4 = st.columns(2)
        mktg_url = mc3.text_input("URL (optional if uploading a file)")
        mktg_by = mc4.text_input("Your Name", key="mktg_by")
        mktg_submit = st.form_submit_button("➕ Add Marketing Entry", type="primary")

        if mktg_submit and mktg_space:
            file_url = mktg_url
            file_name = ''

            # Handle uploaded file — save to Google Drive via Sheets note, or encode
            if uploaded_file is not None:
                import base64
                file_bytes = uploaded_file.read()
                file_name = uploaded_file.name

                # Save file to shared Dropbox marketing folder
                marketing_dir = None
                for mp in [
                    "/home/node/OpenClaw/Share Jason/General Procedures/marketing",
                    os.path.join(os.path.dirname(__file__), "marketing"),
                ]:
                    try:
                        os.makedirs(mp, exist_ok=True)
                        marketing_dir = mp
                        break
                    except Exception:
                        continue

                if marketing_dir:
                    # Add timestamp to avoid conflicts
                    safe_name = f"{datetime.now().strftime('%Y%m%d_%H%M')}_{file_name}"
                    file_path = os.path.join(marketing_dir, safe_name)
                    with open(file_path, 'wb') as f:
                        f.write(file_bytes)
                    file_url = f"[File saved: {safe_name}]"
                    if not mktg_desc:
                        mktg_desc = file_name

            if file_url or file_name:
                success = add_marketing_entry({
                    'space': mktg_space,
                    'description': mktg_desc or file_name,
                    'url': file_url,
                    'added_by': mktg_by,
                })
                if success:
                    st.success("Marketing entry added!")
                    st.cache_resource.clear()
                    st.rerun()
            else:
                st.warning("Please provide a URL or upload a file.")

    marketing = get_marketing_data()
    if marketing:
        import pandas as pd
        mdf = pd.DataFrame(marketing)
        show_grid(mdf, key="marketing", tab_key="marketing")

        # Show download links for uploaded files
        marketing_dir = None
        for mp in [
            "/home/node/OpenClaw/Share Jason/General Procedures/marketing",
            os.path.join(os.path.dirname(__file__), "marketing"),
        ]:
            if os.path.isdir(mp):
                marketing_dir = mp
                break

        if marketing_dir:
            files = sorted(os.listdir(marketing_dir))
            if files:
                st.caption("📎 Uploaded files:")
                for fname in files:
                    fpath = os.path.join(marketing_dir, fname)
                    with open(fpath, 'rb') as f:
                        st.download_button(
                            f"📄 {fname}", f.read(), file_name=fname,
                            key=f"dl_{fname}", mime="application/octet-stream"
                        )
    else:
        st.info("No marketing entries yet.")
        mdf = pd.DataFrame(columns=['Space', 'Description', 'URL', 'Added By', 'Date', 'Status'])

    render_column_config_editor('vacancy', ['Building', 'Space', 'Tenant', 'SF', 'Monthly', 'Net PSF', 'MTE', 'Exp Date', 'Last Tenant', 'Vacant Since', 'Months Vacant', 'Notes'])
    render_column_config_editor('marketing', list(mdf.columns))


# --- COI EMAIL REPORT ---
EMAIL_TEAM_RECIPIENTS = [
    "asollog@gmail.com",
    "richard.b.angel@gmail.com",
    "jason.forster@proventusproperties.com",
]


# --- GENERIC TAB REPORT (PDF + Email) ---
# Reusable across every dashboard tab. `sections` is a list of
# (section_label:str, pandas.DataFrame). Produces a compact PDF capped at 2 pages
# and emails it to the team. Kept independent of the COI-specific functions.

def _smtp_config():
    """Return the [smtp] secrets dict or None."""
    try:
        return dict(st.secrets['smtp'])
    except Exception:
        return None


def _smtp_send(subject, text, html, pdf_bytes, pdf_filename):
    """Shared SMTP sender. Returns (ok, msg). Reads creds from st.secrets['smtp']."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication

    cfg = _smtp_config()
    if not cfg:
        return (False, "SMTP credentials not configured. Add an [smtp] section to "
                       "Streamlit Cloud secrets (host, port, user, password).")
    host = cfg.get('host', 'smtp.gmail.com')
    port = int(cfg.get('port', 587))
    user = cfg.get('user')
    password = cfg.get('password')
    sender = cfg.get('from', user)
    if not user or not password:
        return (False, "SMTP user/password missing from secrets.")

    msg = MIMEMultipart('mixed')
    msg['Subject'] = subject
    msg['From'] = sender
    msg['To'] = ", ".join(EMAIL_TEAM_RECIPIENTS)
    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText(text, 'plain'))
    alt.attach(MIMEText(html, 'html'))
    msg.attach(alt)
    if pdf_bytes is not None:
        part = MIMEApplication(pdf_bytes, _subtype='pdf')
        part.add_header('Content-Disposition', 'attachment', filename=pdf_filename)
        msg.attach(part)

    try:
        if port == 465:
            with smtplib.SMTP_SSL(host, port, timeout=30) as server:
                server.login(user, password)
                server.sendmail(sender, EMAIL_TEAM_RECIPIENTS, msg.as_string())
        else:
            with smtplib.SMTP(host, port, timeout=30) as server:
                server.starttls()
                server.login(user, password)
                server.sendmail(sender, EMAIL_TEAM_RECIPIENTS, msg.as_string())
    except Exception as e:
        return (False, f"SMTP send failed: {e}")
    return (True, f"Report emailed to {', '.join(EMAIL_TEAM_RECIPIENTS)}.")


def generate_tab_pdf(title, sections, meta=None, max_rows_total=150):
    """Render a compact <=2-page PDF from (label, DataFrame) sections.
    Chooses landscape automatically for wide tables. Long DataFrames are capped
    with a '...and N more' note so the PDF never overruns 2 pages.
    Returns PDF bytes.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    import io

    # Normalize sections to (label, df, subtitle) 3-tuples (subtitle optional).
    norm_sections = []
    for s in sections:
        if len(s) == 3:
            norm_sections.append((s[0], s[1], s[2]))
        else:
            norm_sections.append((s[0], s[1], None))
    sections = norm_sections

    # Decide orientation from the widest section.
    max_cols = 1
    for _lbl, df, _sub in sections:
        try:
            max_cols = max(max_cols, len(df.columns))
        except Exception:
            pass
    is_current_tenancy = title == 'Current Tenancy'
    is_vacancy_report = title == 'Vacancy'
    use_landscape = max_cols > 7
    # Vacancy uses the same landscape treatment as the current-tenancy report,
    # even when a section is narrow enough to fit portrait.
    pagesize = landscape(letter) if (is_current_tenancy or is_vacancy_report) else (landscape(letter) if use_landscape else letter)
    margin = 0.18 * inch if (is_current_tenancy or is_vacancy_report) else 0.5 * inch
    avail_w = pagesize[0] - (2 * margin)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=pagesize,
                            leftMargin=margin, rightMargin=margin,
                            topMargin=margin, bottomMargin=margin,
                            title=f"MSP {title}")
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Title'],
                                 fontSize=8 if (is_current_tenancy or is_vacancy_report) else 16,
                                 leading=9 if (is_current_tenancy or is_vacancy_report) else 18,
                                 textColor=colors.HexColor('#1a3a5c'), spaceAfter=1)
    sub = ParagraphStyle('sub', parent=styles['Heading2'],
                         fontSize=5.5 if (is_current_tenancy or is_vacancy_report) else 12,
                         leading=6.5 if (is_current_tenancy or is_vacancy_report) else 14,
                         spaceBefore=3 if (is_current_tenancy or is_vacancy_report) else 10,
                         spaceAfter=1, textColor=colors.HexColor('#1a3a5c'))
    subtitle_style = ParagraphStyle('subtitle', parent=styles['Normal'],
                                    fontSize=4.5 if (is_current_tenancy or is_vacancy_report) else 8,
                                    leading=5 if (is_current_tenancy or is_vacancy_report) else 10,
                                    spaceAfter=1,
                                    textColor=colors.HexColor('#333333'))
    small = ParagraphStyle('S', parent=styles['Normal'],
                           fontSize=5.5 if (is_current_tenancy or is_vacancy_report) else 8,
                           leading=6 if (is_current_tenancy or is_vacancy_report) else 10,
                           textColor=colors.HexColor('#555555'))
    cell = ParagraphStyle('cell', parent=styles['Normal'],
                          fontSize=4.5 if (is_current_tenancy or is_vacancy_report) else 7,
                          leading=5 if (is_current_tenancy or is_vacancy_report) else 8)
    hcell = ParagraphStyle('hcell', parent=styles['Normal'],
                           fontSize=4.5 if (is_current_tenancy or is_vacancy_report) else 7,
                           leading=5 if (is_current_tenancy or is_vacancy_report) else 8,
                           textColor=colors.white, fontName='Helvetica-Bold')
    note = ParagraphStyle('note', parent=styles['Normal'],
                          fontSize=4.5 if (is_current_tenancy or is_vacancy_report) else 7,
                          leading=5 if (is_current_tenancy or is_vacancy_report) else 8,
                          textColor=colors.HexColor('#888888'))

    story = []
    story.append(Paragraph(f"Marion Street Properties — {title}", title_style))
    story.append(Paragraph(f"Generated {datetime.now().strftime('%B %d, %Y %I:%M %p')}", small))
    if meta:
        story.append(Paragraph(meta, small))
    story.append(Spacer(1, 2 if is_current_tenancy else 6))

    # Budget rows across sections so we stay within ~2 pages.
    nonempty = [(lbl, df, sbt) for lbl, df, sbt in sections
                if df is not None and hasattr(df, 'empty') and not df.empty]
    n_sec = max(1, len(nonempty))
    per_section_cap = max(6, max_rows_total // n_sec)

    tenancy_layout = None
    if title == 'Current Tenancy' and nonempty:
        from reportlab.pdfbase.pdfmetrics import stringWidth
        shared_cols = list(nonempty[0][1].columns)
        shared_rows = [list(map(str, shared_cols))]
        for _, section_df, _ in nonempty:
            for _, row in section_df.head(per_section_cap).iterrows():
                shared_rows.append(['' if pd.isna(row[c]) else str(row[c]) for c in shared_cols])
        shared_font_size = 5.5
        shared_widths = [max(stringWidth(row[i], 'Helvetica-Bold' if r == 0 else 'Helvetica', shared_font_size)
                             for r, row in enumerate(shared_rows)) + 3 for i in range(len(shared_cols))]
        total_width = sum(shared_widths)
        if total_width > avail_w:
            shared_font_size = max(4.0, shared_font_size * avail_w / total_width)
            shared_widths = [max(stringWidth(row[i], 'Helvetica-Bold' if r == 0 else 'Helvetica', shared_font_size)
                                 for r, row in enumerate(shared_rows)) + 3 for i in range(len(shared_cols))]
            total_width = sum(shared_widths)
        if total_width > avail_w:
            shared_widths = [w * avail_w / total_width for w in shared_widths]
        elif total_width < avail_w:
            extra = (avail_w - total_width) / len(shared_widths)
            shared_widths = [w + extra for w in shared_widths]
        tenancy_layout = (shared_cols, shared_widths, shared_font_size)

    def _mk_table(df):
        cols = list(df.columns)
        ncol = len(cols)
        shown = df.head(per_section_cap)

        # Current Tenancy and Vacancy are deliberately packed: text is right-
        # justified and each column gets only the width its contents require,
        # with any remaining width distributed evenly.
        is_packed = title in ('Current Tenancy', 'Vacancy')
        if is_packed:
            from reportlab.lib.enums import TA_RIGHT
            packed_cell = ParagraphStyle(
                'packed_cell', parent=cell, fontSize=4.5,
                leading=5.5, alignment=TA_RIGHT, splitLongWords=0,
            )
            packed_hcell = ParagraphStyle(
                'packed_hcell', parent=hcell, fontSize=4.5,
                leading=5.5, alignment=TA_RIGHT, splitLongWords=0,
            )
            text_rows = [
                ['' if pd.isna(row[c]) else str(row[c]) for c in cols]
                for _, row in shown.iterrows()
            ]
            if title == 'Current Tenancy':
                layout_cols, col_widths, font_size = tenancy_layout
                packed_cell.fontSize = font_size
                packed_cell.leading = font_size + 1
                packed_hcell.fontSize = font_size
                packed_hcell.leading = font_size + 1
                header_cols = layout_cols
            else:
                # Vacancy sections have different columns, so size each grid
                # independently instead of forcing every column to equal width.
                from reportlab.pdfbase.pdfmetrics import stringWidth
                font_size = 4.5
                header_cols = cols
                raw_widths = []
                for i, col in enumerate(cols):
                    values = [str(col)] + [row[i] for row in text_rows]
                    raw_widths.append(max(stringWidth(v, 'Helvetica-Bold' if v == str(col) else 'Helvetica', font_size)
                                          for v in values) + 4)
                raw_total = sum(raw_widths)
                if raw_total > avail_w:
                    scale = avail_w / raw_total
                    col_widths = [w * scale for w in raw_widths]
                else:
                    extra = (avail_w - raw_total) / len(raw_widths) if raw_widths else 0
                    col_widths = [w + extra for w in raw_widths]

            data = [[Paragraph(str(c), packed_hcell) for c in header_cols]]
            data.extend([[Paragraph(v, packed_cell) for v in row] for row in text_rows])
            pad_left = pad_right = 1
        else:
            header = [Paragraph(str(c), hcell) for c in cols]
            data = [header]
            for _, row in shown.iterrows():
                data.append([Paragraph('' if pd.isna(row[c]) else str(row[c]), cell)
                             for c in cols])
            col_widths = [avail_w / ncol] * ncol
            pad_left = pad_right = 1

        t = Table(data, repeatRows=1, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5c')),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#f4f6f8')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), pad_left),
            ('RIGHTPADDING', (0, 0), (-1, -1), pad_right),
        ]))
        return t, len(df) - len(shown)

    if not nonempty:
        story.append(Paragraph("No data available for this section.", styles['Normal']))
    for lbl, df, sbt in nonempty:
        story.append(Paragraph(str(lbl), sub))
        if sbt:
            story.append(Paragraph(str(sbt), subtitle_style))
        tbl, remaining = _mk_table(df)
        story.append(tbl)
        if remaining > 0:
            story.append(Paragraph(f"…and {remaining} more row(s) — see dashboard.", note))

    doc.build(story)
    return buf.getvalue()


def _df_to_html(df, cap=25):
    """Small HTML table for the email body."""
    if df is None or not hasattr(df, 'empty') or df.empty:
        return '<p style="color:#2e7d32;">No data.</p>'
    cols = list(df.columns)
    head = ''.join(f'<th style="padding:4px 6px;text-align:left;background:#1a3a5c;'
                   f'color:#fff;font-size:12px;">{c}</th>' for c in cols)
    body = []
    for _, row in df.head(cap).iterrows():
        tds = ''.join(f'<td style="padding:4px 6px;border-bottom:1px solid #eee;'
                      f'font-size:12px;">{"" if pd.isna(row[c]) else row[c]}</td>'
                      for c in cols)
        body.append(f'<tr>{tds}</tr>')
    more = len(df) - min(len(df), cap)
    extra = (f'<p style="color:#888;font-size:11px;">…and {more} more — see PDF.</p>'
             if more > 0 else '')
    return (f'<table style="border-collapse:collapse;width:100%;">'
            f'<tr>{head}</tr>{"".join(body)}</table>{extra}')


def send_tab_email(title, sections, meta=None):
    """Build + send a tab report email (PDF attached) to the team."""
    try:
        pdf_bytes = generate_tab_pdf(title, sections, meta=meta)
    except Exception as e:
        return (False, f"Failed to generate PDF: {e}")

    today_str = datetime.now().strftime('%B %d, %Y')
    text_lines = ["Hi Addison, Richie, and Jason,", "",
                  f"MSP {title} report as of {today_str}.", ""]
    html_secs = []
    for s in sections:
        lbl, df = s[0], s[1]
        sbt = s[2] if len(s) == 3 else None
        n = 0 if (df is None or not hasattr(df, 'empty') or df.empty) else len(df)
        text_lines.append(f"- {lbl}: {n} row(s)" + (f" — {sbt}" if sbt else ""))
        sub_html = (f'<p style="color:#555;font-size:12px;margin:2px 0;">{sbt}</p>'
                    if sbt else '')
        html_secs.append(f'<h3 style="color:#1a3a5c;margin-bottom:4px;">{lbl}</h3>'
                         f'{sub_html}{_df_to_html(df)}')
    text_lines += ["", "Full report attached (PDF).", "", "— Sis",
                   "Marion Street Properties"]
    text = "\n".join(text_lines)
    meta_html = f'<p style="color:#555;font-size:12px;">{meta}</p>' if meta else ''
    html = (f'<html><body style="font-family:Arial,Helvetica,sans-serif;color:#222;">'
            f'<p>Hi Addison, Richie, and Jason,</p>'
            f'<p>MSP <b>{title}</b> report as of <b>{today_str}</b>.</p>'
            f'{meta_html}{"".join(html_secs)}'
            f'<p style="margin-top:16px;">Full report attached (PDF).</p>'
            f'<p>— Sis<br>Marion Street Properties</p></body></html>')

    subject = f"MSP {title} Report — {datetime.now().strftime('%m/%d/%Y')}"
    fname = f"MSP_{title.replace(' ', '_')}_{datetime.now().strftime('%Y-%m-%d')}.pdf"
    return _smtp_send(subject, text, html, pdf_bytes, fname)


def render_report_buttons(tab_key, title, sections_fn, meta=None):
    """Render an '📧 Email Team' + '⬇️ Download PDF' button pair (right-aligned).
    sections_fn is a zero-arg callable returning a list of (label, DataFrame).
    Renders inside a header row: caller passes the title; this draws the title on
    the left and the buttons on the right.
    """
    hdr_col, btn_col = st.columns([3, 1])
    with hdr_col:
        st.markdown(f"### {title}")
    with btn_col:
        if st.button("📧 Email Team", key=f"emailteam_{tab_key}",
                     use_container_width=True,
                     help="Email this report (PDF) to Addison, Richie, and Jason."):
            with st.spinner("Sending report to the team…"):
                try:
                    ok, msg = send_tab_email(title, sections_fn(), meta=meta)
                except Exception as e:
                    ok, msg = False, f"Error: {e}"
            if ok:
                st.success(f"✅ {msg}")
            else:
                st.error(f"❌ {msg}")
        try:
            _pdf = generate_tab_pdf(title, sections_fn(), meta=meta)
            btn_col.download_button(
                "⬇️ Download PDF", data=_pdf,
                file_name=f"MSP_{title.replace(' ', '_')}_{datetime.now().strftime('%Y-%m-%d')}.pdf",
                mime="application/pdf", key=f"pdfdl_{tab_key}",
                use_container_width=True)
        except Exception:
            pass


def build_coi_report_data():
    """Build a unified list of every COI line item (building, PM, tenant) across all
    properties with a normalized status and days_left, for use in the email report
    and the urgent / expiring dashboard sections.

    Returns (items, summary) where items is a list of dicts:
      {building, code, category, entity, unit, coi, expiration (date|None),
       exp_str, days_left (int|None), status}
    """
    tenants, _, summaries = load_tenancy()
    coi_data, building_coi_data, pm_coi_data = scan_coi_files()
    today_dt = datetime.now()
    vacant_keys, vacant_meta = build_vacancy_lookup(tenants, include_auto=False)

    items = []
    summary = {'total': 0, 'covered': 0, 'expired': 0, 'missing': 0, 'expiring_soon': 0}

    def _classify(exp):
        """Return (status, coi_label, days_left) for an expiration date."""
        if not exp:
            return ('No date', '✅ YES', None)
        dl = (exp - today_dt).days
        if exp < today_dt:
            return ('EXPIRED', '❌ EXP', dl)
        if dl < 90:
            return (f'{dl}d left', '✅ YES', dl)
        return ('Active', '✅ YES', dl)

    for building_name in BUILDING_MAP:
        code = BUILDING_MAP[building_name]['code']

        # Building-level certificate
        b_building_certs = building_coi_data.get(building_name, [])
        if b_building_certs:
            for cert in b_building_certs:
                exp = cert.get('exp_date')
                status, coi_label, dl = _classify(exp)
                items.append({
                    'building': building_name, 'code': code, 'category': 'Building',
                    'entity': cert.get('insured_name') or building_name, 'unit': '',
                    'coi': coi_label, 'expiration': exp,
                    'exp_str': exp.strftime('%m/%d/%Y') if exp else 'Unknown',
                    'days_left': dl, 'status': status,
                })
        else:
            items.append({
                'building': building_name, 'code': code, 'category': 'Building',
                'entity': building_name, 'unit': '', 'coi': '❌ NO',
                'expiration': None, 'exp_str': '—', 'days_left': None,
                'status': 'MISSING',
            })

        # Property Manager certificate
        b_pm_certs = pm_coi_data.get(building_name, [])
        if b_pm_certs:
            for cert in b_pm_certs:
                exp = cert.get('exp_date')
                status, coi_label, dl = _classify(exp)
                items.append({
                    'building': building_name, 'code': code, 'category': 'Property Manager',
                    'entity': cert.get('insured_name') or 'Property Manager', 'unit': '',
                    'coi': coi_label, 'expiration': exp,
                    'exp_str': exp.strftime('%m/%d/%Y') if exp else 'Unknown',
                    'days_left': dl, 'status': status,
                })
        else:
            items.append({
                'building': building_name, 'code': code, 'category': 'Property Manager',
                'entity': 'Property Manager', 'unit': '', 'coi': '❌ NO',
                'expiration': None, 'exp_str': '—', 'days_left': None,
                'status': 'MISSING',
            })

    # Tenant certificates
    for s in summaries:
        b = s.get('Buidling', '')
        name = s.get('Tenant Name', '')
        ttype = s.get('Type', '')
        unit = s.get('Unit', '')
        if not b or not name:
            continue
        if ttype and 'apartment' in str(ttype).lower():
            continue
        key = f"{b}|{str(unit).strip()}"
        if key in vacant_keys:
            continue  # vacant units don't need a COI

        summary['total'] += 1
        code = BUILDING_MAP.get(b, {}).get('code', '')
        b_certs = coi_data.get(b, [])
        matches = []
        for cert in b_certs:
            if fuzzy_match_tenant(name, cert.get('insured_name')):
                matches.append(cert)
        if not matches:
            for cert in b_certs:
                if fuzzy_match_tenant(name, cert['filename']):
                    matches.append(cert)
        # When multiple certs match one tenant (e.g. an old + a renewed cert, or a
        # rebrand alias), pick the one with the LATEST expiration date so stale
        # certs never mask a current one.
        matched = None
        if matches:
            matched = max(matches,
                          key=lambda c: c.get('exp_date') or datetime.min)

        if matched:
            exp = matched['exp_date']
            status, coi_label, dl = _classify(exp)
            if status == 'EXPIRED':
                summary['expired'] += 1
            else:
                summary['covered'] += 1
                if dl is not None and dl < 90:
                    summary['expiring_soon'] += 1
            items.append({
                'building': b, 'code': code, 'category': 'Tenant',
                'entity': name, 'unit': str(unit), 'coi': coi_label,
                'expiration': exp,
                'exp_str': exp.strftime('%m/%d/%Y') if exp else 'Unknown',
                'days_left': dl, 'status': status,
            })
        else:
            summary['missing'] += 1
            items.append({
                'building': b, 'code': code, 'category': 'Tenant',
                'entity': name, 'unit': str(unit), 'coi': '❌ NO',
                'expiration': None, 'exp_str': '—', 'days_left': None,
                'status': 'MISSING',
            })

    return items, summary


def split_urgent_expiring(items):
    """Split items into (urgent, expiring). Urgent = expired OR missing.
    Expiring = active certs with days_left < 90, sorted soonest-first."""
    urgent = [it for it in items if it['status'] == 'EXPIRED' or it['status'] == 'MISSING']
    # urgent: expired (by most overdue) then missing
    urgent.sort(key=lambda it: (it['status'] != 'EXPIRED',
                                it['days_left'] if it['days_left'] is not None else 0))
    expiring = [it for it in items
                if it['status'] not in ('EXPIRED', 'MISSING', 'No date')
                and it['days_left'] is not None and 0 <= it['days_left'] < 90]
    expiring.sort(key=lambda it: it['days_left'])
    return urgent, expiring


def generate_coi_pdf(items, summary):
    """Generate a PDF report of the COI status. Returns bytes."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                    Spacer)
    import io

    urgent, expiring = split_urgent_expiring(items)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=0.6 * inch, rightMargin=0.6 * inch,
                            topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Title'], fontSize=18,
                                 textColor=colors.HexColor('#1a3a5c'))
    h2 = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=13,
                        spaceBefore=14, spaceAfter=6)
    small = ParagraphStyle('S', parent=styles['Normal'], fontSize=9,
                           textColor=colors.HexColor('#555555'))
    story = []
    today_str = datetime.now().strftime('%B %d, %Y')
    story.append(Paragraph("Marion Street Properties", title_style))
    story.append(Paragraph("Certificate of Insurance Status Report", h2))
    story.append(Paragraph(f"Generated {today_str}", small))
    story.append(Spacer(1, 10))

    cov = (summary['covered'] / summary['total'] * 100) if summary['total'] else 0
    story.append(Paragraph(
        f"<b>Coverage:</b> {cov:.0f}%  |  <b>Active:</b> {summary['covered'] - summary['expiring_soon']}  |  "
        f"<b>Expiring Soon:</b> {summary['expiring_soon']}  |  <b>Expired:</b> {summary['expired']}  |  "
        f"<b>Missing:</b> {summary['missing']}", styles['Normal']))
    story.append(Spacer(1, 6))

    cell_style = ParagraphStyle('cell', parent=styles['Normal'], fontSize=8, leading=9)
    hdr_cell = ParagraphStyle('hcell', parent=styles['Normal'], fontSize=8,
                              leading=9, textColor=colors.white,
                              fontName='Helvetica-Bold')

    def _table(rows, header_color):
        header = [Paragraph(h, hdr_cell) for h in
                  ['Property', 'Type', 'Entity', 'Unit', 'Expiration', 'Days', 'Status']]
        data = [header]
        for it in rows:
            days = '' if it['days_left'] is None else str(it['days_left'])
            data.append([it['code'], it['category'],
                         Paragraph(it['entity'], cell_style),
                         it['unit'], it['exp_str'], days, it['status']])
        t = Table(data, repeatRows=1,
                  colWidths=[0.7*inch, 0.95*inch, 2.25*inch, 0.55*inch, 0.8*inch, 0.55*inch, 0.85*inch])
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f6f8')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]
        for i, it in enumerate(rows, start=1):
            if it['status'] == 'EXPIRED':
                style.append(('TEXTCOLOR', (6, i), (6, i), colors.HexColor('#c0392b')))
                style.append(('FONTNAME', (6, i), (6, i), 'Helvetica-Bold'))
            elif it['status'] == 'MISSING':
                style.append(('TEXTCOLOR', (6, i), (6, i), colors.HexColor('#b9770e')))
                style.append(('FONTNAME', (6, i), (6, i), 'Helvetica-Bold'))
        t.setStyle(TableStyle(style))
        return t

    urgent_style = ParagraphStyle('HU', parent=h2, textColor=colors.HexColor('#c0392b'))
    exp_style = ParagraphStyle('HE', parent=h2, textColor=colors.HexColor('#b9770e'))
    story.append(Paragraph("URGENT — Expired &amp; Missing Certificates", urgent_style))
    if urgent:
        story.append(_table(urgent, colors.HexColor('#c0392b')))
    else:
        story.append(Paragraph("None — all certificates present and current.", styles['Normal']))

    story.append(Paragraph("Expiring in the Next 3 Months", exp_style))
    if expiring:
        story.append(_table(expiring, colors.HexColor('#b9770e')))
    else:
        story.append(Paragraph("None expiring in the next 90 days.", styles['Normal']))

    doc.build(story)
    return buf.getvalue()


def build_coi_email_text(items, summary):
    """Build the plain-text + HTML email body for the COI report."""
    urgent, expiring = split_urgent_expiring(items)
    today_str = datetime.now().strftime('%B %d, %Y')
    cov = (summary['covered'] / summary['total'] * 100) if summary['total'] else 0

    lines = []
    lines.append("Hi Addison, Richie, and Jason,")
    lines.append("")
    lines.append(f"Certificate of Insurance status as of {today_str}:")
    lines.append("")
    lines.append(f"Coverage: {cov:.0f}%  |  Active: {summary['covered'] - summary['expiring_soon']}  |  "
                 f"Expiring Soon: {summary['expiring_soon']}  |  Expired: {summary['expired']}  |  "
                 f"Missing: {summary['missing']}")
    lines.append("")
    lines.append("=== URGENT — EXPIRED & MISSING ===")
    if urgent:
        for it in urgent:
            if it['status'] == 'EXPIRED':
                od = abs(it['days_left']) if it['days_left'] is not None else '?'
                lines.append(f"  [EXPIRED {od}d ago] {it['code']} — {it['category']}: "
                             f"{it['entity']}{(' (Unit ' + it['unit'] + ')') if it['unit'] else ''} "
                             f"— exp. {it['exp_str']}")
            else:
                lines.append(f"  [MISSING] {it['code']} — {it['category']}: "
                             f"{it['entity']}{(' (Unit ' + it['unit'] + ')') if it['unit'] else ''}")
    else:
        lines.append("  None — all certificates present and current.")
    lines.append("")
    lines.append("=== EXPIRING IN NEXT 3 MONTHS (soonest first) ===")
    if expiring:
        for it in expiring:
            lines.append(f"  [{it['days_left']}d left] {it['code']} — {it['category']}: "
                         f"{it['entity']}{(' (Unit ' + it['unit'] + ')') if it['unit'] else ''} "
                         f"— exp. {it['exp_str']}")
    else:
        lines.append("  None expiring in the next 90 days.")
    lines.append("")
    lines.append("Full status report attached (PDF).")
    lines.append("")
    lines.append("— Sis")
    lines.append("Marion Street Properties")
    text = "\n".join(lines)

    def _rows_html(rows, kind):
        if not rows:
            note = ("None — all certificates present and current." if kind == 'urgent'
                    else "None expiring in the next 90 days.")
            return f'<tr><td colspan="5" style="padding:6px;color:#2e7d32;">{note}</td></tr>'
        out = []
        for it in rows:
            if it['status'] == 'EXPIRED':
                badge = f'<span style="color:#c0392b;font-weight:bold;">EXPIRED</span>'
            elif it['status'] == 'MISSING':
                badge = f'<span style="color:#b9770e;font-weight:bold;">MISSING</span>'
            else:
                badge = f'<span style="color:#b9770e;">{it["days_left"]}d left</span>'
            unit = f" (Unit {it['unit']})" if it['unit'] else ''
            out.append(
                f'<tr>'
                f'<td style="padding:5px 8px;border-bottom:1px solid #eee;">{it["code"]}</td>'
                f'<td style="padding:5px 8px;border-bottom:1px solid #eee;">{it["category"]}</td>'
                f'<td style="padding:5px 8px;border-bottom:1px solid #eee;">{it["entity"]}{unit}</td>'
                f'<td style="padding:5px 8px;border-bottom:1px solid #eee;">{it["exp_str"]}</td>'
                f'<td style="padding:5px 8px;border-bottom:1px solid #eee;">{badge}</td>'
                f'</tr>')
        return "".join(out)

    html = f"""\
<html><body style="font-family:Arial,Helvetica,sans-serif;color:#222;font-size:14px;">
<p>Hi Addison, Richie, and Jason,</p>
<p>Certificate of Insurance status as of <b>{today_str}</b>:</p>
<p style="background:#f4f6f8;padding:10px;border-radius:6px;">
<b>Coverage:</b> {cov:.0f}% &nbsp;|&nbsp; <b>Active:</b> {summary['covered'] - summary['expiring_soon']} &nbsp;|&nbsp;
<b>Expiring Soon:</b> {summary['expiring_soon']} &nbsp;|&nbsp; <b>Expired:</b> {summary['expired']} &nbsp;|&nbsp;
<b>Missing:</b> {summary['missing']}</p>
<h3 style="color:#c0392b;margin-bottom:4px;">⚠️ Urgent — Expired &amp; Missing</h3>
<table style="border-collapse:collapse;width:100%;font-size:13px;">
<tr style="background:#c0392b;color:#fff;">
<th style="padding:6px 8px;text-align:left;">Property</th><th style="padding:6px 8px;text-align:left;">Type</th>
<th style="padding:6px 8px;text-align:left;">Entity</th><th style="padding:6px 8px;text-align:left;">Expiration</th>
<th style="padding:6px 8px;text-align:left;">Status</th></tr>
{_rows_html(urgent, 'urgent')}
</table>
<h3 style="color:#b9770e;margin-bottom:4px;margin-top:18px;">🗓️ Expiring in Next 3 Months</h3>
<table style="border-collapse:collapse;width:100%;font-size:13px;">
<tr style="background:#b9770e;color:#fff;">
<th style="padding:6px 8px;text-align:left;">Property</th><th style="padding:6px 8px;text-align:left;">Type</th>
<th style="padding:6px 8px;text-align:left;">Entity</th><th style="padding:6px 8px;text-align:left;">Expiration</th>
<th style="padding:6px 8px;text-align:left;">Status</th></tr>
{_rows_html(expiring, 'expiring')}
</table>
<p style="margin-top:18px;">Full status report attached (PDF).</p>
<p>— Sis<br>Marion Street Properties</p>
</body></html>"""
    return text, html


def send_coi_email(items, summary):
    """Send the COI report to the team. Returns (ok, message).
    SMTP credentials are read from st.secrets['smtp'].
    Expected secrets:
      [smtp]
      host = "smtp.gmail.com"
      port = 587
      user = "Sis.MarionStreet@gmail.com"
      password = "<app password>"
      from = "Sis.MarionStreet@gmail.com"   # optional, defaults to user
    """
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication

    try:
        smtp_cfg = dict(st.secrets['smtp'])
    except Exception:
        return (False, "SMTP credentials not configured. Add an [smtp] section to "
                       "Streamlit Cloud secrets (host, port, user, password).")

    host = smtp_cfg.get('host', 'smtp.gmail.com')
    port = int(smtp_cfg.get('port', 587))
    user = smtp_cfg.get('user')
    password = smtp_cfg.get('password')
    sender = smtp_cfg.get('from', user)
    if not user or not password:
        return (False, "SMTP user/password missing from secrets.")

    text, html = build_coi_email_text(items, summary)
    subject = f"MSP Certificate of Insurance Report — {datetime.now().strftime('%m/%d/%Y')}"

    msg = MIMEMultipart('mixed')
    msg['Subject'] = subject
    msg['From'] = sender
    msg['To'] = ", ".join(EMAIL_TEAM_RECIPIENTS)
    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText(text, 'plain'))
    alt.attach(MIMEText(html, 'html'))
    msg.attach(alt)

    try:
        pdf_bytes = generate_coi_pdf(items, summary)
        pdf_part = MIMEApplication(pdf_bytes, _subtype='pdf')
        pdf_part.add_header('Content-Disposition', 'attachment',
                            filename=f"MSP_COI_Report_{datetime.now().strftime('%Y-%m-%d')}.pdf")
        msg.attach(pdf_part)
    except Exception as e:
        return (False, f"Failed to generate PDF: {e}")

    try:
        if port == 465:
            with smtplib.SMTP_SSL(host, port, timeout=30) as server:
                server.login(user, password)
                server.sendmail(sender, EMAIL_TEAM_RECIPIENTS, msg.as_string())
        else:
            with smtplib.SMTP(host, port, timeout=30) as server:
                server.starttls()
                server.login(user, password)
                server.sendmail(sender, EMAIL_TEAM_RECIPIENTS, msg.as_string())
    except Exception as e:
        return (False, f"SMTP send failed: {e}")

    return (True, f"Report emailed to {', '.join(EMAIL_TEAM_RECIPIENTS)}.")


def render_insurance_tab():
    tenants, _, summaries = load_tenancy()
    coi_data, building_coi_data, pm_coi_data = scan_coi_files()
    today_dt = datetime.now()
    vacant_keys, vacant_meta = build_vacancy_lookup(tenants, include_auto=False)

    # --- Email Team button + Urgent / Expiring sections (top of tab) ---
    report_items, report_summary = build_coi_report_data()
    urgent_items, expiring_items = split_urgent_expiring(report_items)

    hdr_col, btn_col = st.columns([3, 1])
    with hdr_col:
        st.markdown("### 🛡️ Certificate of Insurance Reconciliation")
    with btn_col:
        if st.button("📧 Email Team", key="coi_email_team", use_container_width=True,
                     help="Email the COI report (PDF + summary) to Addison, Richie, and Jason."):
            with st.spinner("Sending report to the team…"):
                ok, msg = send_coi_email(report_items, report_summary)
            if ok:
                st.success(f"✅ {msg}")
            else:
                st.error(f"❌ {msg}")

    # Downloadable PDF (always available, even if email isn't configured)
    try:
        _pdf = generate_coi_pdf(report_items, report_summary)
        btn_col.download_button(
            "⬇️ Download PDF", data=_pdf,
            file_name=f"MSP_COI_Report_{datetime.now().strftime('%Y-%m-%d')}.pdf",
            mime="application/pdf", key="coi_pdf_dl", use_container_width=True)
    except Exception:
        pass

    # (Urgent / Expiring summary sections are rendered at the BOTTOM of this tab.)

    # Build insurance data
    ins_rows = []
    summary = {'total': 0, 'covered': 0, 'expired': 0, 'missing': 0, 'expiring_soon': 0}

    for s in summaries:
        b = s.get('Buidling', '')
        name = s.get('Tenant Name', '')
        ttype = s.get('Type', '')
        unit = s.get('Unit', '')
        if not b or not name:
            continue
        if ttype and 'apartment' in str(ttype).lower():
            continue

        key = f"{b}|{str(unit).strip()}"
        if key in vacant_keys:
            last_name = vacant_meta.get(key, {}).get('tenant') or name
            ins_rows.append({
                'Building': b,
                'Tenant': f"VACANT — {last_name}",
                'Unit': str(unit),
                'Type': ttype,
                'COI': 'N/A',
                'Expiration': '—',
                'Days Left': '—',
                'Status': 'Vacant',
            })
            continue

        summary['total'] += 1
        b_certs = coi_data.get(b, [])
        matches = []
        for cert in b_certs:
            if fuzzy_match_tenant(name, cert.get('insured_name')):
                matches.append(cert)
        if not matches:
            for cert in b_certs:
                if fuzzy_match_tenant(name, cert['filename']):
                    matches.append(cert)
        # Latest-expiration cert wins so stale certs never mask a current one.
        matched = None
        if matches:
            matched = max(matches,
                          key=lambda c: c.get('exp_date') or datetime.min)

        if matched:
            exp = matched['exp_date']
            if exp:
                days_left = (exp - today_dt).days
                if exp < today_dt:
                    status = 'EXPIRED'
                    summary['expired'] += 1
                    coi_label = '❌ EXP'
                elif days_left < 60:
                    status = f'{days_left}d left'
                    summary['expiring_soon'] += 1
                    summary['covered'] += 1
                    coi_label = '✅ YES'
                else:
                    status = 'Active'
                    summary['covered'] += 1
                    coi_label = '✅ YES'
                exp_str = exp.strftime('%m/%d/%Y')
            else:
                exp_str = 'Unknown'
                days_left = None
                status = 'No date'
                summary['covered'] += 1
                coi_label = '✅ YES'
            ins_rows.append({
                'Building': b, 'Tenant': name, 'Unit': str(unit), 'Type': ttype,
                'COI': coi_label, 'Expiration': exp_str,
                'Days Left': str(days_left) if days_left is not None else '—',
                'Status': status, 'File': matched['filename'],
                'View': build_coi_url(b, matched['filename']),
            })
        else:
            summary['missing'] += 1
            ins_rows.append({
                'Building': b, 'Tenant': name, 'Unit': str(unit), 'Type': ttype,
                'COI': '❌ NO', 'Expiration': '—', 'Days Left': '—',
                'Status': 'MISSING', 'File': '', 'View': '',
            })

    # Summary metrics
    coverage_pct = (summary['covered'] / summary['total'] * 100) if summary['total'] else 0
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Coverage", f"{coverage_pct:.0f}%")
    c2.metric("Active", summary['covered'] - summary['expiring_soon'])
    c3.metric("Expiring Soon", summary['expiring_soon'])
    c4.metric("Expired", summary['expired'])
    c5.metric("Missing", summary['missing'])

    # Display by building
    import pandas as pd
    for building_name in BUILDING_MAP:
        b_rows = [r for r in ins_rows if r['Building'] == building_name]
        if not b_rows:
            continue
        b_covered = sum(1 for r in b_rows if '✅' in r['COI'])
        n_certs = len(coi_data.get(building_name, []))
        code = BUILDING_MAP[building_name]['code']

        st.markdown(f'<div class="building-header"><strong>▎ {building_name} ({code})</strong> — {b_covered}/{len(b_rows)} covered · {n_certs} files on disk</div>', unsafe_allow_html=True)

        display_cols = ['Tenant', 'Unit', 'Type', 'COI', 'Expiration', 'Days Left', 'Status', 'View']

        # --- Building-level certificate section (shown BEFORE tenants) ---
        b_building_certs = building_coi_data.get(building_name, [])
        bldg_rows = []
        if b_building_certs:
            for cert in b_building_certs:
                exp = cert.get('exp_date')
                if exp:
                    days_left = (exp - today_dt).days
                    if exp < today_dt:
                        b_status = 'EXPIRED'
                        b_coi = '❌ EXP'
                    elif days_left < 60:
                        b_status = f'{days_left}d left'
                        b_coi = '✅ YES'
                    else:
                        b_status = 'Active'
                        b_coi = '✅ YES'
                    b_exp_str = exp.strftime('%m/%d/%Y')
                    b_days = str(days_left)
                else:
                    b_exp_str = 'Unknown'
                    b_days = '—'
                    b_status = 'No date'
                    b_coi = '✅ YES'
                bldg_rows.append({
                    'Tenant': cert.get('insured_name') or building_name,
                    'Unit': '', 'Type': 'Building',
                    'COI': b_coi, 'Expiration': b_exp_str,
                    'Days Left': b_days, 'Status': b_status,
                    'View': build_coi_url(building_name, cert.get('filename')),
                })
        else:
            bldg_rows.append({
                'Tenant': building_name, 'Unit': '', 'Type': 'Building',
                'COI': '❌ NO', 'Expiration': '—',
                'Days Left': '—', 'Status': 'MISSING', 'View': '',
            })
        st.markdown('<div style="color:#f0883e;font-size:0.8rem;font-weight:600;margin:4px 0 2px 8px;">🏢 BUILDING CERTIFICATE</div>', unsafe_allow_html=True)
        bldg_df = pd.DataFrame(bldg_rows)
        show_grid(bldg_df[display_cols], key=f"ins_bldg_{building_name}", tab_key="insurance")

        # --- Property Manager certificate section ---
        b_pm_certs = pm_coi_data.get(building_name, [])
        pm_rows = []
        if b_pm_certs:
            for cert in b_pm_certs:
                exp = cert.get('exp_date')
                if exp:
                    days_left = (exp - today_dt).days
                    if exp < today_dt:
                        pm_status = 'EXPIRED'
                        pm_coi = '❌ EXP'
                    elif days_left < 60:
                        pm_status = f'{days_left}d left'
                        pm_coi = '✅ YES'
                    else:
                        pm_status = 'Active'
                        pm_coi = '✅ YES'
                    pm_exp_str = exp.strftime('%m/%d/%Y')
                    pm_days = str(days_left)
                else:
                    pm_exp_str = 'Unknown'
                    pm_days = '—'
                    pm_status = 'No date'
                    pm_coi = '✅ YES'
                pm_rows.append({
                    'Tenant': cert.get('insured_name') or 'Property Manager',
                    'Unit': '', 'Type': 'PM',
                    'COI': pm_coi, 'Expiration': pm_exp_str,
                    'Days Left': pm_days, 'Status': pm_status,
                    'View': build_coi_url(building_name, cert.get('filename')),
                })
        else:
            pm_rows.append({
                'Tenant': 'Property Manager', 'Unit': '', 'Type': 'PM',
                'COI': '❌ NO', 'Expiration': '—',
                'Days Left': '—', 'Status': 'MISSING', 'View': '',
            })
        st.markdown('<div style="color:#9b59b6;font-size:0.8rem;font-weight:600;margin:8px 0 2px 8px;">👤 PROPERTY MANAGER CERTIFICATE</div>', unsafe_allow_html=True)
        pm_df = pd.DataFrame(pm_rows)
        show_grid(pm_df[display_cols], key=f"ins_pm_{building_name}", tab_key="insurance")

        # --- Tenant certificates ---
        st.markdown('<div style="color:#8b949e;font-size:0.8rem;font-weight:600;margin:8px 0 2px 8px;">👥 TENANTS</div>', unsafe_allow_html=True)
        df = pd.DataFrame(b_rows)
        show_grid(df[display_cols], key=f"ins_{building_name}", tab_key="insurance")

    render_column_config_editor('insurance', ['Tenant', 'Unit', 'Type', 'COI', 'Expiration', 'Days Left', 'Status', 'View'])

    # --- Summary sections (bottom of tab): Urgent + Expiring in 3 months ---
    st.divider()
    st.markdown("### 📋 COI Summary")
    _uc = len(urgent_items)
    _ec = len(expiring_items)
    with st.expander(f"⚠️ URGENT — Expired & Missing ({_uc})", expanded=_uc > 0):
        if urgent_items:
            urg_df = pd.DataFrame([{
                'Property': it['code'], 'Type': it['category'],
                'Entity': it['entity'], 'Unit': it['unit'],
                'Expiration': it['exp_str'],
                'Overdue/Days': (f"{abs(it['days_left'])}d ago" if it['status'] == 'EXPIRED'
                                 and it['days_left'] is not None else ''),
                'Status': it['status'],
            } for it in urgent_items])
            st.dataframe(urg_df, hide_index=True, use_container_width=True)
        else:
            st.success("All certificates present and current.")

    with st.expander(f"🗓️ Expiring in Next 3 Months ({_ec})", expanded=_ec > 0):
        if expiring_items:
            exp_df = pd.DataFrame([{
                'Property': it['code'], 'Type': it['category'],
                'Entity': it['entity'], 'Unit': it['unit'],
                'Expiration': it['exp_str'], 'Days Left': it['days_left'],
                'Status': it['status'],
            } for it in expiring_items])
            st.dataframe(exp_df, hide_index=True, use_container_width=True)
        else:
            st.info("Nothing expiring in the next 90 days.")


def render_deposits_tab():
    tenants, details, summaries = load_tenancy()
    active = [t for t in tenants if t['Tenant'] not in ('Easement',)]
    today_dt = datetime.now()
    vacant_keys, vacant_meta = build_vacancy_lookup(tenants, include_auto=False)

    # (Report buttons render after dep_data is built, just before the tables.)

    # Build detailed SD data from term rows
    import pandas as pd
    dep_data = []

    for tid, rows in details.items():
        first = rows[0]
        tenant_name = first.get('Tenant', '')
        building = first.get('Building', '')
        if tenant_name in ('Easement', '-') or not building:
            continue

        # Find current SD using tightest-fit logic (smallest end date where start <= today <= end)
        current_sd = first.get('Sec Dep', 0) or 0
        next_sd = None
        next_sd_date = None
        current_year_row = None
        best_end = None

        rows_with_dates = []
        for r in rows:
            start = r.get('Start Date')
            end = r.get('End Date')
            sd = r.get('Sec Dep', 0) or 0
            if isinstance(start, (datetime, date)) and isinstance(end, (datetime, date)):
                start_d = start.date() if isinstance(start, datetime) else start
                end_d = end.date() if isinstance(end, datetime) else end
                rows_with_dates.append((start_d, end_d, sd, r))
                if start_d <= TODAY <= end_d:
                    if best_end is None or end_d < best_end:
                        best_end = end_d
                        current_sd = sd
                        current_year_row = r

        # Next SD: find the row with the smallest end date AFTER current period's end
        if current_year_row and best_end:
            next_sd_date = best_end  # SD Anniversary = end of current period
            for start_d, end_d, sd, r in sorted(rows_with_dates, key=lambda x: x[1]):
                if end_d > best_end:
                    next_sd = sd
                    break

        # If we didn't find current, use first row's SD
        if current_year_row is None:
            current_sd = first.get('Sec Dep', 0) or 0

        # Get tenant type from summary
        tenant_type = ''
        for s in summaries:
            if s.get('Tenant Name') == tenant_name and s.get('Buidling') == building:
                tenant_type = s.get('Type', '') or ''
                break

        # Find TTE from active tenants list
        tte_str = 'N/A'
        lease_type = first.get('Type', '')
        for t in active:
            if t['Tenant'] == tenant_name and t['Building'] == building:
                tte_str = t['TTE']
                lease_type = t['Lease']
                break

        # SD Anniversary = end of current period (when next term starts), shown for all tenants with future periods
        sd_anniv_date = next_sd_date if next_sd_date and next_sd_date > TODAY else None
        sd_new_amount = next_sd if sd_anniv_date and next_sd is not None else 0
        sd_delta = (sd_new_amount - current_sd) if sd_anniv_date else 0

        space_val = str(first.get('Space', '')).strip()
        key = f"{building}|{space_val}"
        is_vacant = key in vacant_keys
        display_tenant = f"VACANT — {vacant_meta.get(key, {}).get('tenant') or tenant_name}" if is_vacant else tenant_name

        dep_data.append({
            'Building': building,
            'Tenant': tenant_name,
            'Display Tenant': display_tenant,
            'Space': space_val,
            'Type': tenant_type,
            'Lease': lease_type,
            'Current SD': current_sd if isinstance(current_sd, (int, float)) else 0,
            'Current SD Fmt': f"${float(current_sd):,.2f}" if isinstance(current_sd, (int, float)) and current_sd > 0 else '❌ $0.00',
            'SD Anniversary': sd_anniv_date,
            'New SD Amount': sd_new_amount,
            'SD Delta': sd_delta,
            'TTE': tte_str,
        })

    dep_data.sort(key=lambda d: (d['Building'], d['Tenant']))

    # Yardi SD comparison — use Security Deposit Activity page (Deposits On Hand)
    # Use the reconcile name mapping to bridge Yardi spaces to spreadsheet spaces
    import json as _json_dep
    yardi_deposits = parse_yardi_deposit_activity()
    yardi_rent_data = parse_yardi_rent_rolls()

    # Load name mapping from reconcile tab (Yardi key -> spreadsheet tenant name)
    name_map_dep = {}
    try:
        raw = _read_gsheet_config("Config: Yardi Names")
        if raw:
            name_map_dep = _json_dep.loads(raw)
    except Exception:
        pass

    # Build reverse lookup: "Building|SheetSpace" -> yardi deposit amount
    # Strategy: for each Yardi deposit entry, find which spreadsheet tenant it maps to
    sheet_sd_from_yardi = {}  # "Building|SheetSpace" -> deposit amount

    # First, build sheet tenant->space index
    sheet_tenant_to_space = {}
    for d in dep_data:
        sheet_tenant_to_space[f"{d['Building']}|{d['Tenant']}"] = d['Space']

    for yardi_key, deposit_amt in yardi_deposits.items():
        building = yardi_key.split('|', 1)[0] if '|' in yardi_key else ''
        yardi_space = yardi_key.split('|', 1)[1] if '|' in yardi_key else ''

        # Method 1: Check name mapping (reconcile lookup table)
        matched_tenant = name_map_dep.get(yardi_key, '')

        # Method 2: Auto-match via normalized space in rent roll
        if not matched_tenant and yardi_key in yardi_rent_data:
            raw_unit = yardi_rent_data[yardi_key].get('raw_unit', yardi_space)
            norm = normalize_space(raw_unit, building)
            # Find a dep_data entry with this building+space
            for d in dep_data:
                if d['Building'] == building and d['Space'] == norm:
                    matched_tenant = d['Tenant']
                    break

        # Method 3: Direct space match
        if not matched_tenant:
            for d in dep_data:
                if d['Building'] == building and d['Space'] == yardi_space:
                    matched_tenant = d['Tenant']
                    break

        if matched_tenant:
            sheet_space = sheet_tenant_to_space.get(f"{building}|{matched_tenant}")
            if sheet_space:
                sheet_sd_from_yardi[f"{building}|{sheet_space}"] = deposit_amt

    for d in dep_data:
        key = f"{d['Building']}|{d['Space']}"
        yardi_sd = sheet_sd_from_yardi.get(key)
        d['Yardi SD'] = yardi_sd
        dash_sd = d['Current SD']
        if yardi_sd is None:
            d['SD Diff'] = 'Not in Yardi'
        elif abs(dash_sd - yardi_sd) > 0.01:
            d['SD Diff'] = f"${dash_sd:,.0f} vs ${yardi_sd:,.0f}"
        else:
            d['SD Diff'] = ''

    # Summary metrics
    total_deposits = sum(d['Current SD'] for d in dep_data)
    with_dep = sum(1 for d in dep_data if d['Current SD'] > 0)
    without_dep = sum(1 for d in dep_data if d['Current SD'] == 0)
    upcoming_changes = sum(1 for d in dep_data if d['SD Anniversary'])

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Deposits Held", f"${total_deposits:,.2f}")
    c2.metric("Tenants w/ Deposit", str(with_dep))
    c3.metric("Tenants w/o Deposit", str(without_dep))
    c4.metric("Upcoming SD Changes", str(upcoming_changes))

    # --- Report buttons (condensed portfolio SD view for the PDF) ---
    def _dep_sections():
        if not dep_data:
            return [("Security Deposits", pd.DataFrame())]
        rows_out = []
        for d in dep_data:
            rows_out.append({
                'Bldg': BUILDING_MAP.get(d['Building'], {}).get('code', d['Building']),
                'Tenant': d.get('Display Tenant', ''),
                'Space': d.get('Space', ''),
                'Current SD': d.get('Current SD Fmt', ''),
                'SD Anniv': (d['SD Anniversary'].strftime('%m/%d/%Y')
                             if isinstance(d.get('SD Anniversary'), (datetime, date))
                             else (d.get('SD Anniversary') or '-')),
                'Yardi SD': (f"${d['Yardi SD']:,.0f}"
                             if d.get('Yardi SD') is not None
                             and not (isinstance(d.get('Yardi SD'), float) and pd.isna(d.get('Yardi SD')))
                             else 'N/A'),
            })
        return [(f"Security Deposits — Portfolio Total ${total_deposits:,.2f}",
                 pd.DataFrame(rows_out))]
    render_report_buttons("deposits", "Security Deposits", _dep_sections,
                          meta=f"Portfolio total security deposits: ${total_deposits:,.2f}")

    # Display by building
    for building_name in BUILDING_MAP:
        b_deps = [d for d in dep_data if d['Building'] == building_name]
        if not b_deps:
            continue
        b_total = sum(d['Current SD'] for d in b_deps)
        code = BUILDING_MAP[building_name]['code']
        st.markdown(f'<div class="building-header"><strong>▎ {building_name} ({code})</strong> — ${b_total:,.2f} total deposits held</div>', unsafe_allow_html=True)

        df = pd.DataFrame(b_deps)
        display_cols = ['Display Tenant', 'Space', 'Type', 'Lease', 'Current SD Fmt', 'SD Anniversary', 'New SD Amount', 'SD Delta', 'Yardi SD', 'SD Diff', 'TTE']
        display_df = df[display_cols].copy()
        display_df.columns = ['Tenant', 'Space', 'Type', 'Lease', 'Current SD', 'SD Anniversary', 'New SD', 'Δ SD', 'Yardi SD', 'SD Diff', 'MTE']
        display_df['SD Anniversary'] = display_df['SD Anniversary'].apply(
            lambda d: d.strftime('%m/%d/%Y') if isinstance(d, (datetime, date)) else ('-' if not d else str(d))
        )
        display_df['New SD'] = display_df['New SD'].apply(lambda x: f"${x:,.0f}" if x is not None and not (isinstance(x, float) and pd.isna(x)) else '$0')
        display_df['Δ SD'] = display_df['Δ SD'].apply(
            lambda x: f"{'+' if x and x > 0 else ''}${x:,.0f}" if x is not None and not (isinstance(x, float) and pd.isna(x)) else '$0'
        )
        display_df['Yardi SD'] = display_df['Yardi SD'].apply(
            lambda x: f"${x:,.2f}" if x is not None and not (isinstance(x, float) and pd.isna(x)) else 'N/A'
        )
        show_grid(display_df, key=f"dep_{building_name}", tab_key="deposits")

    render_column_config_editor('deposits', ['Tenant', 'Space', 'Type', 'Lease', 'Current SD', 'SD Anniversary', 'New SD', 'Δ SD', 'Yardi SD', 'SD Diff', 'MTE'])

    # Portfolio total
    st.divider()
    st.metric("Portfolio Total Security Deposits", f"${total_deposits:,.2f}")


def render_yardi_tab():
    import base64
    yardi_dir = Path(os.path.dirname(__file__)) / "data" / "Yardi"

    if not yardi_dir.exists():
        st.markdown("### 📊 Yardi Reports")
        st.warning("No Yardi reports found. Place PDF files in data/Yardi/ folder.")
        return

    pdfs = sorted([f for f in yardi_dir.iterdir() if f.suffix.lower() == '.pdf'])
    if not pdfs:
        st.markdown("### 📊 Yardi Reports")
        st.warning("No PDF files found in data/Yardi/.")
        return

    # --- Report buttons (index of available Yardi reports for the PDF) ---
    def _yardi_sections():
        import pandas as pd
        rows_out = []
        for p in pdfs:
            m = re.match(r'^([A-Za-z]{3})-(\d{4})', p.stem)
            period = f"{m.group(1)} {m.group(2)}" if m else ''
            bldg = ''
            for bn, mp in BUILDING_MAP.items():
                fn = p.stem.lower().replace('-', '').replace(' ', '').replace('_', '')
                if any(x in fn for x in [mp['dest_folder'].lower().replace('-', '').replace(' ', ''),
                                         bn.lower().replace(' ', ''), mp['code'].lower()]):
                    bldg = mp['code']
                    break
            rows_out.append({'Period': period, 'Building': bldg, 'File': p.name})
        return [("Available Yardi Reports", pd.DataFrame(rows_out))]
    render_report_buttons("yardi", "Yardi Reports", _yardi_sections,
                          meta=f"{len(pdfs)} monthly statement report(s) on file.")
    st.caption(f"{len(pdfs)} report(s) available")

    # Sort PDFs by date (newest first) using Mon-YYYY prefix
    def _pdf_sort_key(p):
        m = re.match(r'^([a-z]{3})-(\d{4})', p.stem.lower())
        if m:
            return int(m.group(2)) * 100 + MONTH_ORDER.get(m.group(1), 0)
        return 0

    pdfs_sorted = sorted(pdfs, key=_pdf_sort_key, reverse=True)

    # Group by building
    building_files = {}
    for pdf in pdfs_sorted:
        name = pdf.stem
        matched_building = None
        for bldg_name, mapping in BUILDING_MAP.items():
            dest = mapping['dest_folder'].lower().replace('-', '').replace(' ', '')
            fname = name.lower().replace('-', '').replace(' ', '').replace('_', '')
            if any(part in fname for part in [
                dest, bldg_name.lower().replace(' ', ''),
                mapping['code'].lower()
            ]):
                matched_building = bldg_name
                break
        if not matched_building:
            matched_building = "Other"
        building_files.setdefault(matched_building, []).append(pdf)

    for building, files in building_files.items():
        code = BUILDING_MAP.get(building, {}).get('code', '')
        label = f"{building} ({code})" if code else building
        st.markdown(f'<div style="margin:8px 0 2px 0;font-size:0.95em;"><strong>▎ {label}</strong> — {len(files)} report(s)</div>', unsafe_allow_html=True)

        for pdf in files:
            with open(pdf, "rb") as f:
                pdf_bytes = f.read()
            b64 = base64.b64encode(pdf_bytes).decode()

            # Clean display name: "May 2026" from "May-2026_10097-15-South-St_..."
            date_match = re.match(r'^([A-Za-z]{3})-(\d{4})', pdf.stem)
            if date_match:
                display_name = f"{date_match.group(1)} {date_match.group(2)}"
            else:
                display_name = pdf.stem.replace('_', ' ').replace('-', ' ')

            with st.expander(f"📄 {display_name}", expanded=False):
                col_dl_inner, col_spacer = st.columns([2, 5])
                col_dl_inner.download_button("⬇️ Download", data=pdf_bytes, file_name=pdf.name, mime="application/pdf", key=f"dl_{pdf.name}")
                if st.session_state.get('mobile_view', False) and HAS_PYMUPDF:
                    doc = fitz.open(str(pdf))
                    total_pages = len(doc)
                    page_key = f"yardi_page_{pdf.name}"
                    if page_key not in st.session_state:
                        st.session_state[page_key] = 0
                    pg = max(0, min(st.session_state[page_key], total_pages - 1))
                    cp, cn_num, cn = st.columns([1, 2, 1])
                    if cp.button("⬅️", disabled=(pg == 0), key=f"yp_{pdf.name}"):
                        st.session_state[page_key] = pg - 1
                        st.rerun()
                    cn_num.markdown(f"<div style='text-align:center;padding-top:8px;'>{pg+1} / {total_pages}</div>", unsafe_allow_html=True)
                    if cn.button("➡️", disabled=(pg >= total_pages - 1), key=f"yn_{pdf.name}"):
                        st.session_state[page_key] = pg + 1
                        st.rerun()
                    page = doc[pg]
                    pix = page.get_pixmap(matrix=fitz.Matrix(0.9, 0.9))
                    img_b64 = base64.b64encode(pix.tobytes("png")).decode()
                    doc.close()
                    st.markdown(f'<img src="data:image/png;base64,{img_b64}" style="width:100%;border-radius:6px;">', unsafe_allow_html=True)
                else:
                    st.markdown(f'<iframe src="data:application/pdf;base64,{b64}" width="100%" height="600px" style="border:1px solid rgba(255,255,255,0.1);border-radius:6px;"></iframe>', unsafe_allow_html=True)


def render_reconcile_tab():
    """Yardi Reconciliation — Yardi-driven with lookup table at top for name matching."""
    import json as _json
    st.markdown("### 🔄 Yardi Reconciliation")

    tenants, details, summaries = load_tenancy()
    yardi_data = parse_yardi_rent_rolls()
    if not yardi_data:
        st.warning("No Yardi rent roll data found. Place PDF files in data/Yardi/ folder.")
        return

    # --- Name mapping: Yardi key -> spreadsheet tenant name (persisted in Google Sheets) ---
    MAPPING_TAB = "Config: Yardi Names"
    name_map = {}  # key: "Building|YardiSpace" -> spreadsheet tenant name
    try:
        raw = _read_gsheet_config(MAPPING_TAB)
        if raw:
            name_map = _json.loads(raw)
    except Exception:
        pass

    # Build spreadsheet tenant index keyed by "Building|TenantName" for lookup
    active = [t for t in tenants if t['Tenant'] not in ('-',)] if tenants else []
    sheet_by_name = {}  # "Building|TenantName" -> tenant dict
    sheet_by_space = {}  # "Building|Space" -> tenant dict
    for t in active:
        b = t.get('Building', '').strip()
        n = t.get('Tenant', '').strip()
        s = str(t.get('Space', '')).strip()
        if b and n:
            sheet_by_name[f"{b}|{n}"] = t
        if b and s:
            sheet_by_space[f"{b}|{s}"] = t

    # Get all unique spreadsheet tenant names per building for dropdown
    building_tenants = {}
    for t in active:
        b = t.get('Building', '').strip()
        n = t.get('Tenant', '').strip()
        if b and n:
            building_tenants.setdefault(b, []).append(n)
    for b in building_tenants:
        building_tenants[b] = sorted(set(building_tenants[b]))

    # Helper: get spreadsheet SD for a tenant
    def _get_sheet_sd(building, tenant_name):
        sd = 0
        for dk, dv in details.items():
            if not dv:
                continue
            first = dv[0]
            if first.get('Building') == building and first.get('Tenant') == tenant_name:
                best_end = None
                for r in dv:
                    start = r.get('Start Date')
                    end = r.get('End Date')
                    s = r.get('Sec Dep', 0) or 0
                    if isinstance(start, (datetime, date)) and isinstance(end, (datetime, date)):
                        start_d = start.date() if isinstance(start, datetime) else start
                        end_d = end.date() if isinstance(end, datetime) else end
                        if start_d <= TODAY <= end_d:
                            if best_end is None or end_d < best_end:
                                best_end = end_d
                                sd = s
                break
        return sd

    # ===========================
    # SECTION 1: LOOKUP TABLE (compact)
    # ===========================
    with st.expander("🔗 Name Lookup Table", expanded=False):
        st.caption("Match Yardi names → spreadsheet names. Changes save automatically.")

        # Inject compact CSS for lookup table
        st.markdown("""<style>
            .lookup-table { width: 100%; border-collapse: collapse; font-size: 0.8rem; margin-bottom: 0.5rem; }
            .lookup-table th { text-align: left; padding: 2px 6px; border-bottom: 1px solid #444; color: #888; font-weight: 600; }
            .lookup-table td { padding: 1px 6px; border-bottom: 1px solid #2a2a2a; }
            .lookup-table .unit { color: #888; width: 50px; }
            .lookup-table .yname { white-space: nowrap; max-width: 200px; overflow: hidden; text-overflow: ellipsis; }
        </style>""", unsafe_allow_html=True)

        mapping_changed = False

        for building_name in BUILDING_MAP:
            b_yardi = {k: v for k, v in yardi_data.items() if k.startswith(f"{building_name}|")}
            if not b_yardi:
                continue

            code = BUILDING_MAP[building_name]['code']
            st.markdown(f"<small><strong>{building_name} ({code})</strong></small>", unsafe_allow_html=True)

            available_names = ["—"] + building_tenants.get(building_name, [])

            cols = st.columns([1, 3, 3])
            cols[0].markdown("<small>**Unit**</small>", unsafe_allow_html=True)
            cols[1].markdown("<small>**Yardi Name**</small>", unsafe_allow_html=True)
            cols[2].markdown("<small>**Sheet Match**</small>", unsafe_allow_html=True)

            for yardi_key in sorted(b_yardi.keys()):
                yardi_entry = b_yardi[yardi_key]
                yardi_name = yardi_entry.get('tenant', '')
                yardi_space = yardi_key.split('|', 1)[1] if '|' in yardi_key else ''

                current_match = name_map.get(yardi_key, '')
                if not current_match:
                    norm_space = normalize_space(yardi_entry.get('raw_unit', yardi_space), building_name)
                    auto_key = f"{building_name}|{norm_space}"
                    auto_tenant = sheet_by_space.get(auto_key)
                    if auto_tenant:
                        current_match = auto_tenant.get('Tenant', '')

                default_idx = 0
                if current_match in available_names:
                    default_idx = available_names.index(current_match)

                c1, c2, c3 = st.columns([1, 3, 3])
                c1.markdown(f"<small style='color:#888;padding-top:8px;display:block;'>{yardi_space}</small>", unsafe_allow_html=True)
                c2.markdown(f"<small style='padding-top:8px;display:block;'>{yardi_name}</small>", unsafe_allow_html=True)
                selected = c3.selectbox("m", available_names, index=default_idx,
                                        key=f"lu_{yardi_key}", label_visibility="collapsed")

                new_val = '' if selected == "—" else selected
                if new_val != name_map.get(yardi_key, ''):
                    name_map[yardi_key] = new_val
                    mapping_changed = True

        if mapping_changed:
            name_map = {k: v for k, v in name_map.items() if v}
            _write_gsheet_config(MAPPING_TAB, _json.dumps(name_map))
            st.rerun()

    # ===========================
    # SECTION 2: RECONCILIATION GRID
    # ===========================
    st.markdown("---")
    st.markdown("#### 📊 Data Comparison")

    # Get Yardi deposit activity (Deposits On Hand) for SD comparison
    yardi_deposits = parse_yardi_deposit_activity()

    recon_rows = []
    for yardi_key, yardi_entry in sorted(yardi_data.items()):
        building = yardi_key.split('|', 1)[0] if '|' in yardi_key else ''
        yardi_space = yardi_key.split('|', 1)[1] if '|' in yardi_key else ''
        yardi_name = yardi_entry.get('tenant', '')
        yardi_monthly = yardi_entry.get('monthly', 0) or 0
        yardi_exp = yardi_entry.get('expiration')
        yardi_sd = yardi_deposits.get(yardi_key, 0) or 0

        # Find matched spreadsheet tenant
        matched_name = name_map.get(yardi_key, '')
        if not matched_name:
            # Auto-match via normalized space
            norm_space = normalize_space(yardi_entry.get('raw_unit', yardi_space), building)
            auto_key = f"{building}|{norm_space}"
            auto_t = sheet_by_space.get(auto_key)
            if auto_t:
                matched_name = auto_t.get('Tenant', '')

        # Get spreadsheet data for matched tenant
        sheet_t = sheet_by_name.get(f"{building}|{matched_name}") if matched_name else None
        dash_monthly = sheet_t.get('Monthly', 0) if sheet_t else None
        dash_exp_str = sheet_t.get('Exp Date', '') if sheet_t else ''
        dash_sd = _get_sheet_sd(building, matched_name) if matched_name else None

        # Compute diffs
        rent_diff = ''
        if sheet_t and dash_monthly is not None:
            diff_val = dash_monthly - yardi_monthly
            if abs(diff_val) > 0.01:
                rent_diff = f"${diff_val:+,.0f}"

        exp_diff = ''
        yardi_exp_str = yardi_exp.strftime('%m/%d/%Y') if yardi_exp else 'N/A'
        if dash_exp_str and dash_exp_str != 'MTM' and yardi_exp:
            try:
                d_exp = datetime.strptime(dash_exp_str, '%m/%d/%Y').date()
                if d_exp != yardi_exp:
                    delta = (d_exp - yardi_exp).days
                    exp_diff = f"{delta:+d}d"
            except (ValueError, AttributeError):
                pass

        sd_diff = ''
        if dash_sd is not None and abs(dash_sd - yardi_sd) > 0.01:
            sd_diff = f"${dash_sd - yardi_sd:+,.0f}"

        # Status
        if not matched_name:
            status = '❓ Unmatched'
        else:
            diffs = []
            if rent_diff:
                diffs.append('Rent')
            if exp_diff:
                diffs.append('Exp')
            if sd_diff:
                diffs.append('SD')
            status = '⚠️ ' + ', '.join(diffs) if diffs else '✅ Match'

        recon_rows.append({
            'Building': building,
            'Space': yardi_space,
            'Yardi Name': yardi_name,
            'Sheet Name': matched_name or '—',
            'Rent (Yardi)': f"${yardi_monthly:,.0f}",
            'Rent (Sheet)': f"${dash_monthly:,.0f}" if dash_monthly is not None else '—',
            'Rent Δ': rent_diff,
            'Exp (Yardi)': yardi_exp_str,
            'Exp (Sheet)': dash_exp_str if dash_exp_str else '—',
            'Exp Δ': exp_diff,
            'SD (Yardi)': f"${yardi_sd:,.2f}",
            'SD (Sheet)': f"${dash_sd:,.2f}" if dash_sd is not None else '—',
            'SD Δ': sd_diff,
            'Status': status,
        })

    # Append sheet-only tenants (in spreadsheet but not in Yardi) into main table
    matched_sheet_keys_main = set()
    for r in recon_rows:
        if r['Sheet Name'] and r['Sheet Name'] != '—':
            matched_sheet_keys_main.add(f"{r['Building']}|{r['Sheet Name']}")
    for yk, sn in name_map.items():
        b = yk.split('|', 1)[0] if '|' in yk else ''
        if b and sn:
            matched_sheet_keys_main.add(f"{b}|{sn}")

    for t in active:
        t_key = f"{t.get('Building', '')}|{t.get('Tenant', '')}"
        if t_key not in matched_sheet_keys_main and t.get('Tenant', '') not in ('-', '') and 'VACANT' not in t.get('Tenant', '').upper():
            dash_monthly = t.get('Monthly', 0) or 0
            dash_sd = _get_sheet_sd(t.get('Building', ''), t.get('Tenant', ''))
            recon_rows.append({
                'Building': t.get('Building', ''),
                'Space': str(t.get('Space', '')),
                'Yardi Name': '—',
                'Sheet Name': t.get('Tenant', ''),
                'Rent (Yardi)': '—',
                'Rent (Sheet)': f"${dash_monthly:,.0f}",
                'Rent Δ': '',
                'Exp (Yardi)': '—',
                'Exp (Sheet)': t.get('Exp Date', ''),
                'Exp Δ': '',
                'SD (Yardi)': '—',
                'SD (Sheet)': f"${dash_sd:,.2f}" if dash_sd else '—',
                'SD Δ': '',
                'Status': '📋 Sheet Only',
            })

    recon_rows.sort(key=lambda r: (r['Building'], r['Space']))

    # Summary metrics
    total = len(recon_rows)
    matches = sum(1 for r in recon_rows if r['Status'] == '✅ Match')
    mismatches = sum(1 for r in recon_rows if r['Status'].startswith('⚠️'))
    unmatched = sum(1 for r in recon_rows if r['Status'] == '❓ Unmatched')

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Yardi Units", str(total))
    c2.metric("✅ Match", str(matches))
    c3.metric("⚠️ Mismatch", str(mismatches))
    c4.metric("❓ Unmatched", str(unmatched))

    # Filter
    filter_opt = st.radio("Show:", ["All", "Mismatches Only", "Unmatched Only"], horizontal=True, key="recon_filter")
    filtered = recon_rows
    if filter_opt == "Mismatches Only":
        filtered = [r for r in recon_rows if r['Status'].startswith('⚠️')]
    elif filter_opt == "Unmatched Only":
        filtered = [r for r in recon_rows if r['Status'] == '❓ Unmatched']

    # --- Report buttons (condensed reconcile view for the PDF) ---
    def _recon_sections():
        if not filtered:
            return [("Yardi Reconciliation", pd.DataFrame())]
        rows_out = []
        for r in filtered:
            rows_out.append({
                'Bldg': BUILDING_MAP.get(r.get('Building'), {}).get('code', r.get('Building', '')),
                'Space': r.get('Space', ''),
                'Yardi Name': r.get('Yardi Name', ''),
                'Sheet Name': r.get('Sheet Name', ''),
                'Rent Y/S': f"{r.get('Rent (Yardi)', '')} / {r.get('Rent (Sheet)', '')}",
                'ΔRent': r.get('Rent Δ', ''),
                'ΔExp': r.get('Exp Δ', ''),
                'ΔSD': r.get('SD Δ', ''),
                'Status': r.get('Status', ''),
            })
        return [("Yardi vs Sheet Reconciliation", pd.DataFrame(rows_out))]
    render_report_buttons("reconcile", "Yardi Reconcile", _recon_sections)

    for building_name in BUILDING_MAP:
        b_rows = [r for r in filtered if r['Building'] == building_name]
        if not b_rows:
            continue
        b_match = sum(1 for r in b_rows if r['Status'] == '✅ Match')
        code = BUILDING_MAP[building_name]['code']
        st.markdown(f'<div class="building-header"><strong>▎ {building_name} ({code})</strong> — {b_match}/{len(b_rows)} matching</div>', unsafe_allow_html=True)

        df = pd.DataFrame(b_rows)
        display_cols = ['Space', 'Yardi Name', 'Sheet Name', 'Rent (Yardi)', 'Rent (Sheet)', 'Rent Δ',
                        'Exp (Yardi)', 'Exp (Sheet)', 'Exp Δ', 'SD (Yardi)', 'SD (Sheet)', 'SD Δ', 'Status']
        display_df = df[display_cols].copy()
        show_grid(display_df, key=f"recon_{building_name}", tab_key="reconcile")

    render_column_config_editor('reconcile', ['Space', 'Yardi Name', 'Sheet Name', 'Rent (Yardi)', 'Rent (Sheet)', 'Rent Δ',
                                               'Exp (Yardi)', 'Exp (Sheet)', 'Exp Δ', 'SD (Yardi)', 'SD (Sheet)', 'SD Δ', 'Status'])

    # ===========================
    # Sheet-only tenants are now included in the main recon_rows table above


# =====================
# LEAD SHEET
# =====================
LEAD_SHEET_ID = "1Gbl7FM-C4dN_BHVwT_1dRbkODEZW57MdymzBMDvPrCw"
LEAD_SHEET_GID = 1762616490

@st.cache_data(ttl=300)
def _load_lead_changelog():
    """Load the Lead Sheet Changelog and return dict of {(section, row): (timestamp, user)} for most recent edit per row."""
    try:
        sheet = get_gsheet()
        if not sheet:
            return {}
        ws = sheet.worksheet('Lead Sheet Changelog')
        records = ws.get_all_records()
    except Exception:
        # Try the lead sheet's own spreadsheet
        try:
            client = get_gspread_client()
            if not client:
                return {}
            ss = client.open_by_key(LEAD_SHEET_ID)
            ws = ss.worksheet('Lead Sheet Changelog')
            records = ws.get_all_records()
        except Exception:
            return {}

    latest = {}
    for r in records:
        section = r.get('Section', 'Retail')
        row = r.get('Row', 0)
        ts = r.get('Timestamp', '')
        user = r.get('User', '')
        key = (section, row)
        # Keep the latest timestamp per row
        if key not in latest or str(ts) > str(latest[key][0]):
            latest[key] = (ts, user)
    return latest


@st.cache_data(ttl=300)
def load_lead_sheet():
    """Load lead sheet from Google Sheets and return retail_df, office_df."""
    client = get_gspread_client()
    if client is None:
        st.error("No Google service account configured.")
        return pd.DataFrame(), pd.DataFrame()
    try:
        sheet = client.open_by_key(LEAD_SHEET_ID)
        ws = sheet.get_worksheet_by_id(LEAD_SHEET_GID)
        rows = ws.get_all_values()
    except Exception as e:
        st.error(f"Failed to load lead sheet: {e}")
        return pd.DataFrame(), pd.DataFrame()

    if not rows:
        return pd.DataFrame(), pd.DataFrame()

    # Dynamically read column layout from header row
    header = rows[0]

    # Find the blank separator column between retail and office
    sep_col = None
    for i, h in enumerate(header):
        if i > 0 and not h.strip():
            # Check if previous cols had data and next col looks like a header
            if i + 1 < len(header) and header[i + 1].strip():
                sep_col = i
                break

    if sep_col is None:
        # Fallback: assume no office section
        sep_col = len(header)

    retail_headers = [h.strip() for h in header[:sep_col]]
    office_headers = [h.strip() for h in header[sep_col + 1:]] if sep_col < len(header) - 1 else []

    # Normalize office headers to match retail (strip "Office "/"Retail " prefix from first col)
    # Use retail headers as canonical column names, replacing first col name
    col_names = list(retail_headers)
    col_names[0] = "Tenant"

    # Map date columns to age column names, preserving order
    _age_name_map = {"Contacted": "Contact", "Shown": "Show"}
    date_cols = {}
    for c in col_names:
        if 'date' in c.lower():
            raw = c.replace("Date ", "").replace("Date", "").strip()
            friendly = _age_name_map.get(raw, raw)
            date_cols[c] = f"{friendly} Age"

    def parse_rows(rows_data, start_col, num_cols):
        parsed = []
        for row in rows_data[1:]:
            cells = row[start_col:start_col + num_cols] if len(row) > start_col else []
            # Pad if short
            while len(cells) < num_cols:
                cells.append('')
            cells = [c.strip() for c in cells[:num_cols]]
            if any(c for c in cells):
                parsed.append(cells)
        return parsed

    def add_age_columns(df):
        today = date.today()
        for date_col, age_col in date_cols.items():
            ages = []
            for val in df[date_col]:
                if val:
                    try:
                        dt = pd.to_datetime(val, format='mixed', dayfirst=False).date()
                        delta = (today - dt).days
                        ages.append(f"{delta}d")
                    except Exception:
                        ages.append("")
                else:
                    ages.append("")
            # Insert age column right after its date column
            idx = df.columns.get_loc(date_col) + 1
            df.insert(idx, age_col, ages)
        return df

    retail_rows = parse_rows(rows, 0, len(retail_headers))
    office_cols_count = len(office_headers) if office_headers else len(retail_headers)
    office_rows = parse_rows(rows, sep_col + 1, office_cols_count)

    # Office headers may differ slightly; normalize to match retail col_names
    office_col_names = list(col_names)
    if office_headers:
        office_col_names = list(office_headers)
        office_col_names[0] = "Tenant"
    # Ensure same length
    while len(office_col_names) < office_cols_count:
        office_col_names.append(f"Col_{len(office_col_names)}")

    retail_df = pd.DataFrame(retail_rows, columns=col_names) if retail_rows else pd.DataFrame(columns=col_names)
    office_df = pd.DataFrame(office_rows, columns=office_col_names[:office_cols_count]) if office_rows else pd.DataFrame(columns=col_names)

    retail_df = add_age_columns(retail_df)
    office_df = add_age_columns(office_df)

    # Merge changelog data (last modified + modified by)
    changelog = _load_lead_changelog()
    for section, df in [('Retail', retail_df), ('Office', office_df)]:
        last_mod = []
        mod_by = []
        for i in range(len(df)):
            sheet_row = i + 2  # row 1 is header, data starts at row 2
            entry = changelog.get((section, sheet_row))
            if entry:
                ts_str = str(entry[0])
                # Shorten email to just the name part
                user = entry[1].split('@')[0] if '@' in entry[1] else entry[1]
                last_mod.append(ts_str)
                mod_by.append(user)
            else:
                last_mod.append('')
                mod_by.append('')
        df['Last Modified'] = last_mod
        df['Modified By'] = mod_by

    return retail_df, office_df


def render_covenants_tab():
    """Render the Lease Covenants tab directly from MSP Tenancy.xlsx (cols AI-AU, rows 10-40)."""
    st.caption("Special lease provisions from MSP Tenancy.xlsx · Edit the spreadsheet and re-upload to update")

    if not TENANCY_FILE:
        st.warning("MSP Tenancy.xlsx not found.")
        return

    # Column mapping: spreadsheet column letter → display name
    COV_COLUMNS = {
        35: "Renewal Options",
        36: "Option Notice",
        37: "LL Termination",
        38: "Tenant Termination",
        39: "Non-Compete / Exclusive",
        40: "CAM Obligations",
        41: "ROFO/ROFR",
        42: "Assignment/Subletting",
        43: "Personal Guarantee",
        44: "Holdover Rate",
        45: "Late Fee",
        46: "Other Notable",
    }

    try:
        wb = openpyxl.load_workbook(TENANCY_FILE, data_only=True)
        ws = wb.active

        rows = []
        for r in range(11, 41):
            building = ws.cell(row=r, column=1).value
            tenant = ws.cell(row=r, column=3).value
            if not building or not tenant:
                continue
            tenant_str = str(tenant)
            is_vacant = tenant_str.upper().startswith("VACANT") or tenant_str.upper() == "EASEMENT"
            row_data = {
                "Building": str(building),
                "Unit": ws.cell(row=r, column=2).value or "",
                "Tenant": "VACANT" if is_vacant else tenant_str,
            }
            for col_idx, col_name in COV_COLUMNS.items():
                if is_vacant:
                    row_data[col_name] = ""
                else:
                    val = ws.cell(row=r, column=col_idx).value
                    if val is not None and isinstance(val, datetime) and col_idx == 36:
                        row_data[col_name] = val.strftime("%m/%d/%Y")
                    else:
                        row_data[col_name] = str(val) if val is not None else ""
            rows.append(row_data)
        wb.close()
    except Exception as e:
        st.error(f"Error reading tenancy file: {e}")
        return

    if not rows:
        st.info("No covenant data found in spreadsheet.")
        return

    df = pd.DataFrame(rows)

    # Report buttons: condensed key-covenant columns (full set is too wide for PDF)
    def _cov_sections():
        key_cols = [c for c in ["Building", "Unit", "Tenant", "Renewal Options",
                                "Non-Compete / Exclusive", "ROFO/ROFR",
                                "Personal Guarantee", "Assignment/Subletting"]
                    if c in df.columns]
        return [("Lease Covenants (key provisions)", df[key_cols])]
    render_report_buttons("covenants", "Lease Covenants", _cov_sections)

    # --- Building filter ---
    buildings = sorted(df["Building"].unique().tolist())
    selected = st.multiselect("Filter by Building", buildings, default=buildings, key="cov_building_filter")
    filtered = df[df["Building"].isin(selected)].copy()

    # Summary metrics
    total = len(filtered)

    def _count_has(col_name, keyword="YES"):
        if col_name not in filtered.columns:
            return 0
        return filtered[col_name].astype(str).str.upper().str.contains(keyword, na=False).sum()

    has_noncompete = _count_has("Non-Compete / Exclusive")
    has_rofr = _count_has("ROFO/ROFR")
    has_guarantee = _count_has("Personal Guarantee")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Tenants", total)
    c2.metric("Non-Compete", has_noncompete)
    c3.metric("ROFO/ROFR", has_rofr)
    c4.metric("Personal Guarantee", has_guarantee)

    # Display per building
    for bldg in selected:
        bldg_df = filtered[filtered["Building"] == bldg].copy()
        if bldg_df.empty:
            continue
        st.markdown(f"#### 🏢 {bldg}")
        display_df = bldg_df.drop(columns=["Building"]).reset_index(drop=True)
        # Auto-size: 42px header + 42px per row + buffer
        grid_h = 42 + 42 * len(display_df) + 12
        show_grid(display_df, key=f"cov_{bldg}", height=grid_h, tab_key="covenants")

    # Column width config at the bottom
    all_cov_cols = ["Unit", "Tenant"] + list(COV_COLUMNS.values())
    render_column_config_editor('covenants', all_cov_cols)


def render_lead_sheet_tab():
    """Render the Lead Sheet tab with Retail and Office sections."""
    retail_df, office_df = load_lead_sheet()
    render_report_buttons(
        "leads", "Lead Sheet",
        lambda: [("Retail Leads", retail_df), ("Office Leads", office_df)])
    st.caption("Prospect leads from Google Sheets · Auto-refreshes every 5 minutes")

    if retail_df.empty and office_df.empty:
        st.warning("No lead data found. Check Google Sheets connection.")
        return

    # Summary metrics
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(f'<div class="metric-card"><div class="metric-value">{len(retail_df)}</div><div class="metric-label">Retail Leads</div></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-card"><div class="metric-value">{len(office_df)}</div><div class="metric-label">Office Leads</div></div>', unsafe_allow_html=True)
    showing_count = sum(1 for _, r in pd.concat([retail_df, office_df]).iterrows() if r.get("Showing", "").strip().lower() == "yes")
    loi_count = sum(1 for _, r in pd.concat([retail_df, office_df]).iterrows() if r.get("LOI", "").strip().lower() == "yes")
    c3.markdown(f'<div class="metric-card"><div class="metric-value">{showing_count}</div><div class="metric-label">Showings Done</div></div>', unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-card"><div class="metric-value">{loi_count}</div><div class="metric-label">LOIs Received</div></div>', unsafe_allow_html=True)

    st.markdown("")

    # Retail section
    st.markdown('<div class="building-header">🏪 Retail Leads</div>', unsafe_allow_html=True)
    if not retail_df.empty:
        show_grid(retail_df, key="leads_retail", tab_key="leads")
    else:
        st.info("No retail leads.")

    st.markdown("")

    # Office section
    st.markdown('<div class="building-header">🏢 Office Leads</div>', unsafe_allow_html=True)
    if not office_df.empty:
        show_grid(office_df, key="leads_office", tab_key="leads")
    else:
        st.info("No office leads.")

    # Use retail columns (with age cols) as the config source
    all_cols = list(retail_df.columns) if not retail_df.empty else list(office_df.columns)
    render_column_config_editor('leads', all_cols)

    if st.button("🔄 Refresh Lead Sheet", key="refresh_leads"):
        load_lead_sheet.clear()
        st.rerun()


# =====================
# LEASE BUILDER TAB
# =====================

LEASE_MANDATORY_BOOKMARKS = {
    "Tx_BuildingAddress", "Tx_Landlord", "Tx_Tenant", "Tx_Premises", "Tx_Sqft",
    "Tx_PermittedUse", "Tx_LeaseType", "Tx_LeaseCommenceDt", "Tx_RentCommDt",
    "Tx_LeaseExpDt", "Tx_LeaseTerm",
}

LEASE_DEFAULT_LINKS = {
    "Tx_BuildingAddress": "1",
    "Tx_Premises": "1",
    "Tx_Sqft": "1",
    "Tx_TenantShareProperty": "1",
    "Tx_TenantShareFloor": "1",
    "Tx_TenantSharePremisis": "1",
    "Tx_LeaseCommenceDt": "2",
    "Tx_RentCommDt": "2",
    "Tx_LeaseExpDt": "2",
    "Tx_LeaseTerm": "2",
    "Tx_PermittedUse": "4",
    "Tx_LeaseType": "5",
    "Tx_AdditionalRent": "5.1",
    "Tx_OptionPeriod": "5.3",
    "Tx_Utilities": "6",
    "Tx_SecurityDeposit": "8",
    "Tx_LandlordContFitUp": "9",
    "Tx_Brokers": "14",
}



def _lb_bool(value, default=False):
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "yes", "y", "1", "use", "include"}:
        return True
    if text in {"false", "no", "n", "0", "don't use", "exclude"}:
        return False
    return default


def _lb_export_key_provisions_xlsx(rows):
    """Export the editable key-provision table; Bookmark is hidden but retained for re-import."""
    export_rows = []
    for row in rows:
        alternates = list(row.get("Alternates", []))[:10]
        export_row = {
            "Group": row.get("Group", "Optional"),
            "Use": bool(row.get("Include", True)),
            "Key Provision": row.get("Field", ""),
            # Carets become real newlines so the cell reads naturally in Excel;
            # Alt+Enter there comes back as a caret on import.
            "Current Value": kp_value_plain(row.get("Value", "")),
            "Choice": ld.normalize_choice(row),
            "Link": bool(row.get("Link", False)),
            "Target Section": row.get("Section", ""),
            "Bookmark": row.get("Bookmark", ""),
        }
        for index in range(10):
            export_row[f"Alt {index + 1}"] = (
                kp_value_plain(alternates[index]) if index < len(alternates) else ""
            )
        export_rows.append(export_row)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(export_rows).to_excel(writer, index=False, sheet_name="Key Provisions")
        worksheet = writer.book["Key Provisions"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        # Choice sits at E, so everything after it shifts one column right and
        # Bookmark — the hidden internal id — is now H rather than G.
        worksheet.column_dimensions["H"].hidden = True
        widths = {"A": 14, "B": 9, "C": 28, "D": 70, "E": 14, "F": 9, "G": 34}
        # A real dropdown in Excel, so the round trip is a chooser too and not
        # a free-text field where "alt2" silently fails to match.
        if worksheet.max_row > 1:
            choice_rule = DataValidation(
                type="list",
                formula1='"' + ",".join(
                    f"Alt {index}" for index in range(1, 11)
                ) + '"',
                allow_blank=True,
                showDropDown=False,
            )
            worksheet.add_data_validation(choice_rule)
            choice_rule.add(f"E2:E{worksheet.max_row}")
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width
        # Multi-line values are only visible in Excel when the cell wraps.
        for excel_row in worksheet.iter_rows(min_row=2, min_col=4):
            for cell in excel_row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    return output.getvalue()


def _lb_import_key_provisions_xlsx(uploaded_file, existing_rows, section_labels):
    """Merge an uploaded key-provision workbook into the current draft by Bookmark or name."""
    uploaded_file.seek(0)
    imported = pd.read_excel(uploaded_file, sheet_name=0, dtype=object)
    imported = imported.astype(object).where(imported.notna(), "")
    aliases = {
        "Group": "Group", "Use": "Include", "Include": "Include",
        "Key Provision": "Field", "Field": "Field", "Name": "Field",
        "Current Value": "Value", "Value": "Value", "Link": "Link", "Target Section": "Section",
        "Section": "Section", "Bookmark": "Bookmark",
    }
    normalized = {}
    for column in imported.columns:
        key = aliases.get(str(column).strip())
        if key:
            normalized[key] = column
    if "Field" not in normalized and "Bookmark" not in normalized:
        raise ValueError("The workbook needs a Key Provision/Field or Bookmark column.")

    by_bookmark = {str(row.get("Bookmark", "")): row for row in existing_rows}
    by_field = {str(row.get("Field", "")).strip().lower(): row for row in existing_rows}
    # Rows are collected in spreadsheet order, so reordering in Excel reorders
    # the provisions too. Matching is by identity, not by id(), which is not a
    # safe key for dicts that come and go during the loop.
    ordered = []
    matched = set()
    for _, source in imported.iterrows():
        bookmark = str(source.get(normalized.get("Bookmark", "__missing__"), "")).strip()
        field = str(source.get(normalized.get("Field", "__missing__"), "")).strip()
        if not bookmark and not field:
            continue  # a blank spreadsheet row is padding, not a provision
        row = by_bookmark.get(bookmark) if bookmark else by_field.get(field.lower())
        if row is None:
            row = {
                "Group": "Optional",
                "Include": True,
                "Field": field or "New Key Provision",
                "Value": "",
                "Alternates": [],
                "Choice": "Current Value",
                "Link": False,
                "Section": next(iter(section_labels.values())),
                "Bookmark": "Custom_" + uuid4().hex[:12],
            }
        else:
            matched.add(str(row.get("Bookmark", "")) or str(row.get("Field", "")))
        ordered.append(row)
        if "Group" in normalized:
            group = str(source[normalized["Group"]]).strip()
            row["Group"] = group if group in {"Mandatory", "Optional"} else "Optional"
        if "Include" in normalized:
            row["Include"] = _lb_bool(source[normalized["Include"]], row.get("Include", True))
        if "Field" in normalized and field:
            row["Field"] = field
        if "Value" in normalized:
            # An Alt+Enter newline typed in Excel is stored as a caret.
            row["Value"] = normalize_kp_value(source[normalized["Value"]])
        # Keep literal Alt 1–Alt 10 positions. Empty earlier cells must not
        # shift Alt 3 into Alt 1.
        alternate_slots = []
        for index in range(1, 11):
            column = next(
                (column_name for column_name in imported.columns
                 if str(column_name).strip().lower() == f"alt {index}".lower()),
                None,
            )
            alternate_slots.append(
                normalize_kp_value(source[column]) if column is not None else ""
            )
        row["Alternates"] = alternate_slots
        if "Choice" in normalized:
            row["Choice"] = str(source[normalized["Choice"]]).strip()
        # Applied after Alternates are in place, so a choice typed in Excel is
        # validated against the slots from the same upload rather than the ones
        # that happened to be there before it.
        row.update(ld.apply_choice(row))
        if "Link" in normalized:
            row["Link"] = _lb_bool(source[normalized["Link"]], row.get("Link", False))
        if "Section" in normalized:
            target = str(source[normalized["Section"]]).strip()
            target = section_labels.get(target, target) if isinstance(section_labels, dict) else target
            row["Section"] = target if target in section_labels.values() else next(iter(section_labels.values()))
    # The spreadsheet is the whole list, not a set of additions: a provision
    # deleted there has to disappear here, otherwise there is no way to remove
    # one at all. Removals are returned so the caller can report them rather
    # than quietly dropping rows out of a lease.
    removed = [
        str(row.get("Field", ""))
        for row in existing_rows
        if (str(row.get("Bookmark", "")) or str(row.get("Field", ""))) not in matched
    ]
    # A sheet that matched nothing is far more likely to be the wrong file than
    # a deliberate wipe, so the existing rows are kept and the caller is told.
    if not ordered:
        raise ValueError(
            "No key provisions were found in that spreadsheet. Nothing was changed — "
            "check that it has Bookmark and Field columns."
        )
    return ordered, removed


def _lb_section_body(section_number, text):
    text = str(text or "").strip()
    pattern = re.compile(
        rf"^\s*Section\s+{re.escape(str(section_number))}(?:\.(?:\s+|$)|\s+)",
        re.IGNORECASE,
    )
    return pattern.sub("", text, count=1).strip()


def _lb_preview_html(key_provisions, preview_sections, focus_section="", show_all=False):
    selected_provisions = [
        item for item in key_provisions if show_all or bool(item.get("Include"))
    ]
    section_name_to_number = {
        f"Section {section['number']} — {section['title']}": str(section['number'])
        for section in preview_sections
    }
    linked_by_section = {}
    for item in selected_provisions:
        if bool(item.get("Link")) and item.get("Section"):
            target = section_name_to_number.get(str(item["Section"]), str(item["Section"]))
            linked_by_section.setdefault(target, []).append(item)

    summary_rows = "".join(
        "<tr><th>" + html_lib.escape(str(item.get("Field", ""))) + "</th><td>" +
        html_lib.escape(kp_value_plain(item.get("Value", ""))).replace("\n", "<br>") + "</td></tr>"
        for item in selected_provisions
    ) or '<tr><td colspan="2" class="empty">No key provisions selected.</td></tr>'

    section_blocks = []
    for section in preview_sections:
        if not show_all and not bool(section.get("include", True)):
            continue
        number = str(section["number"])
        section_id = "section-" + re.sub(r"[^A-Za-z0-9_-]", "-", number)
        body = _lb_section_body(number, section.get("text", ""))
        paragraphs = "".join(
            f"<p>{html_lib.escape(part.strip())}</p>"
            for part in re.split(r"\n\s*\n", body)
            if part.strip()
        )
        linked_paragraphs = "".join(
            "<p><strong>" + html_lib.escape(str(item.get("Field", ""))) + ".</strong> " +
            html_lib.escape(str(item.get("Value", ""))) + "</p>"
            for item in linked_by_section.get(number, [])
        )
        section_blocks.append(
            f'<section id="{section_id}" class="lease-section">'
            f'<p class="section-heading"><strong>Section {html_lib.escape(number)}. '
            f'{html_lib.escape(str(section.get("title", "")))}.</strong></p>'
            f'{paragraphs}{linked_paragraphs}</section>'
        )

    focus_id = "section-" + re.sub(r"[^A-Za-z0-9_-]", "-", str(focus_section))
    scroll_script = ""
    if focus_section:
        scroll_script = (
            "<script>setTimeout(function(){var e=document.getElementById('" + focus_id +
            "');if(e){e.scrollIntoView({behavior:'smooth',block:'start'});}},100);</script>"
        )

    return f"""
    <!doctype html>
    <html><head><style>
      * {{ box-sizing: border-box; }}
      body {{ margin: 0; padding: 16px; background: #cfd3d8; color: #111; font-family: 'Times New Roman', Times, serif; }}
      .preview-label {{ position: sticky; top: 0; z-index: 5; margin: -16px -16px 14px; padding: 9px 15px; background: #1a2332; color: #e6edf3; font: 700 12px Arial, sans-serif; letter-spacing: .6px; }}
      .page {{ width: 100%; max-width: 780px; min-height: 990px; margin: 0 auto 18px; padding: 58px 62px; background: #fff; box-shadow: 0 2px 12px rgba(0,0,0,.25); }}
      h1 {{ text-align: center; font-size: 19px; margin: 0 0 5px; }}
      h2 {{ text-align: center; font-size: 14px; margin: 0 0 24px; }}
      table {{ width: 100%; border-collapse: collapse; font-size: 10.5px; }}
      th, td {{ border: 1px solid #666; padding: 6px 7px; vertical-align: top; line-height: 1.35; }}
      th {{ width: 31%; text-align: left; background: #f3f3f3; }}
      .empty {{ text-align: center; color: #777; padding: 20px; }}
      .lease-section {{ margin: 0 0 13px; scroll-margin-top: 42px; }}
      .lease-section p {{ font-size: 10.5px; line-height: 1.45; text-align: justify; margin: 0 0 7px; }}
      .section-heading {{ text-align: left !important; margin-top: 12px !important; }}
      .note {{ text-align: center; color: #555; font: 10px Arial, sans-serif; margin-bottom: 10px; }}
    </style></head><body>
      <div class="preview-label">LIVE DRAFT PREVIEW — CURRENT SELECTIONS</div>
      <div class="note">No drafting highlights are added. Word export uses the original layout with the selected content.</div>
      <div class="page"><h1>LEASE AGREEMENT</h1><h2>KEY PROVISIONS SUMMARY</h2><table>{summary_rows}</table></div>
      <div class="page">{''.join(section_blocks)}</div>
      {scroll_script}
    </body></html>
    """


def _lb_money(value):
    try:
        return f"${float(value):,.2f}"
    except (ValueError, TypeError):
        return ""


def _lb_parse_money(value):
    try:
        return float(str(value).replace("$", "").replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def _lb_money_or_none(value):
    """Return a float for anything that reads as currency, else None."""
    text = str(value).strip().replace("$", "").replace(",", "")
    if text in {"", "nan", "NaN", "None", "-"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    try:
        number = float(text)
    except (ValueError, TypeError):
        return None
    return -number if negative else number


def _lb_rent_cell(value):
    """Normalize one rent cell: currency-looking values format as $x,xxx.xx, everything else passes through."""
    number = _lb_money_or_none(value)
    if number is not None:
        return _lb_money(number)
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none"} else text


RENT_SHEET_NAME = "Rent Schedule"
RENT_COLUMNS = ["Period", "Monthly Rent", "Annual Rent"]


def _lb_blank_rent_rows(periods):
    return [{"Period": period, "Monthly Rent": "", "Annual Rent": ""} for period in periods]


def _lb_default_rent_schedules():
    """Empty scaffold so the downloaded workbook always has rows to fill in."""
    return {
        "base": _lb_blank_rent_rows([f"Year {year}" for year in range(1, 11)]),
        "options": _lb_blank_rent_rows([f"Option 1 — Year {year}" for year in range(1, 6)]),
    }


def _lb_normalize_rent_rows(rows):
    clean = []
    for row in rows or []:
        period = str(row.get("Period", "")).strip()
        monthly = _lb_rent_cell(row.get("Monthly Rent", ""))
        annual = _lb_rent_cell(row.get("Annual Rent", ""))
        if not period and not monthly and not annual:
            continue
        clean.append({"Period": period, "Monthly Rent": monthly, "Annual Rent": annual})
    return clean


def _lb_rent_is_absent(monthly, annual):
    """A period with blank or zero rent does not exist.

    The upload sheet uses 0 as a placeholder for options a deal never granted,
    so a zero row is dropped exactly like an empty one.
    """
    for value in (monthly, annual):
        number = _lb_money_or_none(value)
        if number is None:
            if str(value).strip():
                return False  # Text such as "Free" is a real, intentional entry.
        elif number != 0:
            return False
    return True


def _lb_split_option_rows(option_rows):
    """Group flat option rows back into blocks keyed by their option label.

    'Option 1 — Year 2' carries both the block and the year, so the flat list the
    Word builder consumes can be rebuilt into the stacked sheet layout.
    """
    blocks = []
    index = {}
    for row in option_rows or []:
        period = str(row.get("Period", "")).strip()
        match = re.match(r"^\s*(Option\s*\d*)\s*[—\-–:]\s*(.+)$", period, flags=re.IGNORECASE)
        if match:
            label = re.sub(r"\s+", " ", match.group(1)).title()
            year = match.group(2).strip()
        else:
            label, year = "Option 1", period
        if label not in index:
            index[label] = {"label": label, "rows": []}
            blocks.append(index[label])
        index[label]["rows"].append({
            "Period": year,
            "Monthly Rent": row.get("Monthly Rent", ""),
            "Annual Rent": row.get("Annual Rent", ""),
        })
    return blocks


def _lb_export_rent_schedule_xlsx(rent_state):
    """Write the rent schedule in the upload layout.

    One sheet: the base term in columns A-C under a Term header, and each option
    as its own block in columns F-H separated by a blank row.
    """
    rent_state = rent_state or {}
    base_rows = _lb_normalize_rent_rows(rent_state.get("base"))
    option_blocks = _lb_split_option_rows(_lb_normalize_rent_rows(rent_state.get("options")))
    if not base_rows:
        base_rows = _lb_blank_rent_rows([f"Year {year}" for year in range(1, 11)])
    if not option_blocks:
        option_blocks = [
            {"label": "Option 1", "rows": _lb_blank_rent_rows([f"Year {year}" for year in range(1, 6)])},
            {"label": "Option 2", "rows": _lb_blank_rent_rows([f"Year {year}" for year in range(1, 6)])},
        ]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = RENT_SHEET_NAME
    money_format = '"$"#,##0.00'

    def write_cell(row_index, column_index, value, is_money=False):
        cell = worksheet.cell(row=row_index, column=column_index)
        if is_money:
            number = _lb_money_or_none(value)
            cell.value = number if number is not None else (str(value).strip() or None)
            if number is not None:
                cell.number_format = money_format
        else:
            cell.value = value
        return cell

    header_font = _XLFont(bold=True)
    for column, title in ((1, "Term"), (2, "Monthly Amount"), (3, "Annual Amount")):
        write_cell(1, column, title).font = header_font
    for offset, row in enumerate(base_rows, start=2):
        write_cell(offset, 1, row.get("Period", ""))
        write_cell(offset, 2, row.get("Monthly Rent", ""), is_money=True)
        write_cell(offset, 3, row.get("Annual Rent", ""), is_money=True)

    cursor = 1
    for block in option_blocks:
        for column, title in ((6, block["label"]), (7, "Monthly Amount"), (8, "Annual Amount")):
            write_cell(cursor, column, title).font = header_font
        cursor += 1
        for row in block["rows"]:
            write_cell(cursor, 6, row.get("Period", ""))
            write_cell(cursor, 7, row.get("Monthly Rent", ""), is_money=True)
            write_cell(cursor, 8, row.get("Annual Rent", ""), is_money=True)
            cursor += 1
        cursor += 1  # Blank spacer row between option blocks.

    for column, width in (("A", 16), ("B", 18), ("C", 18), ("D", 3), ("E", 3), ("F", 16), ("G", 18), ("H", 18)):
        worksheet.column_dimensions[column].width = width
    worksheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _lb_read_rent_block(grid, start_row, start_col, stop_labels):
    """Read Period/Monthly/Annual down from a header cell until the block ends."""
    rows = []
    row_index = start_row + 1
    while row_index < len(grid):
        line = grid[row_index]
        period = str(line[start_col]).strip() if start_col < len(line) else ""
        monthly = line[start_col + 1] if start_col + 1 < len(line) else ""
        annual = line[start_col + 2] if start_col + 2 < len(line) else ""
        if not period and not str(monthly).strip() and not str(annual).strip():
            break  # Blank row closes the block.
        if period.strip().lower() in stop_labels:
            break  # Next block's header closes this one.
        rows.append({"Period": period, "Monthly Rent": monthly, "Annual Rent": annual})
        row_index += 1
    return rows


def _lb_import_rent_schedule_xlsx(uploaded_file):
    """Read an uploaded rent workbook into {'base': [...], 'options': [...]}.

    Primary layout is the upload sheet: a Term block in columns A-C and any
    number of stacked Option blocks elsewhere on the same sheet. The older
    two-sheet Base Rent / Option Rent workbook is still accepted so previously
    downloaded files keep working. Blank or zero periods are dropped, and an
    option whose every year is blank or zero is treated as not granted.
    """
    uploaded_file.seek(0)
    workbook = openpyxl.load_workbook(uploaded_file, data_only=True)

    base_rows = []
    option_rows = []
    found_block = False

    for worksheet in workbook.worksheets:
        grid = [
            ["" if value is None else value for value in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
        if not grid:
            continue
        width = max(len(row) for row in grid)
        grid = [list(row) + [""] * (width - len(row)) for row in grid]

        # Locate every block header on this sheet before reading any of them, so
        # a block knows where the next one starts.
        base_headers, option_headers = [], []
        for row_index, line in enumerate(grid):
            for col_index, value in enumerate(line):
                label = re.sub(r"\s+", " ", str(value)).strip()
                if not label:
                    continue
                if label.lower() in {"term", "period", "base rent", "base term"}:
                    base_headers.append((row_index, col_index))
                elif re.fullmatch(r"option\s*\d*", label, flags=re.IGNORECASE):
                    option_headers.append((row_index, col_index, label.title()))

        stop_labels = {"term", "period", "base rent", "base term"}
        stop_labels |= {label.lower() for _, _, label in option_headers}

        # A legacy "Option Rent" sheet also heads its table with Period, so the
        # sheet name decides where a generic header's rows belong.
        sheet_is_option = "option" in str(worksheet.title).strip().lower()
        for row_index, col_index in base_headers:
            found_block = True
            rows = _lb_read_rent_block(grid, row_index, col_index, stop_labels)
            (option_rows if sheet_is_option else base_rows).extend(rows)

        for row_index, col_index, label in option_headers:
            found_block = True
            block = _lb_read_rent_block(grid, row_index, col_index, stop_labels)
            live = [
                row for row in block
                if not _lb_rent_is_absent(row.get("Monthly Rent", ""), row.get("Annual Rent", ""))
            ]
            if not live:
                continue  # Every year blank or zero: this option was not granted.
            for row in live:
                year = str(row.get("Period", "")).strip()
                row["Period"] = f"{label} — {year}" if year else label
            option_rows.extend(live)

        # Fall back to the older two-sheet workbook, which has no Term header.
        if not base_headers and not option_headers:
            title = str(worksheet.title).strip().lower()
            legacy = _lb_read_rent_block(grid, 0, 0, set())
            if legacy:
                found_block = True
                (option_rows if "option" in title else base_rows).extend(legacy)

    if not found_block:
        raise ValueError(
            "No rent tables found. Expected a 'Term' header above the base years, "
            "and an 'Option 1' header above each option."
        )

    def live_only(rows):
        return [
            row for row in rows
            if not _lb_rent_is_absent(row.get("Monthly Rent", ""), row.get("Annual Rent", ""))
        ]

    return {
        "base": _lb_normalize_rent_rows(live_only(base_rows)),
        "options": _lb_normalize_rent_rows(live_only(option_rows)),
    }


# ---------------------------------------------------------------------------
# Two-mode model.
#
# TEMPLATE mode authors the menu: every provision and every clause choice is
# visible and editable, and "Use" means "checked on by default" for new leases.
# LEASE mode consumes the menu: pick which provisions and sections are used and
# which choice applies, and the preview shows only what is used.
#
# Templates live in the "Lease Builder Templates" sheet, leases in "Saved
# Leases". A lease stores only its selections (bookmark, on/off, chosen value)
# and re-hydrates the alternates from its parent template, which keeps saved
# leases small and lets a template edit propagate to future leases.
# ---------------------------------------------------------------------------

LEASE_TEMPLATE_SHEET = "Lease Builder Templates"
LEASE_FORMAT_SHEET = "Lease Format Profiles"
LEASE_DOC_SHEET = "Lease Documents"
SAVED_LEASE_SHEET = "Saved Leases"
LB_NEW_TEMPLATE = "➕ New document"


# ---------------------------------------------------------------------------
# Where documents live.
#
# Documents used to be one gzipped JSON blob in cell A1 of the "Lease Documents"
# tab. That was chosen because Streamlit Cloud rebuilds its filesystem on every
# redeploy, so the Sheet was the only durable store the app could write to. It
# cost us a 50,000 character ceiling and left every document in one opaque blob
# with no history.
#
# They now live as one JSON file each in a private GitHub data repo. Each save
# is a commit, so a lease has a real diff and a real history.
#
# The Sheet is still read when the repo holds nothing, which covers both the
# one-time migration and the case where the token is missing or expired. It is
# never written to again — leaving it frozen means a bad migration can always
# be walked back to a known good copy.
# ---------------------------------------------------------------------------

@st.cache_resource(ttl=300)
def get_lease_store():
    """The GitHub-backed document store, or None if it is not configured."""
    try:
        return lstore.build_store(secrets=st.secrets)
    except Exception as exc:
        print(f"Lease store unavailable: {type(exc).__name__}: {exc}")
        return None


@st.cache_data(ttl=300)
def _lb_space_records():
    """Leasable spaces from the tenancy workbook, ready for the picker.

    On Streamlit Cloud the workbook is whatever was last committed to the repo,
    so this is a snapshot rather than live data. The Reload button clears the
    cache; a genuinely stale workbook needs a new commit.
    """
    try:
        _, _, summaries = load_tenancy()
        return lsp.space_records(summaries)
    except Exception as exc:
        print(f"Tenancy spaces unavailable: {type(exc).__name__}: {exc}")
        return []


def _lb_load_documents():
    """Every saved document, and where it came from.

    Returns (documents, source) where source is "repo", "sheet" or "empty".
    The caller needs the source because a save has to go somewhere sensible
    and because falling back to a read-only Sheet is worth saying out loud.
    """
    store = get_lease_store()
    if store is not None:
        try:
            documents = ld.normalize_store(store.load_all())
            if documents:
                return documents, "repo"
        except lstore.StoreError as exc:
            st.warning(f"Could not read the lease data repo — {exc}")
        except Exception as exc:
            st.warning(f"Could not read the lease data repo — {type(exc).__name__}: {exc}")

    legacy = ld.normalize_store(_read_gsheet_config(LEASE_DOC_SHEET))
    if legacy:
        return legacy, "sheet"
    return {}, "empty"


def _lb_save_document(name, payload):
    """Write one document. Returns (ok, message)."""
    store = get_lease_store()
    if store is None:
        return False, (
            "The lease data repo is not configured, so there is nowhere durable "
            "to save. Add a [lease_data] section to secrets and reload."
        )
    try:
        store.save_document(str(name).strip(), payload)
        return True, ""
    except lstore.StoreError as exc:
        return False, str(exc)
    except Exception as exc:
        return False, f"{type(exc).__name__}: {exc}"


def _lb_delete_document(name):
    """Remove one document. Returns (ok, message)."""
    store = get_lease_store()
    if store is None:
        return False, "The lease data repo is not configured."
    try:
        if store.delete_document(str(name).strip()):
            return True, ""
        return False, "That document was not found in the repo."
    except Exception as exc:
        return False, f"{type(exc).__name__}: {exc}"


def _lb_migrate_sheet_to_repo(documents):
    """Copy the Sheet's documents into the repo, once.

    Every document must land before this reports success. A partial migration
    that looked complete would be the one failure here that loses a lease, so
    anything that does not write is named individually and the Sheet is left
    untouched as the fallback.
    """
    store = get_lease_store()
    if store is None or not documents:
        return False, []
    failures = []
    for name, document in documents.items():
        ok, message = _lb_save_document(name, document)
        if not ok:
            failures.append(f"{name} — {message}")
    return (not failures), failures


def _lb_compact_sections(sections, section_state):
    """Store a section's choice, and its text only when it departs from template language."""
    compact = {}
    for section in sections:
        number = str(section["number"])
        config = section_state.get(number, {})
        entry = {
            "include": bool(config.get("include", True)),
            "choice": config.get("choice", "Template language"),
        }
        # Unchanged template text is what previously blew past the cell limit.
        if entry["choice"] != "Template language":
            entry["text"] = config.get("text", "")
        compact[number] = entry
    return compact


def _lb_apply_document(draft_state, document, profiles=None):
    """Load a saved document into draft state.

    Everything the document needs is in the document: its provisions carry their
    own alternates, its sections carry their own chosen language. Nothing is
    re-hydrated from a parent, because there is no parent any more.
    """
    resolved = ld.normalize_document(document)
    if resolved["key_provisions"]:
        draft_state["key_provisions"] = [dict(row) for row in resolved["key_provisions"]]
    for number, config in resolved["sections"].items():
        if number in draft_state["sections"]:
            draft_state["sections"][number].update(config)
    if resolved["rent_schedules"]:
        draft_state["rent_schedules"] = resolved["rent_schedules"]
    profile_name = lf.resolve_profile_name(profiles, resolved["format_profile"])
    draft_state["format_profile"] = profile_name
    draft_state["formatting"] = lf.profile_settings(profiles, profile_name)
    draft_state["copied_from"] = resolved["copied_from"]
    return draft_state


def _lb_section_variant_texts(selected_label, section_number, built_in_library, saved_clause_library):
    """Every approved text for one section: template language plus all saved variants."""
    texts = []
    for variant in built_in_library.get(section_number, {}).get("variants", []):
        texts.append(str(variant.get("text", "")))
    if isinstance(saved_clause_library, dict):
        for variant in saved_clause_library.get(selected_label, {}).get(section_number, []):
            texts.append(str(variant.get("text", "")))
    return texts


def _lb_is_off_menu(text, approved_texts):
    """True when the drafted language matches nothing the template offers."""
    def squash(value):
        return re.sub(r"\s+", " ", str(value or "")).strip().lower()

    candidate = squash(text)
    if not candidate:
        return False
    return all(candidate != squash(approved) for approved in approved_texts)


def _lb_token_preview_html(text, names):
    """Clause text with [KP:...] tokens coloured: green when they resolve, red when not."""
    known = {str(name).strip().lower() for name in names if str(name).strip()}
    pieces = []
    position = 0
    source = str(text or "")
    for match in KP_TOKEN_RE.finditer(source):
        pieces.append(html_lib.escape(source[position:match.start()]))
        resolves = match.group(1).strip().lower() in known
        colour = f"#{KP_LINK_COLOR}" if resolves else "#d13438"
        title = "Resolves to this provision's value" if resolves else "No provision with this name"
        pieces.append(
            f'<span style="color:{colour};font-weight:600;text-decoration:underline;" '
            f'title="{title}">{html_lib.escape(match.group(0))}</span>'
        )
        position = match.end()
    pieces.append(html_lib.escape(source[position:]))
    body = "".join(pieces).replace("\n", "<br>")
    return (
        '<div style="border:1px solid rgba(128,128,128,.35);border-radius:.5rem;'
        'padding:.75rem;font-size:.85rem;line-height:1.5;max-height:260px;overflow-y:auto;'
        f'white-space:pre-wrap;">{body}</div>'
    )


def _lb_token_citations(sections, draft_state, include_only=True):
    """Map provision name -> section numbers whose clause text cites it via KP:.

    Linking is derived from the drafted language rather than set by hand, so a
    provision cited in three sections reports all three and nothing has to be
    kept in sync.
    """
    names = [
        str(row.get("Field", "")).strip()
        for row in draft_state.get("key_provisions", [])
        if str(row.get("Field", "")).strip()
    ]
    citations = {}
    for section in sections:
        number = str(section["number"])
        config = draft_state["sections"].get(number, {})
        if include_only and not bool(config.get("include", True)):
            continue
        text = str(config.get("text", section.get("text", "")))
        for name in find_kp_references(text, names):
            citations.setdefault(name, []).append(number)
    return citations


def _lb_apply_space(draft_state, sections, space):
    """A copy of the draft and the section list with [Space:...] tokens resolved.

    Resolution happens on the way out to Word, never in what is stored. The
    saved document keeps the tokens, so re-picking the space re-renders every
    number instead of baking one unit's figures into the template forever.

    An unresolvable token is left in place rather than dropped — a lease that
    reads "[Space:Sqft]" is obviously wrong, whereas one reading "Approximately
    Total Gross Square Feet" reads as finished.
    """
    if not space:
        return draft_state, sections
    resolved_state = dict(draft_state)
    resolved_state["key_provisions"] = lsp.resolve_provisions(
        draft_state.get("key_provisions", []), space
    )
    resolved_state["sections"] = {
        number: ({**config, "text": lsp.resolve(config["text"], space)}
                 if isinstance(config, dict) and config.get("text") else dict(config or {}))
        for number, config in (draft_state.get("sections") or {}).items()
    }
    resolved_sections = [
        {**section, "text": lsp.resolve(section.get("text", ""), space)}
        for section in (sections or [])
    ]
    return resolved_state, resolved_sections


def _lb_build_current_word(template_path, draft_name, clean_notes, sections, draft_state,
                           force_include_all=False, token_report=None, space=None):
    draft_state, sections = _lb_apply_space(draft_state, sections, space)
    section_labels = {
        f"Section {section['number']} — {section['title']}": str(section['number'])
        for section in sections
    }
    section_choices = {}
    for section in sections:
        number = str(section["number"])
        config = draft_state["sections"][number]
        configured_text = str(config.get("text", section["text"]))
        base_text = str(section["text"])
        replacement_text = "" if configured_text.strip() == base_text.strip() else _lb_section_body(number, configured_text)
        section_choices[number] = {
            # Template mode proofreads the whole menu, so nothing is filtered out.
            "include": True if force_include_all else bool(config.get("include", True)),
            "title": section["title"],
            "replacement_text": replacement_text,
        }

    def is_used(row):
        return True if force_include_all else bool(row.get("Include"))

    bookmark_values = {
        str(row["Bookmark"]): row.get("Value", "")
        for row in draft_state["key_provisions"]
    }
    included_bookmarks = {
        str(row["Bookmark"])
        for row in draft_state["key_provisions"] if is_used(row)
    }
    linked_provisions = [
        {
            "bookmark": str(row["Bookmark"]),
            "field": row.get("Field", "Key Provision"),
            "value": row.get("Value", ""),
            "include": is_used(row),
            "section": section_labels.get(str(row.get("Section", "")), str(row.get("Section", ""))),
        }
        for row in draft_state["key_provisions"] if bool(row.get("Link"))
    ]
    custom_provisions = [
        {
            "field": row.get("Field", "Key Provision"),
            "value": row.get("Value", ""),
            "include": is_used(row),
            "section": section_labels.get(str(row.get("Section", "")), str(row.get("Section", ""))),
        }
        for row in draft_state["key_provisions"]
        if str(row.get("Bookmark", "")).startswith("Custom_")
    ]
    # The provision list is app-owned: the summary table is rebuilt from these
    # rows. Linking is derived from KP: tokens in the clause text, so the row
    # points at the first section that actually cites the provision.
    citations = _lb_token_citations(sections, draft_state)
    key_provision_rows = [
        {
            "bookmark": str(row.get("Bookmark", "")),
            "field": row.get("Field", "Key Provision"),
            "value": row.get("Value", ""),
            "include": is_used(row),
            "link": bool(citations.get(str(row.get("Field", "")))),
            "section": (citations.get(str(row.get("Field", ""))) or [""])[0],
        }
        for row in draft_state["key_provisions"]
    ]
    return build_word_document(
        template_path,
        section_choices=section_choices,
        bookmark_values=bookmark_values,
        included_bookmarks=included_bookmarks,
        linked_provisions=linked_provisions,
        custom_provisions=custom_provisions,
        rent_schedules=draft_state.get("rent_schedules"),
        additional_choices=[],
        clean_drafting_notes=clean_notes,
        document_title=draft_name,
        key_provision_rows=key_provision_rows,
        token_report=token_report,
    )


def _lb_build_rules_word(draft_state, sections, settings, space=None):
    """Generate the lease with the rules-based renderer — no base .docx.

    Clause text comes from the draft's own section state, so edits made in the
    tab are reflected; only the document's construction differs from the legacy
    path. Steps 4 and 6-8 of the spec (Key Provisions table, rent tables,
    signatures, exhibits) are not built yet, so those appear as markers.
    """
    from io import BytesIO

    draft_state, sections = _lb_apply_space(draft_state, sections, space)
    values = {
        str(row.get("Field", "")): str(row.get("Value", "") or "")
        for row in draft_state["key_provisions"]
        if bool(row.get("Include", True))
    }
    payload = []
    for section in sections:
        number = str(section["number"])
        config = draft_state["sections"].get(number, {})
        if not config.get("include", True):
            continue
        payload.append({
            "number": number,
            "title": section.get("title", ""),
            "body": str(config.get("text", section.get("text", ""))),
        })

    document, renderer = lr.render_lease(payload, settings, values)
    buffer = BytesIO()
    document.save(buffer)
    return buffer.getvalue(), renderer


@st.cache_data(show_spinner=False, max_entries=3)
def _lb_render_word_pages(word_bytes):
    """Render the generated DOCX to page images for a faithful live preview.

    Returns (page_images, {section_number: page_index}) so a preview can jump to
    a section in the real rendering instead of an HTML approximation of it.
    """
    office = shutil.which("libreoffice") or shutil.which("soffice")
    if not office or not HAS_PYMUPDF:
        return [], {}
    with tempfile.TemporaryDirectory(prefix="msp-lease-preview-") as temp_dir:
        temp_path = Path(temp_dir)
        docx_path = temp_path / "lease-preview.docx"
        docx_path.write_bytes(word_bytes)
        profile_path = temp_path / "lo-profile"
        command = [
            office,
            "--headless",
            f"-env:UserInstallation=file://{profile_path}",
            "--convert-to", "pdf",
            "--outdir", str(temp_path),
            str(docx_path),
        ]
        result = subprocess.run(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=75,
            check=False,
            env={**os.environ, "HOME": temp_dir},
        )
        pdf_path = temp_path / "lease-preview.pdf"
        if result.returncode != 0 or not pdf_path.exists():
            print("Lease preview conversion failed:", result.stderr.decode("utf-8", errors="ignore")[-500:])
            return [], {}
        document = fitz.open(pdf_path)
        pages = []
        section_pages = {}
        matrix = fitz.Matrix(1.05, 1.05)
        heading_re = re.compile(r"^\s*Section\s+(\d+(?:\.\d+)?)\s*\.", re.IGNORECASE | re.MULTILINE)
        for index, page in enumerate(document):
            pixmap = page.get_pixmap(matrix=matrix, alpha=False)
            pages.append(pixmap.tobytes("png"))
            # Record where each section starts so a preview can jump to it.
            # First occurrence wins: a later cross-reference to "Section 5."
            # must not drag the jump away from the section itself.
            for match in heading_re.finditer(page.get_text()):
                section_pages.setdefault(match.group(1), index)
        document.close()
        return pages, section_pages


def _lb_reset_editor_widget_state():
    prefixes = (
        "lb_kp_editor_", "lb_section_editor_", "lb_section_choice_",
        "lb_section_text_", "lb_new_choice_name_",
        "rent_base_grid", "rent_option_grid", "lb_offmenu_", "lb_fmt_",
    )
    for key in list(st.session_state.keys()):
        if key.startswith(prefixes):
            del st.session_state[key]


def _lb_render_formatting_form(current_settings, editable=True, expanded=False):
    """Document Formatting panel — the page-setup and typography rules the
    template-free renderer follows.

    Pure in, pure out: takes a settings dict and returns the edited one. The
    caller decides where it lives, which is what lets the same panel edit a
    named format profile rather than one template's private copy.
    """
    settings = lf.normalize_settings(current_settings)

    with st.expander("📐 Document Formatting", expanded=expanded):
        if not editable:
            st.caption(
                "Formatting comes from the template's format profile. Switch to "
                "Edit Lease Template to change it."
            )
        st.caption(
            "Defaults measured from the master lease. These rules drive "
            "generation directly — there is no base Word file behind them."
        )

        def number(label, key, step=0.1, fmt="%.2f", help_text=None):
            low, high = lf._NUMERIC_BOUNDS.get(key, (0.0, 100.0))
            return st.number_input(
                label,
                min_value=float(low),
                max_value=float(high),
                value=float(settings[key]),
                step=step,
                format=fmt,
                key=f"lb_fmt_{key}",
                help=help_text,
                disabled=not editable,
            )

        def choice(label, key, options, help_text=None):
            current = settings[key]
            return st.selectbox(
                label,
                options,
                index=options.index(current) if current in options else 0,
                key=f"lb_fmt_{key}",
                help=help_text,
                disabled=not editable,
            )

        def flag(label, key, help_text=None):
            return st.checkbox(
                label,
                value=bool(settings[key]),
                key=f"lb_fmt_{key}",
                help=help_text,
                disabled=not editable,
            )

        def text(label, key, placeholder="", help_text=None):
            return st.text_input(
                label,
                value=str(settings[key]),
                key=f"lb_fmt_{key}",
                placeholder=placeholder,
                help=help_text,
                disabled=not editable,
            )

        updated = dict(settings)
        page_tab, body_tab, kp_tab, section_tab, rent_tab, back_tab = st.tabs(
            ["Page", "Body Text", "Key Provisions", "Sections", "Rent Table", "Signatures & Exhibits"]
        )

        with page_tab:
            col1, col2 = st.columns(2)
            with col1:
                updated["page_size"] = choice("Page size", "page_size", list(lf.PAGE_SIZES))
                updated["margin_top_in"] = number("Top margin (in)", "margin_top_in")
                updated["margin_bottom_in"] = number("Bottom margin (in)", "margin_bottom_in")
            with col2:
                updated["margin_left_in"] = number("Left margin (in)", "margin_left_in")
                updated["margin_right_in"] = number("Right margin (in)", "margin_right_in")
            st.markdown("**Footer**")
            foot1, foot2, foot3 = st.columns([2, 1, 1])
            with foot1:
                updated["footer_doc_id"] = text(
                    "Counsel document ID", "footer_doc_id",
                    placeholder="4928-4211-2690, v. 2",
                    help_text="Prints at the footer left. Per-template, not per-lease.",
                )
            with foot2:
                updated["footer_font_size_pt"] = number("Footer size (pt)", "footer_font_size_pt", step=0.5, fmt="%.1f")
            with foot3:
                updated["page_number_position"] = choice(
                    "Page number", "page_number_position", lf.PAGE_NUMBER_POSITIONS
                )
            st.markdown("**Title block**")
            title1, title2 = st.columns(2)
            with title1:
                updated["title_text"] = text("Document title", "title_text")
                updated["key_provisions_title"] = text("Summary title", "key_provisions_title")
            with title2:
                updated["title_size_pt"] = number("Title size (pt)", "title_size_pt", step=0.5, fmt="%.1f")
                updated["title_bold"] = flag("Title bold", "title_bold")
                updated["title_underline"] = flag("Title underlined", "title_underline")

        with body_tab:
            col1, col2 = st.columns(2)
            with col1:
                updated["body_font"] = choice("Body font", "body_font", lf.BODY_FONTS)
                updated["body_size_pt"] = number("Body size (pt)", "body_size_pt", step=0.5, fmt="%.1f")
                updated["body_alignment"] = choice("Alignment", "body_alignment", lf.ALIGNMENTS)
            with col2:
                updated["body_line_spacing"] = number("Line spacing", "body_line_spacing", step=0.05)
                updated["space_after_pt"] = number("Space after paragraph (pt)", "space_after_pt", step=1.0, fmt="%.1f")
                updated["first_line_indent_in"] = number("First-line indent (in)", "first_line_indent_in")

        with kp_tab:
            col1, col2 = st.columns(2)
            with col1:
                updated["kp_label_width_in"] = number("Label column (in)", "kp_label_width_in")
                updated["kp_value_width_in"] = number("Value column (in)", "kp_value_width_in")
                updated["kp_cell_padding_pt"] = number("Cell padding (pt)", "kp_cell_padding_pt", step=0.5, fmt="%.1f")
                updated["kp_split_left_width_in"] = number(
                    "Split — Landlord half (in)", "kp_split_left_width_in",
                    help_text="Together the two halves should equal the value column.",
                )
                updated["kp_split_right_width_in"] = number("Split — Tenant half (in)", "kp_split_right_width_in")
            with col2:
                updated["kp_borders"] = flag("Bordered table", "kp_borders")
                updated["kp_label_bold"] = flag("Bold labels", "kp_label_bold")
                updated["kp_link_underline"] = flag("Underline cross-references", "kp_link_underline")
                updated["kp_link_color"] = st.color_picker(
                    "Cross-reference color",
                    value=f"#{settings['kp_link_color']}",
                    key="lb_fmt_kp_link_color",
                    disabled=not editable,
                )
            updated["kp_split_fields"] = st.text_input(
                "Provisions that split Landlord | Tenant",
                value=", ".join(settings["kp_split_fields"]),
                key="lb_fmt_kp_split_fields",
                help="Comma-separated. These rows divide the value into two columns.",
                disabled=not editable,
            )

        with section_tab:
            col1, col2 = st.columns(2)
            with col1:
                updated["section_word"] = text("Heading word", "section_word", help_text='Prints as "Section 5." before the title.')
                updated["section_first_line_indent_in"] = number(
                    "Heading first-line indent (in)", "section_first_line_indent_in"
                )
                updated["section_tab_stop_in"] = number("Tab after number (in)", "section_tab_stop_in")
                updated["section_space_before_pt"] = number("Space before section (pt)", "section_space_before_pt", step=1.0, fmt="%.1f")
                updated["section_heading_bold"] = flag("Bold run-in heading", "section_heading_bold")
            with col2:
                updated["subclause_level1_style"] = choice(
                    "Sub-clause level 1", "subclause_level1_style", lf.SUBCLAUSE_LEVEL1_STYLES
                )
                updated["subclause_level1_indent_in"] = number("Level 1 indent (in)", "subclause_level1_indent_in")
                updated["subclause_level1_hanging_in"] = number("Level 1 hanging (in)", "subclause_level1_hanging_in")
                updated["subclause_level2_style"] = choice(
                    "Sub-clause level 2", "subclause_level2_style", lf.SUBCLAUSE_LEVEL2_STYLES,
                    help_text="Runs inline in the paragraph, not as its own list.",
                )
                updated["subclause_lead_in_bold"] = flag("Bold lead-in phrase", "subclause_lead_in_bold")

        with rent_tab:
            col1, col2 = st.columns(2)
            with col1:
                updated["rent_table_label"] = text("Table label", "rent_table_label")
                updated["rent_col_term_width_in"] = number("Term column (in)", "rent_col_term_width_in")
            with col2:
                updated["rent_col_monthly_width_in"] = number("Monthly column (in)", "rent_col_monthly_width_in")
                updated["rent_col_annual_width_in"] = number("Annual column (in)", "rent_col_annual_width_in")
            updated["rent_header_bold"] = flag("Bold header row", "rent_header_bold")
            updated["rent_borders"] = flag("Bordered table", "rent_borders")
            updated["rent_amount_alignment"] = choice(
                "Amount alignment", "rent_amount_alignment", ["right", "left", "center"]
            )

        with back_tab:
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**Signatures**")
                updated["signature_title"] = text("Block title", "signature_title")
                updated["signature_landlord_label"] = text("Landlord label", "signature_landlord_label")
                updated["signature_tenant_label"] = text("Tenant label", "signature_tenant_label")
                updated["signature_keep_together"] = flag("Keep block on one page", "signature_keep_together")
            with col2:
                st.markdown("**Exhibits**")
                updated["exhibit_label_format"] = text(
                    "Label format", "exhibit_label_format",
                    help_text="{letter} is replaced by A, B, C …",
                )
                updated["exhibit_title_size_pt"] = number("Exhibit title size (pt)", "exhibit_title_size_pt", step=0.5, fmt="%.1f")
                updated["exhibit_image_max_width_in"] = number("Max image width (in)", "exhibit_image_max_width_in")
                updated["exhibit_page_break"] = flag("Page break before each exhibit", "exhibit_page_break")

        normalized = lf.normalize_settings(updated) if editable else settings

        for warning in lf.validate_settings(normalized):
            st.warning(warning)

        changed = lf.settings_diff(normalized)
        summary_col, reset_col = st.columns([4, 1])
        summary_col.caption(
            f"{len(changed)} setting{'' if len(changed) == 1 else 's'} differ from the measured defaults."
            if changed else "Matching the measured defaults exactly."
        )
        if editable and reset_col.button("↩︎ Reset", key="lb_fmt_reset", width="stretch"):
            for key in [k for k in st.session_state if k.startswith("lb_fmt_")]:
                del st.session_state[key]
            normalized = lf.default_settings()
            st.rerun()
        if changed:
            with st.popover("Show changes"):
                st.dataframe(
                    pd.DataFrame(
                        [
                            {"Setting": key, "Default": lf.DEFAULTS[key], "This profile": value}
                            for key, value in sorted(changed.items())
                        ]
                    ),
                    hide_index=True,
                    width="stretch",
                )

    return normalized


def _lb_render_document_preview(word_bytes, draft_state=None, preview_sections=None):
    """Full-width, scrollable, zoomable preview of the whole document.

    Collapsible so it can be pushed out of the way, but open by default — a
    preview you have to go find is a preview nobody looks at.
    """
    label = "📄 Document Preview"
    with st.expander(label, expanded=True):
        zoom_col, height_col, note_col = st.columns([2, 2, 4])
        zoom = zoom_col.slider(
            "Zoom", min_value=50, max_value=250, value=100, step=10,
            format="%d%%", key="lb_preview_zoom",
        )
        pane_height = height_col.select_slider(
            "Window height", options=[500, 700, 900, 1200, 1600],
            value=900, key="lb_preview_height",
        )

        rendered_pages, _section_pages = [], {}
        if word_bytes:
            with st.spinner("Rendering the Word document…"):
                rendered_pages, _section_pages = _lb_render_word_pages(word_bytes)

        if rendered_pages:
            note_col.caption(
                f"{len(rendered_pages)} page(s), rendered from the generated Word file. "
                "Numbering, included sections and formatting match the download."
            )
        else:
            note_col.caption(
                "Exact Word rendering is unavailable here, so the content-faithful "
                "fallback preview is shown."
            )

        pane = st.container(height=pane_height, border=True)
        with pane:
            if rendered_pages:
                # Streamlit scales to the container by default; an explicit pixel
                # width is what makes zoom actually zoom.
                for page_number, page_image in enumerate(rendered_pages, start=1):
                    st.image(page_image, caption=f"Page {page_number}",
                             width=int(700 * zoom / 100))
            else:
                st.components.v1.html(
                    _lb_preview_html(
                        (draft_state or {}).get("key_provisions", []),
                        preview_sections or [],
                        "",
                        show_all=False,
                    ),
                    height=max(300, pane_height - 60),
                    scrolling=True,
                )


def _lb_render_section_preview(draft_state, preview_sections, focus_section, word_bytes=None):
    """Preview beside the section editor, opened at the section in hand.

    Uses the same rendered pages as the full preview above — an HTML mock-up
    beside the real thing invites the question of which one is true, and the
    answer must always be "they are the same document".
    """
    pages, section_pages = ([], {})
    if word_bytes:
        pages, section_pages = _lb_render_word_pages(word_bytes)

    if not pages:
        # Only when LibreOffice is unavailable: an approximation, labelled as one.
        st.caption(
            "⚠️ Word rendering unavailable on this server — showing the "
            "approximate HTML preview, which does not reflect the format profile."
        )
        st.components.v1.html(
            _lb_preview_html(draft_state["key_provisions"], preview_sections,
                             focus_section, show_all=False),
            height=860, scrolling=True,
        )
        return

    start = section_pages.get(str(focus_section))
    if start is None:
        start = 0
        st.caption(
            f"Section {focus_section} has no heading in the rendered document yet — "
            "showing from page 1."
            if focus_section else "Select a section on the left to jump to it."
        )
    else:
        st.caption(
            f"Section {focus_section} begins on page {start + 1} of {len(pages)}."
        )

    zoom = st.slider("Zoom", min_value=50, max_value=250, value=100, step=10,
                     format="%d%%", key="lb_section_preview_zoom")
    pane = st.container(height=860, border=True)
    with pane:
        # From the section's page to the end, so reading on past it still works.
        for offset, page_image in enumerate(pages[start:], start=start + 1):
            st.image(page_image, caption=f"Page {offset}",
                     width=int(560 * zoom / 100))


def _lb_render_template_header(working_template, saved_docs, profiles, profile_name,
                               payload_builder, editable=True):
    """The three things that identify what you are editing, at the top.

    1. which template, and how to save it
    2. where its words come from (the extracted JSON, not a Word file)
    3. which format profile styles it

    Rendered into a container reserved at the top of the tab but called late,
    once draft_state exists — otherwise the save buttons could not build a
    payload.
    """
    library = lc.load_content()
    profiles = lf.normalize_profiles(profiles)
    profile_name = lf.resolve_profile_name(profiles, profile_name)
    existing = working_template != LB_NEW_TEMPLATE

    # ---- 1. Template ----------------------------------------------------
    if editable:
        name_col, save_col, saveas_col, delete_col = st.columns([4, 1.2, 1.2, 1.2])
        name_col.markdown(f"#### 📄 {working_template if existing else 'New template'}")
        if save_col.button("💾 Save", type="primary", key="lb_save_template",
                           disabled=not existing, width="stretch"):
            ok, message = _lb_save_document(working_template, payload_builder(profile_name))
            if ok:
                st.session_state["lb_save_as_open"] = False
                st.success(f"Saved: {working_template}")
                st.rerun()
            else:
                st.error(f"Could not save “{working_template}” — {message}")
        if saveas_col.button("💾 Save As…", key="lb_save_template_as", width="stretch"):
            # Opens the rename field rather than saving immediately — Save As
            # without a chance to type a name is just a duplicate.
            st.session_state["lb_save_as_open"] = True
            st.session_state["lb_save_template_as_name"] = (
                working_template if existing else ""
            )
            st.rerun()
        if delete_col.button("🗑 Delete", key="lb_delete_template",
                             disabled=not existing, width="stretch"):
            st.session_state["lb_delete_open"] = True
            st.rerun()

        if st.session_state.get("lb_delete_open") and existing:
            # Two clicks, and the name spelled out in between. A saved lease is
            # the one thing here with no undo short of the repo's history.
            st.warning(f"Delete “{working_template}” permanently?")
            confirm_delete_col, cancel_delete_col, _ = st.columns([1.6, 1.6, 4])
            if confirm_delete_col.button("🗑 Yes, delete", type="primary",
                                         key="lb_delete_confirm", width="stretch"):
                ok, message = _lb_delete_document(working_template)
                st.session_state["lb_delete_open"] = False
                if ok:
                    # The editor is still holding the deleted document's
                    # content. Dropping the picker choice and every per-template
                    # draft stops it being re-saved back into existence.
                    st.session_state.pop("lb_template_edit_choice", None)
                    for _key in [k for k in st.session_state if str(k).startswith("lb_draft_state_")]:
                        st.session_state.pop(_key, None)
                    st.success(f"Deleted: {working_template}")
                    st.rerun()
                else:
                    st.error(f"Could not delete “{working_template}” — {message}")
            if cancel_delete_col.button("✕ Keep it", key="lb_delete_cancel", width="stretch"):
                st.session_state["lb_delete_open"] = False
                st.rerun()

        if st.session_state.get("lb_save_as_open"):
            rename_col, confirm_col, cancel_col = st.columns([4, 1.2, 1.2])
            save_as_name = rename_col.text_input(
                "New template name", key="lb_save_template_as_name",
                placeholder="e.g., MSP NNN Retail — Restaurant",
                label_visibility="collapsed",
            )
            if confirm_col.button("✔ Save copy", type="primary",
                                  key="lb_save_as_confirm", width="stretch"):
                candidate = save_as_name.strip()
                if not candidate:
                    st.warning("Enter a template name first.")
                elif candidate == working_template:
                    st.warning("That is the current name — use Save to overwrite it.")
                elif candidate in saved_docs:
                    st.warning("A document with that name already exists.")
                else:
                    ok, message = _lb_save_document(candidate, payload_builder(profile_name))
                    if ok:
                        st.session_state["lb_save_as_open"] = False
                        # Switch the editor to the copy, which is what you almost
                        # always want after saving one. It has to go through a
                        # plain key: Streamlit forbids assigning to a widget's own
                        # key once that widget has been created this run, and the
                        # picker was built at the top of the tab.
                        st.session_state["lb_pending_template_choice"] = candidate
                        st.session_state.pop("lb_pending_copy", None)
                        st.success(f"Saved: {candidate}")
                        st.rerun()
                    else:
                        st.error(f"Could not save “{candidate}” — {message}")
            if cancel_col.button("✕ Cancel", key="lb_save_as_cancel", width="stretch"):
                st.session_state["lb_save_as_open"] = False
                st.rerun()

        if not existing:
            st.caption("New document — use Save As to name and create it.")
    else:
        st.markdown(f"#### 📄 {working_template if existing else 'New template'}")

    # ---- 2. Content -----------------------------------------------------
    sections = library.get("sections", [])
    provisions = library.get("key_provisions", [])
    extracted = str(library.get("extracted_at", ""))[:10]
    content_col, extract_col = st.columns([4, 1.4])
    if sections:
        revisions = library.get("tracked_changes_resolved") or {}
        note = (
            f" · {revisions.get('insertions', 0)} tracked edits resolved"
            if revisions.get("insertions") else ""
        )
        content_col.caption(
            f"**Content** · `lease_content.json` · {len(sections)} sections · "
            f"{len(provisions)} key provisions · extracted {extracted}{note}"
        )
    else:
        content_col.warning(
            "No extracted content found. Run `python lease_content.py --extract` "
            "to read the master lease into JSON."
        )
    if editable and extract_col.button("🔄 Re-extract", key="lb_reextract", width="stretch"):
        try:
            written = lc.write_content(lc.extract_from_docx())
            st.success(f"Re-extracted into {Path(written).name}.")
            st.rerun()
        except Exception as exc:
            st.error(f"Could not re-extract: {type(exc).__name__}: {exc}")

    # ---- 3. Formatting --------------------------------------------------
    settings = lf.profile_settings(profiles, profile_name)
    profile_col, summary_col = st.columns([2, 4])
    chosen = profile_col.selectbox(
        "Format profile", sorted(profiles),
        index=sorted(profiles).index(profile_name),
        key="lb_format_profile", label_visibility="collapsed", disabled=not editable,
    )
    if chosen != profile_name:
        profile_name = chosen
        settings = lf.profile_settings(profiles, profile_name)
        for key in [k for k in st.session_state if k.startswith("lb_fmt_")]:
            del st.session_state[key]
    summary_col.caption(f"**Formatting** · {lf.describe_settings(settings)}")

    edited = _lb_render_formatting_form(settings, editable=editable)
    if editable and edited != settings:
        fmt_save, fmt_name, fmt_saveas = st.columns([1.4, 2.6, 1.4])
        if fmt_save.button("💾 Save profile", key="lb_fmt_save", width="stretch"):
            updated = dict(profiles)
            updated[profile_name] = lf.settings_diff(edited)
            if _write_gsheet_config(LEASE_FORMAT_SHEET, updated):
                st.success(f"Saved format profile: {profile_name}")
                st.rerun()
            else:
                st.error("Could not save the format profile to Google Sheets.")
        new_profile = fmt_name.text_input(
            "New profile name", key="lb_fmt_profile_name",
            placeholder="e.g., MSP House Style — Compact", label_visibility="collapsed",
        )
        if fmt_saveas.button("💾 Save as new", key="lb_fmt_save_as", width="stretch"):
            if not new_profile.strip():
                st.warning("Name the new profile first.")
            elif new_profile.strip() in profiles:
                st.warning("A profile with that name already exists.")
            else:
                updated = dict(profiles)
                updated[new_profile.strip()] = lf.settings_diff(edited)
                if _write_gsheet_config(LEASE_FORMAT_SHEET, updated):
                    st.success(f"Saved format profile: {new_profile.strip()}")
                    st.rerun()
                else:
                    st.error("Could not save the format profile to Google Sheets.")
        st.caption("Unsaved formatting changes — they apply to the preview now, but save to keep them.")

    st.divider()
    return profile_name, edited


def render_lease_builder_tab():
    st.markdown("## 🧱 Lease Builder")

    if not LEASE_BUILDER_AVAILABLE:
        st.error(f"Lease Builder could not load — {LEASE_BUILDER_ERROR}")
        # The two failure modes need opposite fixes, so tell them apart rather
        # than always suggesting a reinstall.
        if "cannot import name" in LEASE_BUILDER_ERROR:
            stale = re.search(r"from '([\w.]+)'", LEASE_BUILDER_ERROR)
            module_name = stale.group(1) if stale else "lease_builder"
            st.markdown(
                f"`app.py` is newer than the `{module_name}` module currently loaded.\n\n"
                "**On Streamlit Cloud:** the old module is still cached in memory — "
                "*Manage app → ⋮ → Reboot app* forces a fresh import. A rerun will not do it.\n\n"
                "**Running locally:** stop and restart `streamlit run app.py`; a rerun alone "
                "does not re-import a changed module.\n\n"
                f"If it persists after a reboot, `{module_name}.py` on the deployed branch is "
                "genuinely out of date — confirm the push landed."
            )
        else:
            st.markdown(
                "A required package is missing. Run this from the project folder, then restart:\n\n"
                "```\npip install -r requirements.txt\n```\n\n"
                "The package is named `python-docx` but imports as `docx`."
            )
        return

    templates = discover_templates()
    if not templates:
        st.warning("No DOCX lease templates were found in data/Lease Builder.")
        return

    saved_templates = _read_gsheet_config(LEASE_TEMPLATE_SHEET) or {}
    saved_leases = _read_gsheet_config(SAVED_LEASE_SHEET) or {}
    saved_profiles = lf.normalize_profiles(_read_gsheet_config(LEASE_FORMAT_SHEET))
    saved_clause_library = _read_gsheet_config("Lease Clause Library") or {}
    built_in_library = load_clause_library().get("sections", {})

    # One list of documents. "Triple Net Template" and "ABC Bakery Lease" are the
    # same kind of object; only what you leave checked differs. The old template
    # and lease sheets are folded in on first load and left untouched as backup.
    saved_docs, doc_source = _lb_load_documents()
    if not saved_docs and (saved_templates or saved_leases):
        saved_docs, migration_notes = ld.migrate_stores(saved_templates, saved_leases)
        if saved_docs:
            merged_ok, merge_failures = _lb_migrate_sheet_to_repo(saved_docs)
            if merged_ok:
                doc_source = "repo"
                st.success(
                    f"Merged {len(saved_templates)} template(s) and {len(saved_leases)} lease(s) "
                    f"into {len(saved_docs)} document(s). The old sheets are untouched."
                )
                for note in migration_notes:
                    st.warning(note)
            else:
                st.error(
                    "Some documents could not be written to the data repo, so nothing "
                    "was migrated. The old sheets are untouched."
                )
                for failure in merge_failures:
                    st.warning(failure)

    # Moving documents out of the Sheet is one-way, and it happens silently the
    # first time the repo is reachable. Saying so is the difference between a
    # migration you can trust and one you discover later.
    if doc_source == "sheet" and saved_docs:
        if get_lease_store() is None:
            st.warning(
                f"Reading {len(saved_docs)} document(s) from the old Google Sheet. "
                "Saving is disabled until the lease data repo is configured — "
                "add a [lease_data] section to secrets."
            )
        else:
            migrate_col, note_col = st.columns([1.4, 4])
            if migrate_col.button("📦 Move to data repo", type="primary", key="lb_migrate_to_repo"):
                moved_ok, moved_failures = _lb_migrate_sheet_to_repo(saved_docs)
                if moved_ok:
                    get_lease_store.clear()
                    st.success(
                        f"Copied {len(saved_docs)} document(s) into the data repo. "
                        "The Google Sheet is unchanged, so nothing is lost either way."
                    )
                    st.rerun()
                else:
                    st.error("Nothing was migrated — every document has to land first.")
                    for failure in moved_failures:
                        st.warning(failure)
            note_col.info(
                f"{len(saved_docs)} document(s) are still in the old Google Sheet. "
                "Move them to the data repo to get per-document files and version history."
            )

    doc_names = sorted(saved_docs)
    head2, head3 = st.columns([4, 1.4])
    # A "Save As" from the previous run parks its new name here. Applying it
    # before the picker exists is the only legal moment to set a widget key.
    pending_choice = st.session_state.pop("lb_pending_template_choice", None)
    if pending_choice in doc_names:
        st.session_state["lb_template_edit_choice"] = pending_choice
    working_template = head2.selectbox(
        "Document", doc_names + [LB_NEW_TEMPLATE], key="lb_template_edit_choice",
        help="Templates and leases are the same thing — name them however you like.",
    )
    clean_notes = head3.toggle("Clean draft", value=True, key="lb_clean_notes")
    if working_template in saved_docs:
        st.caption(ld.describe_document(saved_docs[working_template]))

    # ---- Space ------------------------------------------------------------
    # The facts a lease repeats about its space already live in the tenancy
    # workbook. Picking one here resolves every [Space:...] token, so the same
    # document produces a correct lease for any unit.
    space_records = _lb_space_records()
    active_space = None
    if space_records:
        space_labels = [record["_label"] for record in space_records]
        saved_space_key = str((saved_docs.get(working_template) or {}).get("space_key", ""))
        saved_space = lsp.find_space(space_records, saved_space_key)
        if "lb_space_choice" not in st.session_state and saved_space:
            st.session_state["lb_space_choice"] = saved_space["_label"]
        space_col, refresh_col = st.columns([4, 1.4])
        chosen_label = space_col.selectbox(
            "Space", ["— none —"] + space_labels, key="lb_space_choice",
            help="From the MSP Tenancy workbook. Resolves [Space:...] tokens in "
                 "key provisions and clause text.",
        )
        if refresh_col.button("🔄 Reload tenancy", key="lb_space_refresh", width="stretch"):
            _lb_space_records.clear()
            st.rerun()
        active_space = next(
            (record for record in space_records if record["_label"] == chosen_label), None
        )
        if active_space:
            fields = [
                f"**{lsp.field_label(name)}:** {active_space[name]}"
                for name in lsp.token_names()
                if active_space.get(name)
            ]
            st.caption(" · ".join(fields))
            blank = [
                lsp.field_label(name) for name in lsp.token_names()
                if not active_space.get(name)
            ]
            if blank:
                # A blank field leaves its token unresolved rather than printing
                # nothing, so it is better to say so before the lease is built.
                st.caption(f"⚠️ Not in the workbook for this space: {', '.join(blank)}")
        else:
            st.caption(
                "No space selected — [Space:...] tokens stay as written. "
                f"Available: {', '.join('[Space:' + name + ']' for name in lsp.token_names())}"
            )
    else:
        st.caption("MSP Tenancy workbook not found — [Space:...] tokens cannot resolve.")

    # ---- New document: a copy of an existing one --------------------------
    if working_template == LB_NEW_TEMPLATE:
        if not doc_names:
            st.info(
                "No saved documents yet. The first one is built from the extracted "
                "master content — name it with Save As once you have it how you want it."
            )
        else:
            copy_col, go_col = st.columns([4, 1.4])
            copy_source = copy_col.selectbox(
                "Start from", doc_names, key="lb_copy_source",
                help="The new document is a full, independent copy — provisions, "
                     "alternates, section choices and format profile.",
            )
            if go_col.button("📋 Copy", key="lb_copy_go", width="stretch"):
                st.session_state["lb_pending_copy"] = copy_source
                st.rerun()

    draft_name = working_template if working_template != LB_NEW_TEMPLATE else "MSP Lease"

    with st.expander("⚙️ Advanced — base Word template", expanded=False):
        st.caption(
            "Legacy path. Clause text now comes from `lease_content.json`; this "
            "file is only still used by the old Word-template generator."
        )
        selected_label = st.selectbox(
            "Base lease template (.docx)", [item["label"] for item in templates],
            key="lb_template",
        )

    template = next(item for item in templates if item["label"] == selected_label)

    try:
        template_data = inspect_template(template["path"])
    except Exception as exc:
        st.error(f"The selected Word template could not be read: {exc}")
        return

    sections = template_data["sections"]
    section_numbers = [str(section["number"]) for section in sections]
    section_labels = {
        str(section["number"]): f"Section {section['number']} — {section['title']}"
        for section in sections
    }
    section_by_number = {str(section["number"]): section for section in sections}

    # ---- Draft state, rebuilt whenever the mode or a selector changes ----
    def fresh_draft_state():
        return {
            "key_provisions": [
                {
                    "Group": "Mandatory" if item["bookmark"] in LEASE_MANDATORY_BOOKMARKS else "Optional",
                    "Include": True,
                    "Field": item["field"],
                    "Value": item["value"],
                    "Alternates": [""] * 10,
                    "Link": item["bookmark"] in LEASE_DEFAULT_LINKS,
                    "Section": section_labels.get(
                        str(LEASE_DEFAULT_LINKS.get(item["bookmark"], section_numbers[0])),
                        section_labels[section_numbers[0]],
                    ),
                    "Bookmark": item["bookmark"],
                }
                for item in template_data["bookmarks"]
            ] + [
                # Cross-referenced in the clause text but never surfaced as a key
                # provision. Without these the citation would silently vanish.
                {
                    "Group": "Optional",
                    "Include": True,
                    "Field": item["field"],
                    "Value": "",
                    "Alternates": [""] * 10,
                    "Link": False,
                    "Section": section_labels[section_numbers[0]],
                    "Bookmark": item["bookmark"],
                }
                for item in template_data.get("orphan_refs", [])
            ],
            "sections": {
                str(section["number"]): {
                    "include": True,
                    "choice": "Template language",
                    "text": section["text"],
                }
                for section in sections
            },
            "rent_schedules": _lb_default_rent_schedules(),
            "format_profile": lf.DEFAULT_PROFILE_NAME,
            "formatting": lf.default_settings(),
            "kp_version": 0,
            "section_version": 0,
        }

    draft_state_key = f"lb_draft_state_{Path(template['path']).stem}"
    pending_copy = st.session_state.get("lb_pending_copy", "")
    load_marker = "|".join([selected_label, working_template, pending_copy])
    if draft_state_key not in st.session_state or st.session_state.get("lb_load_marker") != load_marker:
        new_state = fresh_draft_state()
        # A document stands on its own: provisions, alternates, section choices
        # and format profile all come from the one payload, with no parent.
        source_doc = {}
        if pending_copy and pending_copy in saved_docs:
            source_doc = ld.copy_document(saved_docs[pending_copy], pending_copy)
        elif working_template in saved_docs:
            source_doc = saved_docs[working_template]
        if source_doc:
            _lb_apply_document(new_state, source_doc, saved_profiles)
        st.session_state[draft_state_key] = new_state
        st.session_state["lb_load_marker"] = load_marker
        _lb_reset_editor_widget_state()
        st.rerun()

    draft_state = st.session_state[draft_state_key]
    draft_state.setdefault("rent_schedules", _lb_default_rent_schedules())
    if pending_copy:
        st.info(
            f"Unsaved copy of **{pending_copy}**. Use **Save As…** to name it — "
            "until then nothing has been written."
        )
    draft_state["format_profile"] = lf.resolve_profile_name(
        saved_profiles, draft_state.get("format_profile"))
    draft_state["formatting"] = lf.normalize_settings(draft_state.get("formatting"))
    for provision in draft_state["key_provisions"]:
        bookmark = str(provision.get("Bookmark", ""))
        # Saved drafts may contain old internal field labels (e.g., Tx_BuildingAddress).
        # Always show the human-facing lease label in both drag cards and grids.
        provision["Field"] = BOOKMARK_LABELS.get(bookmark, str(provision.get("Field", "Key Provision")))
        provision.setdefault("Group", "Mandatory" if bookmark in LEASE_MANDATORY_BOOKMARKS else "Optional")
        provision.setdefault("Include", True)
        provision.setdefault("Alternates", [])
        slots = [str(value) for value in provision.get("Alternates", [])[:10]]
        provision["Alternates"] = slots + [""] * (10 - len(slots))
        provision.setdefault("Link", bookmark in LEASE_DEFAULT_LINKS)
        raw_section = provision.get("Section", LEASE_DEFAULT_LINKS.get(bookmark, section_numbers[0]))
        provision["Section"] = section_labels.get(
            str(raw_section),
            str(raw_section) if str(raw_section) in section_labels.values() else section_labels[section_numbers[0]],
        )

    # Clause text saved before tokens were bracketed is repaired in place, so an
    # older template does not keep showing bare KP:Name citations.
    provision_names = [
        str(row.get("Field", "")).strip()
        for row in draft_state["key_provisions"]
        if str(row.get("Field", "")).strip()
    ]
    for config in draft_state["sections"].values():
        current = str(config.get("text", ""))
        upgraded = normalize_kp_tokens(current, provision_names)
        if upgraded != current:
            config["text"] = upgraded

    if False:
        st.info(
            "No lease template selected — you are drafting straight from the base Word file, so no "
            "clause alternates are available. Pick a template above to draft from the approved menu."
        )

    # ---- Template header -------------------------------------------------
    # Rendered here, in document order, now that sections and draft_state exist:
    # the save buttons need both to build a payload. Everything that identifies
    # what you are editing stays above the two-column body.
    def template_payload(profile_name):
        # The whole document, self-contained: alternates included, so it
        # never depends on another document still existing.
        payload = ld.build_document(
            draft_state["key_provisions"],
            _lb_compact_sections(sections, draft_state["sections"]),
            draft_state.get("rent_schedules"),
            profile_name,
            copied_from=draft_state.get("copied_from", ""),
        )
        # Saved by key, not by resolved values: reopening the document re-reads
        # the workbook, so a corrected square footage reaches an existing lease.
        payload["space_key"] = str((active_space or {}).get("_key", ""))
        return payload

    active_profile, active_settings = _lb_render_template_header(
        working_template, saved_docs, saved_profiles,
        draft_state.get("format_profile"), template_payload, editable=True,
    )
    draft_state["format_profile"] = active_profile
    draft_state["formatting"] = active_settings

    # One full-width column: Key Provisions and the document preview span the
    # page, and only the section editor is split with its own preview.
    left = st.container()
    live_word_bytes = None
    use_label = "Use"

    with left:
        st.markdown("### Key Provisions")
        st.caption(
            "Every provision in the base template is listed. Fill in **Alt 1–Alt 10** to define the "
            "values a lease can choose from. **Default On** decides whether a new lease starts with "
            "that provision checked."
        )
        command_col, count_col = st.columns([2, 3])
        master_key = f"lb_kp_master_{Path(template['path']).stem}"
        master_applied_key = master_key + "_applied"
        st.session_state.setdefault(master_key, True)
        st.session_state.setdefault(master_applied_key, True)
        master_value = command_col.checkbox(f"{use_label}: all / none", key=master_key)
        if master_value != st.session_state[master_applied_key]:
            for row in draft_state["key_provisions"]:
                row["Include"] = master_value
            draft_state["kp_version"] += 1
            st.session_state[master_applied_key] = master_value
            st.rerun()

        # Always-visible Excel controls sit directly above the grid. The uploader
        # lives in a popover so it reads as a button rather than a drop zone.
        excel_col, upload_col = st.columns([1, 1])
        excel_col.download_button(
            "⬇️ Download Excel",
            data=_lb_export_key_provisions_xlsx(draft_state["key_provisions"]),
            file_name="MSP_Key_Provisions.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="lb_download_key_provisions_excel",
            width="stretch",
        )
        with upload_col.popover("⬆️ Upload Excel", width="stretch"):
            st.caption("Bulk-edit the Alt 1–Alt 10 choice lists in Excel, then upload the workbook back.")
            uploaded_kp = st.file_uploader(
                "Key Provisions workbook",
                type=["xlsx", "xls"],
                key="lb_upload_key_provisions_excel",
                label_visibility="collapsed",
            )
            if uploaded_kp is not None and st.button(
                "Import Uploaded Table", key="lb_import_key_provisions_excel", width="stretch"
            ):
                try:
                    imported_rows, removed_fields = _lb_import_key_provisions_xlsx(
                        uploaded_kp, draft_state["key_provisions"], section_labels
                    )
                    # A provision still cited by a [KP:] token would leave a hole
                    # in the clause, so say so before it reaches a signed lease.
                    still_cited = sorted({
                        name for name in removed_fields
                        if any(name in find_kp_references(
                                   str(config.get("text", "")), removed_fields)
                               for config in draft_state["sections"].values())
                    })
                    draft_state["key_provisions"] = imported_rows
                    draft_state["kp_version"] += 1
                    if removed_fields:
                        st.warning(
                            f"Imported. {len(removed_fields)} provision(s) were not in the "
                            f"spreadsheet and have been removed: {', '.join(removed_fields)}"
                        )
                    else:
                        st.success("Key Provisions table imported.")
                    if still_cited:
                        st.error(
                            "These removed provisions are still cross-referenced in clause "
                            f"text and will now print as literal tokens: {', '.join(still_cited)}"
                        )
                    st.rerun()
                except Exception as exc:
                    st.error(f"Key Provisions import failed: {exc}")

        # Template mode shows every section, so citations are not filtered there.
        kp_citations = _lb_token_citations(sections, draft_state, include_only=False)

        # Single editable Key Provisions grid. The drag handle is the only reorder control.
        draft_state["key_provisions"] = sorted(
            draft_state["key_provisions"],
            key=lambda row: 0 if bool(row.get("Include")) else 1,
        )
        # Choice picks which text the provision actually uses. It sits between
        # Current Value and Alt 1, offers only the alternates that have text,
        # and copies the winner into Current Value so the cell always shows
        # what will print. Choice itself is saved, so a lease records that it
        # uses "Alt 2" rather than only the words that happened to be there.
        grid_rows = []
        for source_row in draft_state["key_provisions"]:
            row = ld.apply_choice(source_row)
            slots = [str(value) for value in row.get("Alternates", [])[:10]]
            slots += [""] * (10 - len(slots))
            row["Current Value"] = row.get("Value", "")
            row["Choice"] = ld.normalize_choice(row)
            for index, value in enumerate(slots, start=1):
                row[f"Alt {index}"] = value
            cited_in = kp_citations.get(str(row.get("Field", "")), [])
            row["Linked"] = (
                "Yes — " + ", ".join(f"§{number}" for number in cited_in) if cited_in else ""
            )
            row.pop("Value", None)
            row.pop("Alternates", None)
            row.pop("Link", None)
            row.pop("Section", None)
            grid_rows.append(row)
        kp_df = pd.DataFrame(grid_rows)
        # Drag must stay the first column (it is the reorder handle) and Bookmark
        # is an internal id, so neither is offered to the column config editor.
        KP_TAB_KEY = "key provisions"
        configurable_columns = [
            column for column in kp_df.columns if column not in ("Drag", "Bookmark")
        ]
        ordered_columns = get_column_order(KP_TAB_KEY, configurable_columns)
        # A saved column order predates Choice, and get_column_order appends
        # anything it has never seen — which would strand the chooser out past
        # Alt 10. Put it back beside the value it controls.
        if "Choice" in ordered_columns and "Current Value" in ordered_columns:
            ordered_columns = [column for column in ordered_columns if column != "Choice"]
            ordered_columns.insert(ordered_columns.index("Current Value") + 1, "Choice")
        tail = ["Bookmark"] if "Bookmark" in kp_df.columns else []
        kp_df = kp_df[[column for column in ordered_columns if column in kp_df.columns] + tail]
        kp_df.insert(0, "Drag", "")
        grid_builder = GridOptionsBuilder.from_dataframe(kp_df)
        grid_builder.configure_default_column(resizable=True, sortable=False, filter=False, editable=False)
        grid_builder.configure_column("Drag", header_name="", rowDrag=True, editable=False, width=42, suppressMenu=True)
        # Group, Link and Target Section describe the template's structure, so a
        # lease can read them but only template mode can change them.
        grid_builder.configure_column("Group", header_name="Group", editable=True, width=105,
                                      cellEditor="agSelectCellEditor",
                                      cellEditorParams={"values": ["Mandatory", "Optional"]})
        grid_builder.configure_column("Include", header_name=use_label, editable=True, width=88,
                                      cellRenderer="agCheckboxCellRenderer", cellEditor="agCheckboxCellEditor")
        grid_builder.configure_column("Field", header_name="Key Provision", editable=True, width=165)
        # Only slots that actually hold text are offered. Choosing an empty
        # "Alt 7" would blank the provision in the generated lease.
        choice_options_js = JsCode("""
            function(params) {
                var values = [''];
                for (var i = 1; i <= 10; i++) {
                    var slot = params.data['Alt ' + i];
                    if (slot !== null && slot !== undefined && String(slot).trim() !== '') {
                        values.push('Alt ' + i);
                    }
                }
                return { values: values };
            }
        """)
        # Free text: template mode is where the default value itself is authored.
        grid_builder.configure_column("Current Value", header_name="Default Value", editable=True, width=330)
        grid_builder.configure_column(
            "Choice", header_name="Choice", editable=True, width=120,
            cellEditor="agSelectCellEditor", cellEditorParams=choice_options_js,
        )
        # Linking is read-only: it reflects the KP: tokens found in clause text.
        grid_builder.configure_column("Linked", header_name="Linked", editable=False, width=170)
        for index in range(1, 11):
            grid_builder.configure_column(f"Alt {index}", header_name=f"Alt {index}",
                                          editable=True, width=185)
        grid_builder.configure_column("Bookmark", hide=True)

        # Saved widths win over the defaults above, matching the Tenancy tab.
        kp_width_overrides = get_column_width_overrides(KP_TAB_KEY)
        for column, width in kp_width_overrides.items():
            if column in kp_df.columns:
                grid_builder.configure_column(
                    column, width=width, initialWidth=width, minWidth=width,
                    suppressSizeToFit=True, suppressAutoSize=True,
                )

        grid_builder.configure_grid_options(
            rowDragManaged=True,
            animateRows=True,
            suppressMoveWhenRowDragging=False,
            rowDragEntireRow=False,
        )
        grid_result = AgGrid(
            kp_df,
            gridOptions=grid_builder.build(),
            height=min(760, 96 + len(kp_df) * 42),
            theme="streamlit",
            update_mode=GridUpdateMode.MODEL_CHANGED,
            data_return_mode=DataReturnMode.AS_INPUT,
            fit_columns_on_grid_load=False,
            allow_unsafe_jscode=True,
            key=f"lb_kp_grid_v8_{Path(template['path']).stem}",
        )
        edited_kp = grid_result.data if hasattr(grid_result, "data") else grid_result["data"]
        if edited_kp is None:
            edited_kp = kp_df
        edited_rows = []
        previous_by_bookmark = {
            str(row.get("Bookmark", "")): row for row in draft_state["key_provisions"]
        }
        for row in edited_kp.drop(columns=["Drag"], errors="ignore").to_dict("records"):
            slots = [str(row.pop(f"Alt {index}", "")).strip() for index in range(1, 11)]
            current_value = str(row.pop("Current Value", ""))
            choice = str(row.pop("Choice", "") or ld.NO_CHOICE)
            row.pop("Linked", None)  # Derived for display only.
            previous = previous_by_bookmark.get(str(row.get("Bookmark", "")), {})
            row["Link"] = bool(previous.get("Link", False))
            row["Section"] = previous.get("Section", section_labels[section_numbers[0]])
            row["Value"] = current_value
            row["Alternates"] = slots
            row["Choice"] = choice
            # Re-applied every edit, so retyping a chosen alternate updates the
            # value too rather than leaving the lease on a stale copy.
            edited_rows.append(ld.apply_choice(row))
        draft_state["key_provisions"] = edited_rows
        selected_count = sum(1 for row in edited_rows if bool(row.get("Include")))
        count_col.caption(f"{len(edited_rows)} provisions · {selected_count} on by default")
        st.caption(
            f"Use **{KP_LINE_BREAK}** inside a value to start a new line — "
            f"`123 Main St{KP_LINE_BREAK}Westfield, NJ{KP_LINE_BREAK}07090`. "
            "The grid commits on Enter so it cannot hold a real line break; in Excel you can use "
            "Alt+Enter instead and it converts on import."
        )
        render_column_config_editor(KP_TAB_KEY, configurable_columns)

        # Adding and removing provisions is done through the Excel round trip
        # above, so the whole list can be edited at once and reviewed offline.
        # The importer still creates and drops rows; only the in-app buttons are
        # gone.

        # ---- Off-menu value check ----------------------------------------
        if False:  # provision-level off-menu check retired with the two-mode split
            off_menu_rows = [
                row for row in edited_rows
                if bool(row.get("Include"))
                and str(row.get("Value", "")).strip()
                and _lb_is_off_menu(row.get("Value", ""), [
                    value for value in row.get("Alternates", []) if str(value).strip()
                ])
            ]
            if off_menu_rows:
                with st.expander(f"⚠️ {len(off_menu_rows)} value(s) not in the template", expanded=False):
                    st.caption(
                        "These values were typed for this deal and are not among the template's "
                        "alternates. Add any that should become a standing choice."
                    )
                    add_target = st.selectbox(
                        "Add to which template?", template_names, key="lb_offmenu_value_target"
                    )
                    for row in off_menu_rows:
                        bookmark = str(row.get("Bookmark", ""))
                        value_col, button_col = st.columns([4, 1.3])
                        value_col.markdown(f"**{row.get('Field', '')}** — {row.get('Value', '')}")
                        if button_col.button("Add as Alt", key=f"lb_offmenu_add_{bookmark}"):
                            target = dict(saved_templates.get(add_target, {}))
                            target_rows = [dict(item) for item in target.get("key_provisions", [])]
                            match = next(
                                (item for item in target_rows if str(item.get("Bookmark", "")) == bookmark), None
                            )
                            if match is None:
                                st.error("That provision does not exist in the selected template.")
                            else:
                                slots = [str(value) for value in match.get("Alternates", [])[:10]]
                                slots += [""] * (10 - len(slots))
                                if str(row.get("Value", "")) in slots:
                                    st.info("That value is already a choice in the template.")
                                elif "" not in slots:
                                    st.error("All ten Alt slots are full in that template.")
                                else:
                                    slots[slots.index("")] = str(row.get("Value", ""))
                                    match["Alternates"] = slots
                                    target["key_provisions"] = target_rows
                                    target["saved_at"] = datetime.now().isoformat(timespec="seconds")
                                    updated = dict(saved_templates)
                                    updated[add_target] = target
                                    if _write_gsheet_config(LEASE_TEMPLATE_SHEET, updated):
                                        st.success(f"Added to {add_target}.")
                                        st.rerun()
                                    else:
                                        st.error("Could not write to Google Sheets.")

        # ---- Rent schedule -------------------------------------------------
        if True:
            with st.expander("💵 Rent Schedule Excel Import / Export", expanded=False):
                rent_state = draft_state.setdefault("rent_schedules", _lb_default_rent_schedules())
                rent_state.pop("settings", None)  # Retired generator settings from older drafts.
                st.caption(
                    "Download the workbook, type the monthly and annual rent for each period in Excel, "
                    "then upload it back. Whatever you enter is what prints in the lease — nothing is recalculated."
                )
                download_col, upload_col = st.columns([1, 1])
                download_col.download_button(
                    "⬇️ Download Rent Schedule Excel",
                    data=_lb_export_rent_schedule_xlsx(rent_state),
                    file_name="MSP_Rent_Schedule.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="lb_download_rent_schedule_excel",
                    width="stretch",
                )
                uploaded_rent = upload_col.file_uploader(
                    "Upload Rent Schedule Excel",
                    type=["xlsx", "xls"],
                    key="lb_upload_rent_schedule_excel",
                )
                if uploaded_rent is not None and st.button("Import Uploaded Rent Schedule", key="lb_import_rent_schedule_excel"):
                    try:
                        draft_state["rent_schedules"] = _lb_import_rent_schedule_xlsx(uploaded_rent)
                        imported = draft_state["rent_schedules"]
                        st.success(
                            f"Rent schedule imported — {len(imported['base'])} base rows, "
                            f"{len(imported['options'])} option rows."
                        )
                        st.rerun()
                    except Exception as exc:
                        st.error(f"Rent schedule import failed: {exc}")

                st.caption(
                    "Workbook layout: the base term under a **Term** header in columns A–C, and each "
                    "option under an **Option 1 / Option 2 …** header in columns F–H, separated by a "
                    "blank row. Add or delete rows and option blocks freely. **A year left blank or "
                    "set to 0 does not exist**, and an option whose every year is blank or 0 is "
                    "dropped entirely."
                )

                st.markdown("#### Base Rent Table")
                base_df = st.data_editor(
                    pd.DataFrame(rent_state.get("base", []), columns=RENT_COLUMNS),
                    hide_index=True,
                    num_rows="dynamic",
                    width="stretch",
                    key="rent_base_grid",
                )
                rent_state["base"] = base_df.to_dict("records")

                st.markdown("#### Option Rent Table")
                option_df = st.data_editor(
                    pd.DataFrame(rent_state.get("options", []), columns=RENT_COLUMNS),
                    hide_index=True,
                    num_rows="dynamic",
                    width="stretch",
                    key="rent_option_grid",
                )
                rent_state["options"] = option_df.to_dict("records")
                st.caption("These two tables replace the template rent grids in the generated lease.")

        # The preview belongs here visually, but the Word bytes are not built
        # until the editors below have run — reserve the space, fill it last.
        template_preview_slot = st.container()

        sec_col, prev_col = st.columns([1.05, 0.95], gap="large")
        with sec_col:
            # ---- Lease sections ------------------------------------------------
            st.markdown("### Lease Sections")
            st.caption(
                "Every numbered section is listed. **Default On** decides whether a new lease starts "
                "with the section included. Select a section below to author its language and save "
                "additional clause choices."
            )
            section_rows = [
                {
                    "Include": bool(draft_state["sections"][str(section["number"])].get("include", True)),
                    "Number": str(section["number"]),
                    "Section": section["title"],
                }
                for section in sections
            ]
            section_df = pd.DataFrame(section_rows)
            edited_sections = st.data_editor(
                section_df,
                hide_index=True,
                width="stretch",
                height=520,
                disabled=["Number", "Section"],
                column_config={
                    "Include": st.column_config.CheckboxColumn(use_label, width="small"),
                    "Number": st.column_config.TextColumn("#", width="small"),
                    "Section": st.column_config.TextColumn("Section", width="large"),
                },
                key=f"lb_section_editor_{Path(template['path']).stem}_{draft_state['section_version']}",
            )
            for row in edited_sections.to_dict("records"):
                draft_state["sections"][str(row["Number"])]["include"] = bool(row["Include"])

            whole_numbers = sorted(int(number) for number in section_numbers if number.isdigit())
            missing_numbers = [number for number in range(whole_numbers[0], whole_numbers[-1] + 1) if number not in whole_numbers]
            if missing_numbers:
                st.warning("Source-template numbering gap: " + ", ".join(f"Section {number}" for number in missing_numbers))

            selected_section = st.selectbox(
                "Edit section",
                section_numbers,
                format_func=lambda number: section_labels[number],
                key="lb_selected_section",
            )
            st.session_state["lb_preview_focus_section"] = selected_section
            section_config = draft_state["sections"][selected_section]

            built_variants = built_in_library.get(selected_section, {}).get("variants", [])
            cloud_variants = (
                saved_clause_library.get(selected_label, {}).get(selected_section, [])
                if isinstance(saved_clause_library, dict) else []
            )
            variants = []
            seen_names = set()
            for variant in built_variants + cloud_variants:
                if variant.get("name") and variant["name"] not in seen_names:
                    variants.append(variant)
                    seen_names.add(variant["name"])
            variant_map = {variant["name"]: variant["text"] for variant in variants}
            choice_options = ["Template language"] + list(variant_map.keys()) + ["Custom language"]
            current_choice = section_config.get("choice", "Template language")
            if current_choice not in choice_options:
                current_choice = "Custom language"
            st.caption(f"{len(variant_map)} saved clause choice(s) for this section.")
            choice = st.selectbox(
                "Clause choice",
                choice_options,
                index=choice_options.index(current_choice),
                key=f"lb_section_choice_{selected_section}",
            )
            if choice == "Template language":
                default_text = section_by_number[selected_section]["text"]
            elif choice == "Custom language":
                default_text = section_config.get("text", section_by_number[selected_section]["text"])
            else:
                default_text = variant_map[choice]
            text_key = f"lb_section_text_{selected_section}_{re.sub(r'[^A-Za-z0-9]+', '_', choice)}"
            section_text = st.text_area("Section language", value=default_text, height=260, key=text_key)
            section_config.update({"choice": choice, "text": section_text})

            all_provision_names = [str(r.get("Field", "")) for r in draft_state["key_provisions"]]
            if KP_TOKEN_RE.search(section_text or ""):
                st.caption("Cross-references — green resolves, red matches no provision:")
                st.markdown(
                    _lb_token_preview_html(section_text, all_provision_names),
                    unsafe_allow_html=True,
                )
            st.caption(
                "Cite a key provision with **[KP:Name]** — for example `[KP:Lease Execution Date]`. "
                "The value from the Key Provisions summary is substituted when the document is built, "
                "printed in green as a link back up to that row, and the Linked column updates itself."
            )
            cited_here = sorted(find_kp_references(section_text, all_provision_names))
            if cited_here:
                st.caption("Cited in this section: " + ", ".join(f"`[KP:{name}]`" for name in cited_here))
            with st.popover("📋 Key provision tokens", width="stretch"):
                st.caption("Copy a token into the clause text.")
                for row in draft_state["key_provisions"]:
                    st.code(f"[KP:{row.get('Field', '')}]", language=None)

            approved_texts = [section_by_number[selected_section]["text"]] + _lb_section_variant_texts(
                selected_label, selected_section, built_in_library, saved_clause_library
            )
            off_menu_text = _lb_is_off_menu(section_text, approved_texts)

            save_choice_col, choice_name_col = st.columns([2, 3])
            new_choice_name = choice_name_col.text_input(
                "New clause-choice name",
                key=f"lb_new_choice_name_{selected_section}",
                placeholder="e.g., Customer parking only",
            )
            if save_choice_col.button("Save as clause choice", key=f"lb_save_choice_{selected_section}"):
                if not new_choice_name.strip():
                    st.warning("Enter a name for the clause choice.")
                else:
                    updated_library = dict(saved_clause_library) if isinstance(saved_clause_library, dict) else {}
                    template_library = updated_library.setdefault(selected_label, {})
                    section_library = template_library.setdefault(selected_section, [])
                    section_library = [item for item in section_library if item.get("name") != new_choice_name.strip()]
                    section_library.append({"name": new_choice_name.strip(), "text": section_text})
                    template_library[selected_section] = section_library
                    if _write_gsheet_config("Lease Clause Library", updated_library):
                        st.success(f"Saved clause choice: {new_choice_name.strip()}")
                        st.rerun()
                    else:
                        st.error("Could not save the clause choice to Google Sheets.")
            elif off_menu_text:
                st.warning(
                    "This language is not in the template. It will still print in this lease — "
                    "should it also become a standing clause choice?"
                )
                off1, off2, off3 = st.columns([2, 2, 1.4])
                off_target = working_template
                off1.caption(f"Will be added to **{working_template}**.")
                off_name = off2.text_input(
                    "Clause-choice name",
                    key=f"lb_offmenu_section_name_{selected_section}",
                    placeholder="e.g., Restaurant venting carve-out",
                )
                if off3.button("Add choice", key=f"lb_offmenu_section_add_{selected_section}",
                               disabled=working_template == LB_NEW_TEMPLATE):
                    if not off_name.strip():
                        st.warning("Name the clause choice first.")
                    else:
                        # Clause choices are keyed by base .docx label, so the whole
                        # library for this base template picks the new choice up.
                        updated_library = dict(saved_clause_library) if isinstance(saved_clause_library, dict) else {}
                        template_library = updated_library.setdefault(selected_label, {})
                        section_library = template_library.setdefault(selected_section, [])
                        section_library = [item for item in section_library if item.get("name") != off_name.strip()]
                        section_library.append({"name": off_name.strip(), "text": section_text})
                        template_library[selected_section] = section_library
                        if _write_gsheet_config("Lease Clause Library", updated_library):
                            st.success(f"Added '{off_name.strip()}' as a clause choice for {off_target}.")
                            st.rerun()
                        else:
                            st.error("Could not save the clause choice to Google Sheets.")

            # ---- Publish --------------------------------------------------------
            with st.expander("📦 Publish as a Word Template", expanded=False):
                st.caption(
                    "Writes a real .docx into data/Lease Builder/Published with every choice applied. "
                    "That file then appears in the base-template picker, so it becomes the starting "
                    "point going forward — and you can open it in Word to adjust formatting by hand. "
                    "Publishing keeps all sections and every key provision."
                )
                publish_name_col, publish_button_col = st.columns([3, 1.6])
                publish_name = publish_name_col.text_input(
                    "Published file name",
                    value=working_template if working_template != LB_NEW_TEMPLATE else "",
                    key="lb_publish_name",
                    placeholder="e.g., MSP NNN Retail 2026",
                )
                if publish_button_col.button("📦 Publish", key="lb_publish_docx", width="stretch"):
                    if not publish_name.strip():
                        st.warning("Name the published file first.")
                    else:
                        try:
                            publish_choices = {}
                            for section in sections:
                                number = str(section["number"])
                                config = draft_state["sections"][number]
                                configured = str(config.get("text", section["text"]))
                                base_text = str(section["text"])
                                publish_choices[number] = {
                                    "include": True,
                                    "title": section["title"],
                                    "replacement_text": (
                                        "" if configured.strip() == base_text.strip()
                                        else _lb_section_body(number, configured)
                                    ),
                                }
                            published_path = publish_template_docx(
                                template["path"],
                                publish_name.strip(),
                                publish_choices,
                                None,
                                None,
                                clean_notes,
                                key_provision_rows=[
                                    {
                                        "field": row.get("Field", "Key Provision"),
                                        "value": row.get("Value", ""),
                                        "include": True,
                                        "link": bool(row.get("Link")),
                                        "section": section_labels.get(
                                            str(row.get("Section", "")), str(row.get("Section", ""))
                                        ),
                                    }
                                    for row in draft_state["key_provisions"]
                                ],
                            )
                            # publish_template_docx writes into data/Lease Builder/
                            # Published, which on Streamlit Cloud is wiped by the
                            # next redeploy. Copying the bytes into the data repo
                            # is what actually makes a published file durable.
                            published_bytes = published_path.read_bytes()
                            store = get_lease_store()
                            if store is not None:
                                try:
                                    store.publish(published_path.name, published_bytes)
                                    st.success(
                                        f"Published {published_path.name} to the data repo. "
                                        "It survives redeploys and is downloadable below."
                                    )
                                except Exception as exc:
                                    st.warning(
                                        f"Built {published_path.name}, but it could not be saved "
                                        f"to the data repo — {exc}. Download it now; the local "
                                        "copy is lost on the next redeploy."
                                    )
                            else:
                                st.warning(
                                    f"Built {published_path.name}, but the data repo is not "
                                    "configured. Download it now — the local copy is lost on "
                                    "the next redeploy."
                                )
                            st.download_button(
                                f"⬇️ Download {published_path.name}",
                                data=published_bytes,
                                file_name=published_path.name,
                                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                key="lb_download_published",
                                width="stretch",
                            )
                        except Exception as exc:
                            st.error(f"Publish failed: {exc}")
            st.caption("Drafting tool only. Final lease language should be reviewed by New Jersey counsel.")

            # Two generators run side by side while the rules-based one is finished:
            # the legacy path edits a copy of the base .docx, the new one builds from
            # lease_format + lease_markup and actually obeys the format profile.
            engine = st.radio(
                "Generator",
                ["Word template (current)", "Rules-based (new)"],
                horizontal=True,
                key="lb_engine",
                help=(
                    "The rules-based renderer needs no base .docx and applies the format "
                    "profile above. Key Provisions table, rent tables, signatures and "
                    "exhibits are still being built, so they appear as markers for now."
                ),
            )
            use_rules_engine = engine.startswith("Rules")

            token_report = {}
            rules_renderer = None
            try:
                if use_rules_engine:
                    live_word_bytes, rules_renderer = _lb_build_rules_word(
                        draft_state, sections, draft_state.get("formatting"),
                        space=active_space,
                    )
                    token_report = {"unresolved": sorted(rules_renderer.unresolved), "blank": []}
                else:
                    live_word_bytes = _lb_build_current_word(
                        template["path"], draft_name, clean_notes, sections, draft_state,
                        force_include_all=False, token_report=token_report,
                        space=active_space,
                    )
            except Exception as exc:
                st.error(f"Current draft could not be generated: {type(exc).__name__}: {exc}")

            if use_rules_engine:
                st.caption(
                    f"Rules-based · {lf.describe_settings(draft_state.get('formatting'))}"
                )

            unresolved = token_report.get("unresolved") or []
            blank_refs = token_report.get("blank") or []
            if unresolved or blank_refs:
                with st.expander(
                    f"⚠️ {len(unresolved) + len(blank_refs)} cross-reference issue(s)", expanded=bool(unresolved)
                ):
                    if unresolved:
                        st.markdown("**Unrecognised tokens** — these print literally, so fix the spelling "
                                    "or add the provision:")
                        for name in unresolved:
                            st.markdown(f"- `[KP:{name}]`")
                    if blank_refs:
                        st.markdown("**Cited but not used in this lease** — these resolve to nothing, "
                                    "leaving a gap in the clause:")
                        for name in blank_refs:
                            st.markdown(f"- {name}")

            build_label = "📝 Build Word Document"
            if st.button(build_label, type="primary", key="lb_build_word", width="stretch"):
                if live_word_bytes:
                    st.session_state["lb_word_bytes"] = live_word_bytes
                    st.session_state["lb_word_filename"] = (
                        re.sub(r"[^A-Za-z0-9._-]+", "_", draft_name).strip("_") or "MSP_Lease_Draft"
                    ) + ".docx"
                    st.success("Word draft built from the current selections.")
                else:
                    st.error("The Word draft is not available because generation failed.")

            if st.session_state.get("lb_word_bytes"):
                st.download_button(
                    "⬇️ Download Word Draft",
                    data=st.session_state["lb_word_bytes"],
                    file_name=st.session_state.get("lb_word_filename", "MSP_Lease_Draft.docx"),
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    type="primary",
                    key="lb_download_word",
                    width="stretch",
                )

        preview_sections = []
        for section in sections:
            number = str(section["number"])
            config = draft_state["sections"][number]
            preview_sections.append({
                "number": number,
                "title": section["title"],
                "include": bool(config.get("include", True)),
                "text": config.get("text", section["text"]),
            })


        with prev_col:
            st.markdown("### Section Preview")
            _lb_render_section_preview(
                draft_state, preview_sections,
                st.session_state.get("lb_preview_focus_section", ""),
                word_bytes=live_word_bytes,
            )

        # Filled last because the Word bytes only exist once the editors above
        # have run, but rendered into the slot reserved under Key Provisions.
        with template_preview_slot:
            _lb_render_document_preview(
                live_word_bytes,
                draft_state=draft_state, preview_sections=preview_sections,
            )


# =====================
# MAIN APP
# =====================

col_title, col_toggle = st.columns([8, 2])
col_title.markdown("## 🏢 MSP Property Dashboard")
st.session_state.mobile_view = col_toggle.toggle("📱 Mobile", value=st.session_state.mobile_view)
st.caption(f"Marion Street Properties · {TODAY.strftime('%B %d, %Y')}")
if get_gsheet() is None:
    st.warning("Google Sheets is temporarily unavailable. Local portfolio data remains visible, but Sheets-backed edits and activity may be unavailable until the connection recovers.")

tab_tenancy, tab_vacancy, tab_leads, tab_covenants, tab_lease_builder, tab_insurance, tab_deposits, tab_reconcile, tab_yardi, tab_sop = st.tabs([
    "🏠 Current Tenancy", "🏚️ Vacancy", "📋 Lead Sheet", "📜 Lease Covenants", "🧱 Lease Builder", "🛡️ Insurance", "💰 Security Deposits", "🔄 Yardi Reconcile", "📊 Yardi Reports", "📋 SOPs"
])

with tab_sop:
    render_sop_tab()

with tab_lease_builder:
    render_lease_builder_tab()

with tab_tenancy:
    render_tenancy_tab()

with tab_vacancy:
    render_vacancy_tab()

with tab_leads:
    render_lead_sheet_tab()

with tab_covenants:
    render_covenants_tab()

with tab_insurance:
    render_insurance_tab()

with tab_deposits:
    render_deposits_tab()

with tab_reconcile:
    render_reconcile_tab()

with tab_yardi:
    render_yardi_tab()
